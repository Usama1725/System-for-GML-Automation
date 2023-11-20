# flask imports
import io
import json
import os
import re
from flask import Flask, render_template, request, jsonify, make_response,send_file, send_from_directory

# imports for PyJWT authentication
import jwt
import numpy as np
from datetime import datetime, timedelta
from functools import wraps
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill,Font,colors
from openpyxl.styles import Alignment
from flask_cors import CORS, cross_origin
import pickle

from sqlalchemy import create_engine
# creates Flask object
app = Flask(__name__)
cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'
app.config['CORS_HEADERS'] = 'x-access-token'
app.config['JSON_SORT_KEYS'] = False
# configuration
# NEVER HARDCODE YOUR CONFIGURATION IN YOUR CODE
# INSTEAD CREATE A .env FILE AND STORE IN IT
app.config['SECRET_KEY'] = 'your secret key'
# database name
# database_ip = 'localhost'
# database_username = 'root'
# database_password = ''
# database_name = 'media_campaign_dashboard'
# engine = create_engine(f'mysql+pymysql://{database_username}:{database_password}@{database_ip}/{database_name}')
my_dir = os.path.dirname(__file__)
pickle_file_path = os.path.join(my_dir, 'summary_testing.xlsx')

def get_column_letter(col_num):
    dividend = col_num
    column_letter = ''
    while dividend > 0:
        modulo = (dividend - 1) % 26
        column_letter = chr(65 + modulo) + column_letter
        dividend = (dividend - modulo) // 26
    return column_letter
def set_border_and_align(cell_info):
        thin_border = Border(left=Side(style='thin'), 
                          right=Side(style='thin'), 
                          top=Side(style='thin'), 
                          bottom=Side(style='thin'))

        TEXT_ALIGNMENT = Alignment(horizontal='center',wrap_text=True)
        cell_info.border = thin_border
        cell_info.alignment = TEXT_ALIGNMENT
    
def set_border_and_align_weekly(cell_info):
        thin_border = Border(left=Side(style='thin'), 
                          right=Side(style='thin'), 
                          top=Side(style='thin'), 
                          bottom=Side(style='thin'))

        weekly_total_bg = "EBF1DE"

        BG_PATTERN = PatternFill(start_color=weekly_total_bg,end_color=weekly_total_bg, fill_type = "solid")
            
        TEXT_ALIGNMENT = Alignment(horizontal='center',wrap_text=True)
        cell_info.border = thin_border
        cell_info.alignment = TEXT_ALIGNMENT
        cell_info.fill = BG_PATTERN

# def set_border_and_align_phase(cell_info):
#         thin_border = Border(left=Side(style='thin'), 
#                           right=Side(style='thin'), 
#                           top=Side(style='thin'), 
#                           bottom=Side(style='thin'))

#         phase_color = "FDE9D9"

#         PHASE_BG_PATTERN =  PatternFill(start_color=phase_color,end_color=phase_color, fill_type = "solid")
            
#         TEXT_ALIGNMENT = Alignment(horizontal='center',wrap_text=True)
#         cell_info.border = thin_border
#         cell_info.alignment = TEXT_ALIGNMENT
#         cell_info.fill = PHASE_BG_PATTERN

def set_border_and_red_bg(cell_info):
        thin_border = Border(left=Side(style='thin'), 
                          right=Side(style='thin'), 
                          top=Side(style='thin'), 
                          bottom=Side(style='thin'))

        phase_color = "8E1600"

        PHASE_BG_PATTERN =  PatternFill(start_color=phase_color,end_color=phase_color, fill_type = "solid")
        
        TEXT_ALIGNMENT = Alignment(horizontal='center',wrap_text=True)
        cell_info.border = thin_border
        cell_info.font = Font(color = colors.WHITE)
        cell_info.alignment = TEXT_ALIGNMENT
        cell_info.fill = PHASE_BG_PATTERN

# def set_border_and_black_bg(cell_info):
#         thin_border = Border(left=Side(style='thin'), 
#                           right=Side(style='thin'), 
#                           top=Side(style='thin'), 
#                           bottom=Side(style='thin'))

#         phase_color = "000000"

#         PHASE_BG_PATTERN =  PatternFill(start_color=phase_color,end_color=phase_color, fill_type = "solid")
        
#         TEXT_ALIGNMENT = Alignment(horizontal='center',wrap_text=True)
#         cell_info.border = thin_border
#         cell_info.font = Font(color = colors.WHITE)
#         cell_info.alignment = TEXT_ALIGNMENT
#         cell_info.fill = PHASE_BG_PATTERN

# def set_number_format(cell_info,type_of_format):

#     Formats = {
#         "percentage": '0%',
#         "comma_format": '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-',
#         "two_decimal_percentage": '0.00%',
#         "two_decimal_dollar":'_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
#         "five_digit":'#,##0',
#         "four_decimal_dollar":'_("$"* #,##0.0000_);_("$"* (#,##0.0000);_("$"* "-"????_);_(@_)'
        
#     }
#     cell_info.number_format = Formats[type_of_format]

# def check_filename_already_exist(filename,token_user):



#     res = db.session.query(media_files).filter_by(filename = filename, username = token_user.username).first()
#     db.session.close()
#     if res is None:
#         return False
#     return True




# @app.route("/latest",methods=["POST"])
# @cross_origin()
# def latest():

    


#     data_files = request.files
#     # Fetching files from the request 
#     data = request.form
#     # Fetching data from request 


#     total_raw_files = data["total_raw"]
#     # Storing total raw data 

#     raw = []
#     # storing the data in the array for different media 
    
#     # Total Raw Files for the loop 
#     for i in range(1,int(total_raw_files)+1):
#     # Looping through each raw file 

#         label = "raw-"+str(i)+"-total_media"
#         total_media = data[label]
#         # Then Loop Total Media for each raw file


#         # Data arr for each raw file 
#         data_raw = []

#         media = []

#         for j in range(1,int(total_media)+1):
#             # Storing each data in the dictionary and appending to array 

#             label_media = "raw-"+str(i)+"-media-"+str(j)
#             media_name = label_media + "-name"
#             media_budget = label_media + "-total-budget"
#             media_monthly_view_kpi = label_media + "-monthly-view-kpi"

#             total_identifiers = data[label_media+"-total-identifiers"]
            
#             identifiers = []

#             for k in range(1,int(total_identifiers)+1):
#                 label_identifier = "raw-"+str(i)+"-media-"+str(j)+"-identifier-"+str(k)

#                 identifier_name = label_identifier + "-name"
#                 identifier_type = label_identifier+"-type"

#                 identifier_columns = []
#                 if data[identifier_type] == "multiple":
#                     # If its multiple identifier then use this

#                     for z in range(1,3):
#                         column_name = data[label_identifier+"-column-"+str(z)]
#                         column_identifier = data[label_identifier+"-column-"+str(z)+"-value"]
#                         column_info = {
#                              column_name: column_identifier
#                         }
#                         identifier_columns.append(column_info)
                        
                
#                 else:
#                     # Else if its only single identifier then use this 

#                     column_name = data[label_identifier+"-column-"+str(1)]
#                     column_identifier = data[label_identifier+"-column-"+str(1)+"-value"]
#                     column_info = {
#                              column_name: column_identifier
#                     }
#                     identifier_columns.append(column_info)
                
#                 # Your code here 
                
#                 rawimage_validity = data[label_identifier+"-rawimage"]
#                 if rawimage_validity == "1":
#                     raw_image_file = data_files[label_identifier+"-rawimage-file"]
#                     identifiers_info = {
#                         "name": data[identifier_name],
#                         "type": data[identifier_type],
#                         "columns": identifier_columns,
#                         "rawimage": 1,
#                         "rawimage-file": raw_image_file
#                     }
#                 else:
#                     identifiers_info = {
#                         "name": data[identifier_name],
#                         "type": data[identifier_type],
#                         "columns": identifier_columns,
#                         "rawimage": 0
#                     }

#                 lg_validity = data[label_identifier+"-lg"]
#                 if lg_validity == "1":
#                     lg_file = data_files[label_identifier+"-lg-file"]
#                     identifiers_info['lg'] = 1
#                     identifiers_info["lg-file"] = lg_file
#                 else:
#                     identifiers_info['lg'] = 0
#                 # --------------------------
#                 identifiers.append(identifiers_info)


                
#             media_info = {
#                 "name":data[media_name],
#                 "budget": data[media_budget],
#                 "monthly_view_kpi":data[media_monthly_view_kpi],
#                 "total_identifiers": total_identifiers,
#                 "identifiers" : identifiers
#             }
#             media.append(media_info)



        

#         file = data_files["raw-"+str(i)] 
        
       
        
#         raw_info = {
#             "raw": file,
#             "total_media": total_media,
#             "media": media
#         }  

#         raw.append(raw_info)  
#         # This is the raw data with all the fields and information 
        

#     # print(raw)
#     TotalBudgetForDaily = 0 
#     TotalMonthlyViewKPI = 0

#     list_of_dataframes = []
#     # List to store the dataframes 
#     deleting_table = False
    
#     my_dir = os.path.dirname(__file__)
#     pickle_file_path = os.path.join(my_dir, 'summary_testing.xlsx')
#     with pd.ExcelWriter(pickle_file_path) as writer:
#         list_of_summary_medias_with_data = [] 
#         for dt in raw: 
#         # Loop for going through each object in the dictionary to fetch the files and then all other information     
#             file_data = pd.read_excel(dt["raw"])
#             # reading excel sheet and storing the dataframe in the file_data variable 
        
#             # Calculate the output according to the dataframe as before we were doing it 
            
            
            
#             for x in range(0,int(dt["total_media"])):
#             # Loop for going through each of the media and taking the data 

#                 TotalBudgetForDaily += int(dt["media"][x]["budget"])
#                 # Adding all the budgets for daily 
#                 TotalMonthlyViewKPI += int(dt["media"][x]["monthly_view_kpi"])
#                 # Adding all the monthly view kpi for daily 
                
#                 for y in range(0,int(dt["media"][x]["total_identifiers"])):
#                 # Loop for going through each of the identifiers as for now it is 2 

#                     label_media_identifier = dt["media"][x]["name"]+"_"+dt["media"][x]["identifiers"][y]["name"]
#                     # Getting the name of the identifier 

#                     TotalBudget = dt["media"][x]["budget"]
#                     # Getting the Budget of the identifier 
                    
#                     MonthlyViewKPI = dt["media"][x]["monthly_view_kpi"]
#                     # Getting the monthly view kpi for each identifier 

#                     key_1 = list(dt["media"][x]["identifiers"][y]["columns"][0].keys())[0]
#                     # Getting the first key as it is necessary in both cases either it is single or multiple 

#                     value_1 = dt["media"][x]["identifiers"][y]["columns"][0][key_1]
#                     # Getting the first key value and storing in the variable as it is necessary in both cases either it is single or multiple 
                    
#                     key_2 = None
#                     value_2 = None
#                     # Setting key 2 and value 2  to none for now so if there is no key 2 then we can pass none 


#                     if dt["media"][x]["identifiers"][y]["type"] == "multiple":
#                     # If it is multiple then we can save the key 2 and value 2 values in the variable and replace none to that value 

#                             key_2 = list(dt["media"][x]["identifiers"][y]["columns"][1].keys())[0]
#                             # key 2 value is stored in the key 2 variable 

#                             value_2 = dt["media"][x]["identifiers"][y]["columns"][1][key_2]
#                             # value 2 value is stored in the value 2 variable 

#                     # print(key_2,":",value_2)
#                     # print(file_data[key_2].str.contains(value_2))
#                     # Keys and values and full data from the raw file will be passed in the function to get the dataframe sorted into separate types  
#                     # Find out the index column from the file instead of hard coding 
                    
#                     index_col = file_data.columns[0]
#                     # Storing the date column either it is Starting date or Date or something else 
#                     # print(file_data)
#                     dFrame = SeparatingDataframesForDifferentCategories(file_data=file_data,key_1=key_1,value_1=value_1,key_2=key_2,value_2=value_2, type = label_media_identifier,index_column=index_col,budget=TotalBudget,monthly_view_kpi=MonthlyViewKPI)
#                     # Calling the function to set each type of dataframe like facebook or ig or sc or twitter and setting it to the dictionary object and naming it to the label of that  

#                     dFrame_info = {
#                         "name": label_media_identifier,
#                         "data": dFrame
#                     }
#                     # print(label_media_identifier)
#                     # Dataframe info is stored in this object 

#                     # Weekly for each dataframe 
#                     weekly_data_for_each_dFrame = CalculatingWeeklyForDailyTotalDataFrame(dFrame)
#                     weekly_data_for_each_dFrame.index.name = "Weeks"
#                     weekly_data_for_each_dFrame.index +=1


#                     annual_total = weekly_data_for_each_dFrame.copy()
                    
                
#                     annual_total.drop(weekly_data_for_each_dFrame.index, inplace=True)
#                     # print(gdn_df_daily)
#                     for i in range(0, 1):
#                         row = weekly_data_for_each_dFrame.iloc[i:len(weekly_data_for_each_dFrame)].select_dtypes(include=['int64','double']).sum()
#                         row['CTR'] = row['Clicks'] / row['Impressions']
#                         # print(totalBudget)
#                         row['Budget'] = round(row['Budget'],2)
#                         row['SPENT_BUDGET'] = round((row['Budget']/TotalBudgetForDaily)*100)
#                         row['CTR'] = round(row['CTR']*100,2)
                    
#                         annual_total = annual_total.append(row, ignore_index=True)
                    
                    
#                     annual_total = annual_total.assign(MVKPI_Total = MonthlyViewKPI)





#                     # try:
#                     # weekly_data_for_each_dFrame.to_sql(label_media_identifier+"_weekly",con= db.engine,if_exists="replace",index=True,index_label="Weeks")
                    
#                     annual_total = annual_total.assign(CPC = round((annual_total["Budget"] / annual_total["Clicks"]),2))
#                     # annual_total.to_sql(label_media_identifier+"_final",con= db.engine,if_exists="replace",index=True)
#                     list_of_summary_medias_with_data.append({
#                         "label": label_media_identifier,
#                         "data":annual_total,
#                         "budget": TotalBudget,
#                     })
#                     # Data is stored in the excel file 
#                     # print(writer.sheets[label_media_identifier])
#                     # Find out number of columns 
                    
#                     weekly_data_for_each_dFrame.to_excel(writer, sheet_name=label_media_identifier, index=True,index_label=index_col)
#                     # print(weekly_data_for_each_dFrame)
                    
#                     max_row_weekly = len(list(weekly_data_for_each_dFrame.iloc[:, 0]))+1
#                     # type(writer.sheets[label_media_identifier])
#                     # try:
#                     #     if writer.sheets[label_media_identifier].max_row is not None:
#                     #         max_row = writer.sheets[label_media_identifier].max_row
#                     #     else:
#                     #         max_row = 0
#                     # except:
#                     #     max_row = 0
#                     annual_total.to_excel(writer, sheet_name=label_media_identifier,startrow=max_row_weekly, index=True,header=False)
#                     max_row_weekly = max_row_weekly+2

                    
#                     total_columns_dt = len(dFrame.columns)+2
#                     dFrame.to_excel(writer, sheet_name=label_media_identifier,startrow=max_row_weekly, index=True)

#                     # After adding the dataframe in the sheet - Work on LG.com -------------------------------------------
#                     lg_data_validity  = dt["media"][x]["identifiers"][y]["lg"]

#                     if lg_data_validity == 1:

#                         lg_data_files = dt["media"][x]["identifiers"][y]["lg-file"]

#                         lg_raw_file = pd.read_excel(lg_data_files)
#                         try:
#                             lg_raw_file['Date'] = pd.to_datetime(lg_raw_file['Date'], format="%Y%m%d")

#                             # Convert datetime objects to the desired format
#                             lg_raw_file['Date'] = lg_raw_file['Date'].dt.strftime("%Y-%m-%d %H:%M:%S")
#                         except:
#                             pass
#                         # File is read successfully having all the data 
#                         # Next step is to fetch only those data which is needed by using queries on the raw data
#                         # After getting the relavent data we will add it in dataframe to use it for daily and weekly report 
#                         # Calculating summary from the above weekly and daily info 
#                         # Saving in the file to see the output 
#                         # print(campaign_raw)
#                         # Get the unique column values
                
#                         my_dir = os.path.dirname(__file__)
                        
#                         sum_of_data = lg_raw_file.groupby('Date').sum()
#                         print(dFrame["Clicks"].index)
#                         date_format = "%Y-%m-%d"
                        
#                         sum_of_data['Bounce Rate'] = round((sum_of_data["Bounces"] / sum_of_data["Visits"])*100) 
#                         sum_of_data['CVR'] = round((sum_of_data["Visits"] / dFrame["Clicks"])*100) 
#                         sum_of_data['CPT'] = round((dFrame["Budget"] / sum_of_data["Visits"])*100) 
#                         # print("\n\nImage:",img_id," Record:\n")
#                         total_columns_dt = total_columns_dt +1
#                         sum_of_data.to_excel(writer, sheet_name=label_media_identifier, startrow=max_row_weekly, startcol=total_columns_dt, index=True)
                    
#                     # ------------------------- LG.com till here --------------------------------------------------
                    
#                     rawimage_validity = dt["media"][x]["identifiers"][y]["rawimage"]
#                     if rawimage_validity == 1:
#                         # Do everything related to rawimage here 

#                         rawimage_file = dt["media"][x]["identifiers"][y]["rawimage-file"]
#                         campaign_raw = pd.read_excel(rawimage_file)
#                         # File is read successfully having all the data 
#                         # Next step is to fetch only those data which is needed by using queries on the raw data
#                         # After getting the relavent data we will add it in dataframe to use it for daily and weekly report 
#                         # Calculating summary from the above weekly and daily info 
#                         # Saving in the file to see the output 
#                         # print(campaign_raw)
#                         # Get the unique column values
#                         dataframes_list = campaign_raw.groupby(['Campaign'])
#                         campaign_names = dataframes_list.groups.keys()  
#                         list_of_campaigns = list(campaign_names)
#                         my_dir = os.path.dirname(__file__)

#                         if lg_data_validity == 1:
#                             prevColumns = total_columns_dt + len(sum_of_data.columns)+1
#                         else:
#                             prevColumns = total_columns_dt +1
                     
#                         for camp_id in range(0,len(list_of_campaigns)):
#                             # Campaign_Dataframe
#                             camp = dataframes_list.get_group(list_of_campaigns[camp_id])
#                             # print(camp)
#                             # First Ad group dataframe 
#                             list_of_messages = camp.groupby(['Ad group'])
#                             message_names = list_of_messages.groups.keys()
#                             # print(message_names)
#                             message_names = list(message_names)
#                             for msg_id in range(0,len(message_names)):
#                                 msg = list_of_messages.get_group(message_names[msg_id])
#                                 # print(msg)
#                                 # first msg records
#                                 list_of_sizes = msg.groupby(['Images'])
#                                 sizes_img  = list(list_of_sizes.groups.keys())
#                                 # list_of_sizes_images = []
#                                 value = list_of_campaigns[camp_id].split("_")[-1]
                                
#                                 for img_id in range(0,len(sizes_img)):
#                                     sz_img = list_of_sizes.get_group(sizes_img[img_id])
#                                     # print(sz_img)
#                                     # first size of image records 
#                                     sum_of_data = sz_img.groupby('Day').sum()
#                                     sum_of_data['CTR'] = round(sum_of_data['Clicks'] / sum_of_data['Impr.']*100,2)
                                
                                
                                
                                    
#                                     # print("\n\nImage:",img_id," Record:\n")
#                                     # sheet_name = "gdn"
#                                     # list_of_sizes_images.append(sum_of_data)
#                                     newColumnLength = len(sum_of_data.columns) +1
#                                     # print(len(sum_of_data.columns) + 1 )
#                                     # Find Number of columns 
#                                     # sum_of_data.set
#                                     # Merge cells and name the row 2 with that name 
#                                     sum_of_data.columns = [[value for i in range(0,len(sum_of_data.columns))],[sizes_img[img_id] for i in range(0,len(sum_of_data.columns))],[message_names[msg_id] for i in range(0,len(sum_of_data.columns))],[key for key in sum_of_data.columns]]
#                                     sum_of_data.to_excel(writer, sheet_name=label_media_identifier, startrow=1, startcol=prevColumns)
#                                     prevColumns = prevColumns + newColumnLength +1

#                         # ---------------------------------------

        
#                     # ----------------------------
#                     # except:
#                     #     # db.engine.execute(f"DROP table IF EXISTS `{label_media_identifier}`")
#                     #     label_media_identifier_weekly=label_media_identifier+"_weekly"
#                     #     # db.engine.execute(f"DROP table IF EXISTS `{label_media_identifier_weekly}`")
#                     #     label_media_identifier_final=label_media_identifier+"_final"
#                     #     # db.engine.execute(f"DROP table IF EXISTS `{label_media_identifier_final}`")
#                     #     deleting_table = True


#                     list_of_dataframes.append(dFrame_info)
#                     # Appending the dataframe in the list of dataframe

#         # print(list_of_dataframes)
#         # Total from raw data according to date or reporting starts
#         temp = pd.DataFrame()
#         # Empty dataframe for combining all the dataframes in the daily total tab 

#         # Empty dataframe to append in all other dataframes 
        
#         # From here we are trying to concat all the dataframes for the daily total
        
#         for dt in list_of_dataframes:
#             # Loop to go through each dataframe and concat it to the previous dataframe 

#             if deleting_table == True:
#                 pass
#                 # db.engine.execute(f"DROP table IF EXISTS `{dt['name']}`")

#             else:

#                 # Commenting weekly label                   (will use later)
#                 # weekly_label = dt["name"] + "-weekly"
                

#                 try:
#                     if dt["data"].index.name != "Date":
#                         dt["data"].index.name = "Date"
#                     # dt["data"].to_sql(dt["name"],con=db.engine,if_exists="replace",index=True,index_label="Date")
                    
                    
#                     # Commenting weekly dataframe (Will use it later) 
#                     # weekly_dataframe = CalculatingWeeklyForDailyDataFrame(dt["data"],column_type=dt["name"])
#                     # weekly_dataframe.to_sql(weekly_label,con=db.engine,if_exists="replace",index=True)


#                     temp = pd.concat([temp, dt["data"]])
                    
#                 except:
#                     # db.engine.execute(f"DROP table IF EXISTS `{dt['name']}`")
#                     pass

        
#         # print(temp)
        

#                 # Commenting drop table weekly label (Use it later)
#                 # db.engine.execute(f"DROP table IF EXISTS `{weekly_label}`")
#         # print(temp)
#         # print(temp)
#         # For daily report tab and weekly for daily tab  
#         # print(temp)
        
#         # We are getting all the data together in one dataframe now we have to sum according to date 
#         # To do that we have to use pivot table function and add sum function 
#         # But first we have to create a variable in which we will store the columns 
#         # Also we need a variable for storing the index 
#         if deleting_table == True:
#             label_daily_total = "daily_total"
#             label_daily_total_weeks = "daily_total_weekly"
#             label_summary_total = "summary_total"

#             # db.engine.execute(f"DROP table IF EXISTS `{label_daily_total}`")
#             # db.engine.execute(f"DROP table IF EXISTS `{label_daily_total_weeks}`")
#             # db.engine.execute(f"DROP table IF EXISTS `{label_summary_total}`")

#         else:

#             index_name = "Date"
#             # print(temp)
#             column_values = temp.columns.ravel()
#             # print(column_values)
#             daily_total=  pd.pivot_table(temp, index=[index_name],values=column_values,aggfunc='sum')
#             # print(daily_total)


#             # print("====================================================")
#             # print(daily_total["Impressions"])
#             # print("====================================================")

#             # Now we need to calculate other columns of the daily total which has formulas 
#             # daily_total = SettingRemainingColumnsOfDailyTotalDataFrame(daily_total,TotalBudgetForDaily,TotalMonthlyViewKPI)

#             label_daily_total = "daily_total"
#             # daily_total.to_sql(label_daily_total,con = db.engine,if_exists="replace",index=True)
#             # daily_total.to_excel(writer, sheet_name=label_daily_total, index=True)
#             # print(daily_total)
#             daily_total_weekly_dataframe = CalculatingWeeklyForDailyTotalDataFrame(daily_total)
#             daily_total_weekly_dataframe.index.name = "Weeks"
#             daily_total_weekly_dataframe.index +=1
#             # daily_total_weekly_dataframe.to_sql("daily_total_weekly",con= db.engine,if_exists="replace",index=True,index_label="Weeks")

           
#             # daily_total_weekly_dataframe.to_excel(writer, sheet_name="daily_total_weekly", index=True)



#             # Calculating the total for daily total weekly dataframe 
#             #  
#             # daily_total_weekly_dataframe.index = 0
#             summary_total =  daily_total_weekly_dataframe.sum()
#             # Add formulas for formula columns 
            

#             # summary_total.to_sql("summary_total",con= db.engine, if_exists="replace",index=False)
            
            
            
#             # print(summary_total)


#             # 
#             # *********** 
#             # Summary logic =============================================================
#             # 

#             summary = daily_total_weekly_dataframe.copy()
                        
                    
#             summary.drop(summary.index, inplace=True)
#             # print(gdn_df_daily)
#             for i in range(0, 1):
#                 row = daily_total_weekly_dataframe.iloc[i:len(daily_total_weekly_dataframe)].select_dtypes(include=['int64','double']).sum()
#                 row['CTR'] = row['Clicks'] / row['Impressions']
#                 # print(totalBudget)
#                 row['Budget'] = round(row['Budget'],2)
#                 row['SPENT_BUDGET'] = round((row['Budget']/TotalBudgetForDaily)*100)
#                 row['CTR'] = round(row['CTR']*100,2)
            
#                 summary = summary.append(row, ignore_index=True)
            
            
#             summary = summary.assign(MVKPI_Total = TotalMonthlyViewKPI)
#             summary = summary.assign(CPC = round((summary["Budget"] / summary["Clicks"]),2))
#             # summary.to_sql("summary_total",con=db.engine,if_exists="replace",index=False)
#             list_of_summary_medias_with_data.append({
#                 "label": "Total",
#                 "data":summary,
#                 "budget": int(TotalBudgetForDaily),

#             })
#             daily_total_weekly_dataframe.to_excel(writer, sheet_name=label_daily_total, index=True)
       
#             max_row_summary = len(list(daily_total_weekly_dataframe.iloc[:, 0]))+1
#             summary.to_excel(writer, sheet_name=label_daily_total,startrow=max_row_summary, index=True,header=False)
#             max_row_summary +=2
           

#             daily_total.to_excel(writer, sheet_name=label_daily_total,startrow=max_row_summary, index=True)
#             # summary.to_excel(writer, sheet_name="summary_total", index=False)

#         # summary.to_excel(writer, sheet_name='summary', index=False)
#     # summary.to_excel(writer,con=db.engine,if_exist="replace",index=False)
    
#     # new_df.set_index('Date', inplace=True)
#     # daily_total_weekly_dataframe.to_excel(writer, sheet_name='daily_report', index=True,index_label="Date")
#     # writer = StylingSheets(writer, "daily_report")
#     #=============== Summary logic ends ============================================================
#         count_start = 0
        
#         df2 = pd.DataFrame() 
#         for dt in list_of_summary_medias_with_data:
#                 print(dt["data"]["Budget"].values[0])
#                 Media_Column = dt["label"]
#                 Budget_Column = dt["budget"]
#                 Spent_Column = dt["data"]["Budget"].values[0]
#                 SPENT_BUDGET_Column = dt["data"]["SPENT_BUDGET"].values[0]
#                 Clicks_Column = dt["data"]["Clicks"].values[0]
#                 CTR_Column = dt["data"]["CTR"].values[0]
#                 CPC = dt["data"]["CPC"].values[0] 
#                 df = pd.DataFrame([[Media_Column,Budget_Column,Spent_Column,SPENT_BUDGET_Column,Clicks_Column,CTR_Column,CPC]],columns=['Media','Budget','Spent','Budget Spent','Clicks','CTR','CPC'])
#                 df2 = df2.append(df)

#         # print(df2)
#         df2.to_excel(writer, sheet_name="Summary",index=False)

#     # Now next is weekly _total so we can move on to summary 

#     # daily_total_weekly_total =  pd.pivot_table(daily_total_weekly_dataframe,values=column_values,aggfunc = 'sum')


#     # print(list_of_dataframes)
    
#     return send_file(writer),200
#     # return jsonify({
#     #      "message": "Got data and stored in object successfully",
#     #      "status": 200
#     # },200)
#     # Returning the request to show that the data is receiving or not 


#  Formatting excel sheet function 

def FormatExcel():
    # In this function we will first send a file 
    # After fetching the file we will then apply our formating :
        # General Format : 
            #  - Font size : 10 px
            #  - Font family : 맑은 고딕 
            #  - Text color : black
            #  - Wrap Text : Enabled 
            #  - Border : All side border 
        # General Headings: 
            #  - Background color: 218, 238, 243
        # Weekly Headings: 
            #  - Background color: 235, 241, 222
        
        # First heading: 
            #  - Background color: black
            #  - Merged columns : True
            #  - Capital : True 
            # 

    pass




def SeparatingDataframesForDifferentCategories(file_data,key_1,value_1,key_2,value_2,type,index_column,budget,monthly_view_kpi):
    
    # getting dataframe to compare for first column to compare 
    data_frame = file_data[file_data[key_1].str.contains(value_1)]
    # So we are creating a dataframe and filtering the value 1 is present in the key 1 dataframe column 

    # print(data_frame[key_2].str.contains(value_2))
    if key_2 != None and value_2 != None:
        # If key 2 and value 2 is not none then we can filter 2nd column according to that and get the dataframe 
    # getting dataframe to compare for second column to compare 
        data_frame = data_frame[data_frame[key_2].str.contains(value_2)]    
        # Setting dataframe ...
    # print(data_frame)
    column_values = data_frame.columns.ravel()
    # Column names for the dataframe will be stored in the column values 

    # print(column_values)
    data_frame[index_column] = pd.to_datetime(data_frame[index_column])
    # Converting the index column to date time so we can have a same format for all dates and not mixed like text and date .........

    # data_frame[index_column] = data_frame[index_column].astype("datetime")
    data_frame = pd.pivot_table(data_frame, index=[index_column],values=column_values,aggfunc='sum')  
    # We are summing the dataframe and storing the dataframe with sum of similar dates record 
    # print(data_frame)
    data_frame.index.name = "Date"
    data_frame = SettingRemainingColumnsOfDailyTotalDataFrame(df=data_frame,budget=budget,monthly_view_kpi = monthly_view_kpi)
    # Now we want to calculate the remaining columns here so we are passing dataframe and budget and monthly view kpi in the function .....
    
    data_frame = data_frame.assign(type = type)
 
    # returning the dataframe with comparing two columns where its same 
    return data_frame


def SettingRemainingColumnsOfDailyTotalDataFrame(df,budget,monthly_view_kpi):
    df['Budget'] = (round(df['Budget'],2)).replace([np.inf, -np.inf,np.nan], 0)
    df = df.assign(DAILY_VIEW_KPI = (round(float(monthly_view_kpi) / 30)))
    df = df.assign(DAILY_KPI_ACHIEVEMENT = (round((df['View 100%'] / df['DAILY_VIEW_KPI'])*100)).replace([np.inf, -np.inf], 0))
    df = df.assign(SPENT_BUDGET = (round((df['Budget'].cumsum()/int(budget))*100)).replace([np.inf, -np.inf], 0))
    df = df.assign(CPV_COMPLETE = (round(df["Budget"]/df["View 100%"],4)).replace([np.inf, -np.inf], 0))
    df = df.assign(CTR = (round(((df["Clicks"]/df["Impressions"])*100),2).replace([np.inf, -np.inf], 0)))
    try:
        df = df.assign(CPV_TRUEVIEW = (round(df["Budget"]/df["ThruPlays"],4)).replace([np.inf, -np.inf], 0))
    except:
        df = df.assign(CPV_TRUEVIEW = (round(df["Budget"]/df["TrueView: Views"],4)).replace([np.inf, -np.inf], 0))
    
        
    return df

def CalculatingWeeklyForDailyDataFrame(df_daily,column_type):
    date_rng = pd.date_range(start=df_daily.index.min(), end=df_daily.index.max(), freq='D')
    # print("Fine")
    # Fetching the start and end date so we can loop through it for weeks 
    df_weekly = df_daily.copy()
    # Making a copy of daily table so we can use the same columns for weekly as well
    # print("Still fine")
    df_weekly.drop(df_weekly.index, inplace=True)
    # We are dropping all the data so we can add new data inside weekly 
    # print("still still fine")

    # print(gdn_df_daily)
    # Starting the loop for weeks 
    for i in range(0, len(date_rng), 7):
        try:
            row = df_daily.iloc[i:i+7].select_dtypes(include=['int64','double']).sum()
        except:
            print("not fine 1")
        # Doing sum for the row from 1 - 7 days and storing in row 
        try:
            row['Budget'] = round(row['Budget'],2)
        # Changing the Format of the budget 
        except:
            print("not fine 2")
        try:
            row['CTR']=round(((row['Clicks']/row['Impressions'])*100),2)
        except:
            print("not fine 3")
        # Changing the CTR for the week 
        # row['CTR'] = round(row['CTR'],2)
        # Changing the format for CTR

        try:
            row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
        except:
            print("not fine 4")
        
        try:
            try:
                if row["ThruPlays"] == 0:
                    row['CPV_TRUEVIEW'] = 0
                else:
                    row['CPV_TRUEVIEW'] = round(row["Budget"]/row["ThruPlays"],4)

            except:
                if row["TrueView: Views"] == 0:
                    row['CPV_TRUEVIEW'] = 0
                else:
                    row['CPV_TRUEVIEW'] = round(row["Budget"]/row["TrueView: Views"],4)
        except:
            print("not fine 5")
        # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
        if row["View 100%"] == 0:
            row['CPV_COMPLETE'] = 0
        else:
            row['CPV_COMPLETE'] = round(row["Budget"]/row["View 100%"],4)

        if i+6 > len(date_rng):
          
            row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][len(date_rng)-1]
        else : 
            # row['Budget'] = round(row['Budget'],2)
            row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][i+6]
            # row['CTR'] = round(row['CTR'],2)
            # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
            # row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
            row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
            # if type == "fb_video" or type == "fb_image" or type=="ig_image" or type=="ig_video":
            # else:
            #     row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
        row["type"] = column_type 
        df_weekly = df_weekly.append(row, ignore_index=True)
    # print("absolutely fine")
    # del df_weekly['type']
    return df_weekly 


def CalculatingWeeklyForDailyTotalDataFrame(df_daily):
    date_rng = pd.date_range(start=df_daily.index.min(), end=df_daily.index.max(), freq='D')
    # print("Fine")
    # Fetching the start and end date so we can loop through it for weeks 
    df_weekly = df_daily.copy()
    # Making a copy of daily table so we can use the same columns for weekly as well
    # print("Still fine")
    df_weekly.drop(df_weekly.index, inplace=True)
    # We are dropping all the data so we can add new data inside weekly 
    # print("still still fine")

    # print(gdn_df_daily)
    # Starting the loop for weeks 
    for i in range(0, len(date_rng), 7):
        try:
            row = df_daily.iloc[i:i+7].select_dtypes(include=['int64','double']).sum()
        except:
            print("not fine 1")
        # Doing sum for the row from 1 - 7 days and storing in row 
        try:
            row['Budget'] = round(row['Budget'],2)
        # Changing the Format of the budget 
        except:
            print("not fine 2")
        try:
            row['CTR']=round(((row['Clicks']/row['Impressions'])*100),2)
        except:
            print("not fine 3")
        # Changing the CTR for the week 
        # row['CTR'] = round(row['CTR'],2)
        # Changing the format for CTR

        try:
            row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
        except:
            print("not fine 4")
        
        try:
            try:
                if row["ThruPlays"] == 0:
                    row['CPV_TRUEVIEW'] = 0
                else:
                    row['CPV_TRUEVIEW'] = round(row["Budget"]/row["ThruPlays"],4)

            except:
                if row["TrueView: Views"] == 0:
                    row['CPV_TRUEVIEW'] = 0
                else:
                    row['CPV_TRUEVIEW'] = round(row["Budget"]/row["TrueView: Views"],4)
        except:
            print("not fine 5")
        # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
        if row["View 100%"] == 0:
            row['CPV_COMPLETE'] = 0
        else:
            row['CPV_COMPLETE'] = round(row["Budget"]/row["View 100%"],4)

        print("Position of index:",i+6)
        print(len(date_rng))
        if i+6 >= len(date_rng):
            # print(row["SPENT_BUDGET"])
            row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][len(date_rng)-1]
        else : 
            # row['Budget'] = round(row['Budget'],2)
            row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][i+6]
            # row['CTR'] = round(row['CTR'],2)
            # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
            # row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
            row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
            # if type == "fb_video" or type == "fb_image" or type=="ig_image" or type=="ig_video":
            # else:
            #     row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
        
        df_weekly = df_weekly.append(row, ignore_index=True)
    # print("absolutely fine")
    # del df_weekly['type']
    return df_weekly 



# New update from here
def MatchingColumnValue(file,columns_list):
        list_of_columns  = list(file.columns)
        matchedColumnSet = set(list_of_columns) & set(columns_list)
        matchedColumn = False
        for i in matchedColumnSet:
            matchedColumn = i
            break
        return matchedColumn


#  Raw Column set 
class Raw_Columns:
    def __init__(self, date,platform,product,budget,impressions,clicks,view,percent_25,percent_50,percent_75,percent_100,language,country,asset_type):
        self.date = date
        self.platform = platform
        self.product = product
        self.impressions = impressions
        self.asset_type=asset_type
        self.clicks = clicks
        self.budget = budget
        self.view = view
        self.percent_25 = percent_25
        self.percent_50 = percent_50
        self.percent_75 = percent_75
        self.percent_100 = percent_100
        self.language = language
        self.country = country

    def format_cells(self):
        Formats = {
            "percentage": '0%',
            "comma_format": '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-',
            "two_decimal_percentage": '0.00%',
            "two_decimal_dollar":'_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
            "five_digit":'#,##0',
            "four_decimal_dollar":'_("$"* #,##0.0000_);_("$"* (#,##0.0000);_("$"* "-"????_);_(@_)'
            
        }
        return [
            "$#,##0.00"
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
        ]
    def ListOfColumns(self):
        return [
            "date"
            ,"platform"
            ,"product"
            ,"budget"
            ,"impressions"
            ,"clicks"
            ,"view"
            ,"percent_25"
            ,"percent_50"
            ,"percent_75"
            ,"percent_100"
            ,"language"
            ,"country"
            ,"asset_type"
        ]
    def ListOfColumnsForPlatformWeekly(self):
        return [
           
            "budget"
            ,"impressions"
            ,"clicks"
            ,"view"
            ,"percent_25"
            ,"percent_50"
            ,"percent_75"
            ,"percent_100"
        ]
    def ListOfColumnsForPlatform(self):
        return [
            "date"
            ,"budget"
            ,"impressions"
            ,"clicks"
            ,"view"
            ,"percent_25"
            ,"percent_50"
            ,"percent_75"
            ,"percent_100"
        ]
    
    def ListOfColumnsAssets(self):
        return [
            "budget"
            ,"impressions"
            ,"clicks"
            ,"view"
            
        ]
    def Format_for_Weekly_ListOfColumnsForPlatform(self):
        Formats = {
            "percentage": '0%',
            "comma_format": '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-',
            "two_decimal_percentage": '0.00%',
            "two_decimal_dollar":'_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
            "five_digit":'#,##0',
            "four_decimal_dollar":'_("$"* #,##0.0000_);_("$"* (#,##0.0000);_("$"* "-"????_);_(@_)'
            
        }   
        return ["$#,##0.00"
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]]
    
    def Format_for_ListOfColumnsForPlatform(self):
        Formats = {
            "percentage": '0%',
            "comma_format": '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-',
            "two_decimal_percentage": '0.00%',
            "two_decimal_dollar":'_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
            "five_digit":'#,##0',
            "four_decimal_dollar":'_("$"* #,##0.0000_);_("$"* (#,##0.0000);_("$"* "-"????_);_(@_)'
            
        }   
        return ["$#,##0.00"
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]]

    def Format_for_ListOfColumnsForPlatformSummary(self):
        Formats = {
            "percentage": '0%',
            "comma_format": '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-',
            "two_decimal_percentage": '0.00%',
            "two_decimal_dollar":'_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
            "five_digit":'#,##0',
            "four_decimal_dollar":'_("$"* #,##0.0000_);_("$"* (#,##0.0000);_("$"* "-"????_);_(@_)'
            
        }   
        return [
            Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]]
    def format_cells_daily(self):
        Formats = {
            "percentage": '0%',
            "comma_format": '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-',
            "two_decimal_percentage": '0.00%',
            "two_decimal_dollar":'_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
            "five_digit":'#,##0',
            "four_decimal_dollar":'_("$"* #,##0.0000_);_("$"* (#,##0.0000);_("$"* "-"????_);_(@_)'
            
        }

        return ["$#,##0.00"
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]]
    
    def ListOfColumnsDaily(self):
        
        return [
            "budget"
            ,"impressions"
            ,"clicks"
            ,"view"
        ]

def MatchingColumn(file):
        Date_columns = ["Reporting starts","Start Time","Time period", "Date","Day","date","Start"]
        Platform_columns = ["Platform","platform","Plaform"]
        Product_columns = ["Product", "Products"]
        budget_columns = ["Amount spent (USD)","Amount spent","Amount Spent","Cost","Media Cost (Advertiser Currency)","Media Cost"]
        impression_columns = ["Impressions","Paid Impressions","impressions","Impr.","Impressions","Impression"]
        click_columns = ["Link clicks","Swipe Ups","Clicks","clicks"]
        view = ["3-second video plays","2 Second Video Views (View Time Only)", "2 Second Video Views","6s video views","TrueView: Views","Video first quartile","view","View (Video)"]
        percent_25_columns = ["Video Plays at 25%","Video plays at 25%","Video Attachment Plays at 25%", "Video played 25%", "First-Quartile Views (Video)","Video first quartile","percent_25","25%e Views (Video)","25% Views (Video)"]
        percent_50_columns = ["Video Plays at 50%","Video plays at 50%","Video Attachment Plays at 50%", "Video played 50%", "Midpoint Views (Video)","Video midpoint","percent_50","50% View (Video)"]
        percent_75_columns = ["Video Plays at 75%","Video plays at 75%","Video Attachment Plays at 75%", "Video played 75%", "Third-Quartile Views (Video)","Video third quartile","percent_75","75% View (Video)"]
        percent_100_columns = ["Video Completions","Video plays at 100%","Video Attachment Completions", "Video completions", "Complete Views (Video)","Video completes","percent_100","Complete Views (Video)"]
        country_columns = ["Market","Tget Mket","Country"]
        language_columns = ["Language","language","Langauge"]
        asset_type_columns = ["Creative Type","Creative"]


        obj = Raw_Columns(MatchingColumnValue(file,Date_columns),
        MatchingColumnValue(file,Platform_columns),
        MatchingColumnValue(file,Product_columns),
        MatchingColumnValue(file,budget_columns),
        MatchingColumnValue(file,impression_columns),
        MatchingColumnValue(file,click_columns),
        MatchingColumnValue(file,view),
        MatchingColumnValue(file,percent_25_columns),
        MatchingColumnValue(file,percent_50_columns),
        MatchingColumnValue(file,percent_75_columns),
        MatchingColumnValue(file,percent_100_columns),
        MatchingColumnValue(file,language_columns),
        MatchingColumnValue(file,country_columns),
        MatchingColumnValue(file,asset_type_columns))

        return obj

# -------------------------------



# Raw Column Assets set 

class Raw_Columns_Assets:
    def __init__(self, date, ad_name,campaign_name,creative,budget,impressions,clicks,view,percent_25,percent_50,percent_75,percent_100):
        self.date = date
        self.ad_name = ad_name
        self.campaign_name = campaign_name
        self.creative = creative
        self.budget = budget
        self.impressions = impressions
        self.clicks = clicks
        self.view = view
        self.percent_25 = percent_25
        self.percent_50 = percent_50
        self.percent_75 = percent_75
        self.percent_100 = percent_100

    def format_cells(self):
        Formats = {
            "percentage": '0%',
            "comma_format": '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-',
            "two_decimal_percentage": '0.00%',
            "two_decimal_dollar":'_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
            "five_digit":'#,##0',
            "four_decimal_dollar":'_("$"* #,##0.0000_);_("$"* (#,##0.0000);_("$"* "-"????_);_(@_)'
            
        }
        return [
            "$#,##0.00"
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            # ,Formats["comma_format"]
            # ,Formats["comma_format"]
            # ,Formats["comma_format"]
            # ,Formats["comma_format"]
        ]
    def ListOfColumns(self):
        return [
            "budget"
            ,"impressions"
            ,"clicks"
            ,"view"
            # ,"percent_25"
            # ,"percent_50"
            # ,"percent_75"
            # ,"percent_100"
        ]
    def format_cells_daily(self):
        Formats = {
            "percentage": '0%',
            "comma_format": '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-',
            "two_decimal_percentage": '0.00%',
            "two_decimal_dollar":'_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
            "five_digit":'#,##0',
            "four_decimal_dollar":'_("$"* #,##0.0000_);_("$"* (#,##0.0000);_("$"* "-"????_);_(@_)'
            
        }

        return ["$#,##0.00"
            ,Formats["comma_format"]
            ,Formats["comma_format"]
            ,Formats["comma_format"]]
    
    def ListOfColumnsDaily(self):
        
        return [
            "budget"
            ,"impressions"
            ,"clicks"
            ,"view"
        ]

def MatchingColumnAsset(file):
        Date_columns = ["Reporting starts","Start Time","Time period", "Date","Day","date"]
        Platform_columns = ["Platform","platform"]
        objective_columns = ["Objective","objective"]
        ad_group_columns = ["Ad group/Ad set name /Line item","Ad group/Ad set name /Line item ","Ad group / Ad set name / Line item","Ad group / Ad set name / Line item "]
        ad_name_columns = ["Ad name","Creative name","ad_name","Ad name /YouTube Ad","Ad name / YouTube Ad"]
        camp_name_columns = ["Campaign name","Insertion Order","campaign_name","Campaign / Campaign name   / Insertion Order","Campaign / Campaign name / Insertion Order"]
        budget_columns = ["Amount spent (USD)","Amount spent","Amount Spent","Spend","Total Media Cost (Advertiser Currency)", "Total ad cost","budget","Revenue (Adv Currency)","Spend / Revenue","Spend /Revenue "]
        impression_columns = ["Impressions","Paid Impressions","impressions","impr.","Impre.","Impression"]
        click_columns = ["Link clicks","Swipe Ups","Clicks","clicks"]
        conversion_columns = ["Conversion"]
        view = ["3-second video plays", "2 Second Video Views","6s video views","TrueView: Views","Video first quartile","view","View (Video)"]
        percent_25_columns = ["Video plays at 25%","Video Attachment Plays at 25%", "Video played 25%", "First-Quartile Views (Video)","Video first quartile","percent_25","25%e Views (Video)","25% Views (Video)"]
        percent_50_columns = ["Video plays at 50%","Video Attachment Plays at 50%", "Video played 50%", "Midpoint Views (Video)","Video midpoint","percent_50","50% View (Video)"]
        percent_75_columns = ["Video plays at 75%","Video Attachment Plays at 75%", "Video played 75%", "Third-Quartile Views (Video)","Video third quartile","percent_75","75% View (Video)"]
        percent_100_columns = ["Video plays at 100%","Video Attachment Completions", "Video completions", "Complete Views (Video)","Video completes","percent_100","Complete Views (Video)"]
        asset_type_columns = ["Asset type"]
        size_columns = ["Size"]
        language_columns = ["Language"]
        country_columns = ["Country"]
        message_columns = ["Message"]

        obj = Raw_Columns(MatchingColumnValue(file,Date_columns),
        MatchingColumnValue(file,Platform_columns),
        MatchingColumnValue(file,objective_columns),
        MatchingColumnValue(file,ad_group_columns),
        MatchingColumnValue(file,ad_name_columns),
        MatchingColumnValue(file,camp_name_columns),
        MatchingColumnValue(file,budget_columns),
        MatchingColumnValue(file,impression_columns),
        MatchingColumnValue(file,click_columns),
        MatchingColumnValue(file,conversion_columns),
        MatchingColumnValue(file,view),
        MatchingColumnValue(file,percent_25_columns),
        MatchingColumnValue(file,percent_50_columns),
        MatchingColumnValue(file,percent_75_columns),
        MatchingColumnValue(file,percent_100_columns),
        MatchingColumnValue(file,asset_type_columns),
        MatchingColumnValue(file,size_columns),
        MatchingColumnValue(file,language_columns),
        MatchingColumnValue(file,country_columns),
        MatchingColumnValue(file,message_columns))

        return obj

#  -----------------------------   
def summary_calculation(wb,ws,i,media_name,fb_obj,fb_total,have_column,media_label):
    #...................................................
    # Start working on the summary and show to LJ
    #...................................................
        
        if have_column == True:
            columns_label = ["Media","Period","Budget","Budget Spent","Budget Spent %","Impression","Click","CTR","CPC","Visit from Media","Visit KPI","KPI%","CPT","CVR (Media to LG.COM)","Bounce Rate"]
            total_columns = len(columns_label)
            for j in range(0,total_columns):
                cell_info = ws.cell(row=1, column=j+1, value=columns_label[j])
                set_border_and_align_weekly(cell_info)

            
        cell_info=ws.cell(row=i,column=1,value=media_label)
        if media_label == "Total":
            set_border_and_red_bg(cell_info)
        else:
            set_border_and_align(cell_info)

        
        cell_info=ws.cell(row=i,column=2,value="")
        if media_label == "Total":
            set_border_and_red_bg(cell_info)
        else:
            set_border_and_align(cell_info)
        
        set_border_and_align(cell_info)
        Budget = ws.cell(row=i,column=3,value=f"={media_name}!T2")
        Budget.number_format = '$#,##0.00'
        if media_label == "Total":
            set_border_and_red_bg(Budget)
        else:
            set_border_and_align(Budget)

        Budget_Spent = ws.cell(row=i,column=4,value=fb_total[fb_obj.budget])
        Budget_Spent.number_format = '$#,##0.00'
        if media_label == "Total":
            set_border_and_red_bg(Budget_Spent)
        else:
            set_border_and_align(Budget_Spent)

        Budget_Spent_Percentage = ws.cell(row=i,column=5,value=f"=IFERROR((D{i}/C{i}),0)")
        Budget_Spent_Percentage.number_format = '0.00%'
        set_border_and_align(Budget_Spent_Percentage)
        if media_label == "Total":
            set_border_and_red_bg(Budget_Spent_Percentage)
        else:
            set_border_and_align(Budget_Spent_Percentage)

        cell_info = ws.cell(row=i,column=6,value=fb_total[fb_obj.impressions])
        if media_label == "Total":
            set_border_and_red_bg(cell_info)
        else:
            set_border_and_align(cell_info)

        cell_info = ws.cell(row=i,column=7,value=fb_total[fb_obj.clicks])
        if media_label == "Total":
            set_border_and_red_bg(cell_info)
        else:
            set_border_and_align(cell_info)

        CTR = ws.cell(row=i,column=8,value=f"=IFERROR((G{i}/F{i}),0)")
        CTR.number_format = '0.00%'
        set_border_and_align(CTR)
        if media_label == "Total":
            set_border_and_red_bg(CTR)
        else:
            set_border_and_align(CTR)

        CPC=ws.cell(row=i,column=9,value=f"=IFERROR(ROUND(D{i}/G{i},2),0)")
        CPC.number_format = '$#,##0.00'
        set_border_and_align(CPC)
        if media_label == "Total":
            set_border_and_red_bg(CPC)
        else:
            set_border_and_align(CPC)


        cell_info = ws.cell(row=i,column=10,value="")
        set_border_and_align(cell_info)
        if media_label == "Total":
            set_border_and_red_bg(cell_info)
        else:
            set_border_and_align(cell_info)


        cell_info = ws.cell(row=i,column=11,value="")
        set_border_and_align(cell_info)
        if media_label == "Total":
            set_border_and_red_bg(cell_info)
        else:
            set_border_and_align(cell_info)

        cell_info = ws.cell(row=i,column=12,value="")
      
        if media_label == "Total":
            set_border_and_red_bg(cell_info)
        else:
            set_border_and_align(cell_info)

        CPT = ws.cell(row=i,column=13,value=f"=IFERROR(D{i}/J{i},0)")
        CPT.number_format = '$#,##0.00'
        if media_label == "Total":
            set_border_and_red_bg(CPT)
        else:
            set_border_and_align(CPT)
        
        CVR = ws.cell(row=i,column=14,value=f"=IFERROR(J{i}/G{i},0)")
        CVR.number_format = '0.00%'
        if media_label == "Total":
            set_border_and_red_bg(CVR)
        else:
            set_border_and_align(CVR)

        Bounce_rate = ws.cell(row=i,column=15,value = 0)
        Bounce_rate.number_format = '0.00%'
        Bounce_rate.number_format = '0.00%'
        if media_label == "Total":
            set_border_and_red_bg(Bounce_rate)
        else:
            set_border_and_align(Bounce_rate)
        # i = 1
        
        # ws.cell(row=i,column=20,value="Total Budget")
        # # Loop to Print the columns label with index

        # counter= 0 
        # for j in range(total_columns+1,total_columns+total_columns_calculated+1):
        #     ws.cell(row=1, column=j+1, value=list_of_calculated_columns[counter])
        #     counter+=1

        # i=1 
        # starting_row = 1
        # total_cost = "0"
        # for index, row in fb_weekly.iterrows():

        #     for j in range(0,total_columns):
        #         if j == 0:
        #             ws.cell(row=starting_row+1,column=j+1,value=str(index))
        #         try:  
        #             value = fb_weekly[getattr(fb_obj, list_of_columns[j])][i-1]
        #         except:
        #             value = 0
        #         # value = getattr(fb_obj, list_of_columns[j])
        #         ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")

        #     #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        #     # Formulas here 
        #     ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR(ROUND((D{str(i+1)}/C{str(i+1)})*100,2),0)")
        #     ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR(ROUND((({total_cost}+B{str(i+1)})/T2),2)*100,0)")
        #     CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(ROUND(B{str(i+1)}/D{str(i+1)},2),0)")
        #     CPC.number_format = '$#,##0.00'
        #     CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
        #     CPV.number_format = '$#,##0.00'
        #     total_cost += f"+B{str(starting_row+1)}"
        #     starting_row+=1
        #     i+=1
        #     # ------------------------

        # # Show total here -----------------
        # i=1 
        
        # #  We have to go in each of the attribute and store it 
        # # print(list_of_columns)
        # # for 
        # total_cost = "0"
        
        # print(fb_total.index)
        # for j in range(0,total_columns):
        #     if j == 0:
        #         ws.cell(row=starting_row+1,column=j+1,value=str("Total"))
        #     try:  
        #         value = fb_total[getattr(fb_obj, list_of_columns[j])]
        #     except:
        #         value = 0
        #     # value = getattr(fb_obj, list_of_columns[j])
        #     ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")
        # #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        # # Formulas here 
        # ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR(ROUND((D{str(starting_row+1)}/C{str(starting_row+1)})*100,2),0)")
        # ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR(ROUND((({total_cost}+B{str(starting_row+1)})/T2),2)*100,0)")
        # CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/D{str(starting_row+1)},2),0)")
        # CPC.number_format = '$#,##0.00'
        # CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
        # CPV.number_format = '$#,##0.00'
        # total_cost += f"+B{str(starting_row+1)}"
        # starting_row+=1
        # # ---------------------------------

        # i = 1
        # starting_row = starting_row+2
        # total_cost = "0"

        # for index, row in fb.iterrows():
        #     # print(index," rows: ",row)

        #     for j in range(0,total_columns):
        #         if j == 0:
        #             ws.cell(row=starting_row+1,column=j+1,value=str(index))
        #         try:   
        #             value = fb[getattr(fb_obj, list_of_columns[j])][i-1]
        #         except:
        #             value = 0
        #         # value = getattr(fb_obj, list_of_columns[j])
        #         ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")

        #     # list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        #     # Formulas here 
        #     ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR(ROUND((D{str(starting_row+1)}/C{str(starting_row+1)})*100,2),0)")
        #     ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR(ROUND((({total_cost}+B{str(starting_row+1)})/T2),2)*100,0)")
        #     CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/D{str(starting_row+1)},2),0)")
        #     CPC.number_format = '$#,##0.00'
        #     CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
        #     CPV.number_format = '$#,##0.00'

        #     total_cost += f"+B{str(starting_row+1)}"
        #     starting_row+=1
        #     # ------------------------


            
            
        #     i+=1
        
        
        
        
        
        
        wb.save('summary_testing.xlsx')




def media_calculation(wb,ws,i,fb_ig_obj,fb,fb_weekly,fb_total,daily = False):
    #...................................................
    # Start working on the summary and show to LJ
    #...................................................

    # Loop for columns 
        # print(fb_total)
        if daily == True:
            list_of_columns = fb_ig_obj.ListOfColumnsDaily()
            list_of_columns_number_format=fb_ig_obj.format_cells_daily()
        else:
            list_of_columns = fb_ig_obj.ListOfColumnsForPlatformWeekly()
            
            list_of_columns_daily_record = fb_ig_obj.ListOfColumnsForPlatformWeekly()
            columns_daily = fb_ig_obj.ListOfColumnsForPlatform()
            list_of_columns_number_format = fb_ig_obj.Format_for_ListOfColumnsForPlatform()
            list_of_columns_summary_number_format = fb_ig_obj.Format_for_ListOfColumnsForPlatformSummary()

        # print(len(list_of_columns_number_format))
        # List of Columns of the class
        list_of_calculated_columns = ["CTR","Spent Budget %", "CPC","CPV"]
        # List of Columns Which has formulas
        # ws.cell(row=i+1, column=total_columns+1, value=f"=A{str(i+1)}/T2")

        total_columns = len(list_of_columns)
        
        # print(list_of_columns)
        # 9  - list of total columns 
        # Total Columns length
        total_columns_calculated =len(list_of_calculated_columns)
        # 4 - columns 
        # Total Columns calculated length 

        i = 1
        for j in range(0,total_columns):
        # Going to each column 
            if j==0:
                # If the column is date 
                cell_info = ws.cell(row=i,column=j+1,value="date")
                # We are setting column name and setting the value date 
                set_border_and_align_weekly(cell_info)
                # Formatting the cell and adding border and weekly 
            cell_info = ws.cell(row=i, column=j+2, value=list_of_columns[j])
            # cell value will be other columns 
            set_border_and_align_weekly(cell_info)
        # We are setting columns in this loop 



        cell_info = ws.cell(row=i,column=20,value="Total Budget")
        # Then we are setting the box with the text total budget and will add the value manually there 


        set_border_and_align(cell_info)
        # Formatting the cell alignment and border 
        # Loop to Print the columns label with index

        counter= 0 
        for j in range(total_columns+1,total_columns+total_columns_calculated+1):
            cell_info = ws.cell(row=1, column=j+1, value=list_of_calculated_columns[counter])
            set_border_and_align_weekly(cell_info)
            counter+=1
        # We are setting the column border and alignment for other columns that are needed 
        i=1 
        starting_row = 1
        total_cost = "0"
        for index, row in fb_weekly.iterrows():
            # Looping through weekly record and adding value and formating the cell border and alignment 
            for j in range(0,total_columns):
                if j == 0:
                    cell_info = ws.cell(row=starting_row+1,column=j+1,value=str(index+1))
                    set_border_and_align(cell_info)
                # try:  
                # print(getattr(fb_ig_obj, list_of_columns[j]))
                # print(fb_weekly.columns)
                value = fb_weekly[list_of_columns[j]][i-1]
                    
                    
                # except:
                #     value = 0
                # value = getattr(fb_ig_obj, list_of_columns[j])

                cell_info = ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")
                cell_info.number_format = list_of_columns_summary_number_format[j]
                set_border_and_align(cell_info)

            #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
            # Formulas here 
            cell_info= ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR((D{str(i+1)}/C{str(i+1)}),0)")
            cell_info.number_format = '0.00%'
            set_border_and_align(cell_info)
            cell_info = ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR((({total_cost}+B{str(i+1)})/T2),0)")
            cell_info.number_format = '0.00%'
            set_border_and_align(cell_info)
            CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(ROUND(B{str(i+1)}/D{str(i+1)},2),0)")
            CPC.number_format = '$#,##0.00'
            set_border_and_align(CPC)
            set_border_and_align(CPC)
            CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
            CPV.number_format = '$#,##0.00'
            set_border_and_align(CPV)
            set_border_and_align(CPV)
            total_cost += f"+B{str(starting_row+1)}"
            starting_row+=1
            i+=1
            # ------------------------

# Show total here -----------------
        
        start_row = starting_row
        i=1 
        
        #  We have to go in each of the attribute and store it 
        # print(list_of_columns)
        # for 
        total_cost = "0"
        
        # print(fb_total.index)
        for j in range(0,total_columns):
            if j == 0:
                cell_info = ws.cell(row=starting_row+1,column=j+1,value=str("Total"))
                set_border_and_red_bg(cell_info)
            value = fb_total[list_of_columns[j]][0]
            # print(value)
            
            # value = getattr(fb_ig_obj, list_of_columns[j])
            cell_info = ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")
            cell_info.number_format = list_of_columns_number_format[j]
            set_border_and_red_bg(cell_info)
        #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        # Formulas here 
        cell_info = ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR((D{str(starting_row+1)}/C{str(starting_row+1)}),0)")
        cell_info.number_format = '0.00%'
        set_border_and_red_bg(cell_info)
        cell_info = ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR((({total_cost}+B{str(starting_row+1)})/T2),0)")
        cell_info.number_format = '0.00%'
        set_border_and_red_bg(cell_info)
        
        CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(B{str(starting_row+1)}/D{str(starting_row+1)},0)")
        CPC.number_format = '$#,##0.00'
        set_border_and_red_bg(CPC)

        CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
        CPV.number_format = '$#,##0.00'
        set_border_and_red_bg(CPV)
        total_cost += f"+B{str(starting_row+1)}"
        starting_row+=1
# ---------------------------------

        i = starting_row+2
        # Column printing for daily table 
        for j in range(0,total_columns):
            if j==0:
                cell_info = ws.cell(row=i,column=j+1,value="date")
                set_border_and_align_weekly(cell_info)

            cell_info = ws.cell(row=i, column=j+2, value=list_of_columns_daily_record[j])
            set_border_and_align_weekly(cell_info)
        
        # ------------------------------------
        counter= 0 
        
        # Formulated columns printing for daily table
        for j in range(total_columns+1,total_columns+total_columns_calculated+1):
            cell_info = ws.cell(row=i, column=j+1, value=list_of_calculated_columns[counter])
            set_border_and_align_weekly(cell_info)
            counter+=1

        # Daily table printing-------
        i = 1
        start_col = starting_row + 2
        starting_row = starting_row+2
        total_cost = "0"

        weekly_check = 1
        for index, row in fb.iterrows():
            # print(index," rows: ",row)
            
            for j in range(0,total_columns):
                if j == 0:
                    value_date = fb[columns_daily[0]][i-1]
                    cell_info = ws.cell(row=starting_row+1,column=j+1,value=str(value_date))
                    set_border_and_align(cell_info)
                 
                value = fb[list_of_columns_daily_record[j]][i-1]
                    
              
                # value = getattr(fb_ig_obj, list_of_columns[j])
                cell_info = ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")
                cell_info.number_format = list_of_columns_number_format[j]
                set_border_and_align(cell_info)

            # list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
            # Formulas here 
            cell_info = ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR((D{str(starting_row+1)}/C{str(starting_row+1)}),0)")
            cell_info.number_format = '0.00%'
            set_border_and_align(cell_info)
            if weekly_check%7 == 1 and weekly_check != 1 and weekly_check != 0:
                cell_info = ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR((({total_cost})/T2),0)")
            else:
                cell_info = ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR((({total_cost}+B{str(starting_row+1)})/T2),0)")
            cell_info.number_format = '0.00%'
            set_border_and_align(cell_info)

            CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/D{str(starting_row+1)},2),0)")
            CPC.number_format = '$#,##0.00'
            set_border_and_align(CPC)
            CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
            CPV.number_format = '$#,##0.00'
            set_border_and_align(CPV)
            if weekly_check%7 == 1 and weekly_check != 1 and weekly_check != 0:
                weekly_check=0
                pass
            else:
                total_cost += f"+B{str(starting_row+1)}"

            weekly_check+=1
            starting_row+=1
            # ------------------------

            
            i+=1
        
        
        
        
        # Show total here -----------------
        

        i=1 
        
        #  We have to go in each of the attribute and store it 
        # print(list_of_columns)
        # for 
        total_cost = "0"
        
        # print(fb_total.index)
        for j in range(0,total_columns):
            if j == 0:
                cell_info = ws.cell(row=starting_row+1,column=j+1,value=str("Total"))
                set_border_and_red_bg(cell_info)
            value = fb_total[list_of_columns[j]][0]
            # print(value)
            
            # value = getattr(fb_ig_obj, list_of_columns[j])
            cell_info = ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")
            cell_info.number_format = list_of_columns_number_format[j]
            set_border_and_red_bg(cell_info)
        #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        # Formulas here 
        cell_info = ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR((D{str(starting_row+1)}/C{str(starting_row+1)}),0)")
        cell_info.number_format = '0.00%'
        set_border_and_red_bg(cell_info)
        cell_info = ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR((({total_cost}+B{str(starting_row+1)})/T2),0)")
        cell_info.number_format = '0.00%'
        set_border_and_red_bg(cell_info)
        
        CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(B{str(starting_row+1)}/D{str(starting_row+1)},0)")
        CPC.number_format = '$#,##0.00'
        set_border_and_red_bg(CPC)

        CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
        CPV.number_format = '$#,##0.00'
        set_border_and_red_bg(CPV)
        total_cost += f"+B{str(starting_row+1)}"
        starting_row+=1
# ---------------------------------
        
        
        # wb.save('summary_testing.xlsx')
        return wb,start_col,start_row

def media_calculation_assets(wb,ws,i,fb_ig_obj,fb,fb_weekly,fb_total,start_row,start_col,product_name,language_name,daily = False):
    #...................................................
    # Start working on the summary and show to LJ
    #...................................................

    # Loop for columns 
        # print(fb_total)
        if daily == True:
            list_of_columns = fb_ig_obj.ListOfColumnsDaily()
            list_of_columns_number_format=fb_ig_obj.format_cells_daily()
      
        else:
            list_of_columns = fb_ig_obj.ListOfColumnsForPlatformWeekly()
            
            list_of_columns_daily_record = fb_ig_obj.ListOfColumnsForPlatformWeekly()
            columns_daily = fb_ig_obj.ListOfColumnsForPlatform()
            list_of_columns_number_format = fb_ig_obj.Format_for_ListOfColumnsForPlatform()
            list_of_columns_summary_number_format = fb_ig_obj.Format_for_ListOfColumnsForPlatformSummary()

        # print(len(list_of_columns_number_format))
        # List of Columns of the class
        list_of_calculated_columns = ["CTR","Spent Budget %", "CPC","CPV"]
        # List of Columns Which has formulas
        # ws.cell(row=i+1, column=total_columns+1, value=f"=A{str(i+1)}/T2")

        total_columns = len(list_of_columns)
        
        # print(list_of_columns)
        # 9  - list of total columns 
        # Total Columns length
        total_columns_calculated =len(list_of_calculated_columns)
        # 4 - columns 
        # Total Columns calculated length 

        i = 1
        

        
        # counter= 0 
        # for j in range(total_columns+1,total_columns+total_columns_calculated+1):
        #     cell_info = ws.cell(row=1, column=start_col+j+1, value=list_of_calculated_columns[counter])
        #     set_border_and_align_weekly(cell_info)
        #     counter+=1
        # We are setting the column border and alignment for other columns that are needed 
        # i=1 
        # starting_row = 1
        # total_cost = "0"
        # for index, row in fb_weekly.iterrows():
        #     # Looping through weekly record and adding value and formating the cell border and alignment 
        #     for j in range(0,total_columns):
        #         if j == 0:
        #             cell_info = ws.cell(row=starting_row+1,column=start_col+j+1,value=str(index))
        #             set_border_and_align(cell_info)
        #         # try:  
        #         # print(getattr(fb_ig_obj, list_of_columns[j]))
        #         # print(fb_weekly.columns)
        #         value = fb_weekly[list_of_columns[j]][i-1]
                    
                    
        #         # except:
        #         #     value = 0
        #         # value = getattr(fb_ig_obj, list_of_columns[j])

        #         cell_info = ws.cell(row=starting_row+1, column=start_col+j+2, value=f"=ROUND({value},2)")
        #         cell_info.number_format = list_of_columns_summary_number_format[j]
        #         set_border_and_align(cell_info)

        #     #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        #     # Formulas here 
        #     cell_info= ws.cell(row=starting_row+1,column=start_col+total_columns+2,value=f"=IFERROR((D{str(i+1)}/C{str(i+1)}),0)")
        #     cell_info.number_format = '0.00%'
        #     set_border_and_align(cell_info)
        #     cell_info = ws.cell(row=starting_row+1,column=start_col+total_columns+3,value=f"=IFERROR((({total_cost}+B{str(i+1)})/T2),0)")
        #     cell_info.number_format = '0.00%'
        #     set_border_and_align(cell_info)
        #     CPC = ws.cell(row=starting_row+1,column=start_col+total_columns+4,value=f"=IFERROR(ROUND(B{str(i+1)}/D{str(i+1)},2),0)")
        #     CPC.number_format = '$#,##0.00'
        #     set_border_and_align(CPC)
        #     set_border_and_align(CPC)
        #     CPV = ws.cell(row=starting_row+1,column=start_col+total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
        #     CPV.number_format = '$#,##0.00'
        #     set_border_and_align(CPV)
        #     set_border_and_align(CPV)
        #     total_cost += f"+B{str(starting_row+1)}"
        #     starting_row+=1
        #     i+=1
        #     # ------------------------

# Show total here -----------------
        

        i=1 
        starting_row= start_row
        #  We have to go in each of the attribute and store it 
        # print(list_of_columns)
        # for 
        total_cost = "0"
        
        
        starting_row+=1
# ---------------------------------


        i = starting_row+2
        # Column printing for daily table 
        for j in range(0,total_columns):
            if j==0:
                
                cell_info = ws.cell(row=i-2,column=start_col+j+1,value=f"{language_name}")
                set_border_and_align_weekly(cell_info)

                ws.merge_cells(start_column=start_col+j+1, start_row=i-1, end_column = start_col+j+13, end_row=i-1)
                cell_info = ws.cell(row=i-1,column=start_col+j+1,value=f"{product_name}")
                set_border_and_align_weekly(cell_info)
                cell_info = ws.cell(row=i,column=start_col+j+1,value="date")
                set_border_and_align_weekly(cell_info)

            cell_info = ws.cell(row=i, column=start_col+j+2, value=list_of_columns_daily_record[j])
            set_border_and_align_weekly(cell_info)
        
        # ------------------------------------
        counter= 0 
        
        # Formulated columns printing for daily table
        for j in range(total_columns+1,total_columns+total_columns_calculated+1):
            cell_info = ws.cell(row=i, column=start_col+j+1, value=list_of_calculated_columns[counter])
            set_border_and_align_weekly(cell_info)
            counter+=1

        # Daily table printing-------
        i = 1
        start_row = starting_row + 2
        starting_row = starting_row+2
        total_cost = "0"

        weekly_check = 1
        for index, row in fb.iterrows():
            # print(index," rows: ",row)
            
            for j in range(0,total_columns):
                if j == 0:
                    value_date = fb[columns_daily[0]][i-1]
                    cell_info = ws.cell(row=starting_row+1,column=start_col+j+1,value=str(value_date))
                    set_border_and_align(cell_info)
                 
                value = fb[list_of_columns_daily_record[j]][i-1]
                    
              
                # value = getattr(fb_ig_obj, list_of_columns[j])
                cell_info = ws.cell(row=starting_row+1, column=start_col+j+2, value=f"=ROUND({value},2)")
                cell_info.number_format = list_of_columns_number_format[j]
                set_border_and_align(cell_info)

            # list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
            # Formulas here 
            column_budget = get_column_letter(2+start_col)
            column_click = get_column_letter(4+start_col)
            column_impression = get_column_letter(3+start_col)
            column_view = get_column_letter(5+start_col)

            cell_info = ws.cell(row=starting_row+1,column=start_col+total_columns+2,value=f"=IFERROR(({column_click}{str(starting_row+1)}/{column_impression}{str(starting_row+1)}),0)")
            cell_info.number_format = '0.00%'
            set_border_and_align(cell_info)
            if weekly_check%7 == 1 and weekly_check != 1 and weekly_check != 0:
                cell_info = ws.cell(row=starting_row+1,column=start_col+total_columns+3,value=f"=IFERROR((({total_cost})/T2),0)")
            else:
                cell_info = ws.cell(row=starting_row+1,column=start_col+total_columns+3,value=f"=IFERROR((({total_cost}+{column_budget}{str(starting_row+1)})/T2),0)")
            cell_info.number_format = '0.00%'
            set_border_and_align(cell_info)
            
            

            CPC = ws.cell(row=starting_row+1,column=start_col+total_columns+4,value=f"=IFERROR(ROUND({column_budget}{str(starting_row+1)}/{column_click}{str(starting_row+1)},2),0)")
            CPC.number_format = '$#,##0.00'
            set_border_and_align(CPC)
            CPV = ws.cell(row=starting_row+1,column=start_col+total_columns+5,value=f"=IFERROR(ROUND({column_budget}{str(starting_row+1)}/{column_view}{str(starting_row+1)},2),0)")
            CPV.number_format = '$#,##0.00'
            set_border_and_align(CPV)
            if weekly_check%7 == 1 and weekly_check != 1 and weekly_check != 0:
                weekly_check=0
                pass
            else:
                total_cost += f"+{column_budget}{str(starting_row+1)}"

            weekly_check+=1
            starting_row+=1
            # ------------------------

            
            i+=1
        
        
        
        
        # Show total here -----------------
        

        i=1 
        
        #  We have to go in each of the attribute and store it 
        # print(list_of_columns)
        # for 
        total_cost = "0"
        
        # print(fb_total.index)
        for j in range(0,total_columns):
            if j == 0:
                cell_info = ws.cell(row=starting_row+1,column=start_col+j+1,value=str("Total"))
                set_border_and_red_bg(cell_info)
            value = fb_total[list_of_columns[j]][0]
            # print(value)
            
            # value = getattr(fb_ig_obj, list_of_columns[j])
            cell_info = ws.cell(row=starting_row+1, column=start_col+j+2, value=f"=ROUND({value},2)")
            cell_info.number_format = list_of_columns_number_format[j]
            set_border_and_red_bg(cell_info)
        #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        # Formulas here 
        cell_info = ws.cell(row=starting_row+1,column=start_col+total_columns+2,value=f"=IFERROR(({column_click}{str(starting_row+1)}/{column_impression}{str(starting_row+1)}),0)")
        cell_info.number_format = '0.00%'
        set_border_and_red_bg(cell_info)
        cell_info = ws.cell(row=starting_row+1,column=start_col+total_columns+3,value=f"=IFERROR((({total_cost}+{column_budget}{str(starting_row+1)})/T2),0)")
        cell_info.number_format = '0.00%'
        set_border_and_red_bg(cell_info)
        
        CPC = ws.cell(row=starting_row+1,column=start_col+total_columns+4,value=f"=IFERROR({column_budget}{str(starting_row+1)}/{column_click}{str(starting_row+1)},0)")
        CPC.number_format = '$#,##0.00'
        set_border_and_red_bg(CPC)

        CPV = ws.cell(row=starting_row+1,column=start_col+total_columns+5,value=f"=IFERROR(ROUND({column_budget}{str(starting_row+1)}/{column_view}{str(starting_row+1)},2),0)")
        CPV.number_format = '$#,##0.00'
        set_border_and_red_bg(CPV)
        total_cost += f"+B{str(starting_row+1)}"
        starting_row+=1
# ---------------------------------
        
        
        # wb.save('summary_testing.xlsx')
        start_col+=13
        return wb,start_col


# This is for assets

def media_calculation_asset(image_size, wb,ws,start_row,start_col,fb_obj,fb,fb_weekly,fb_total,daily = False):
    #...................................................
    # Start working on the summary and show to LJ
    #...................................................
        # print("Start Column:",start_col)
    # Loop for columns 
        # print(fb_total)
        if daily == True:
            list_of_columns = fb_obj.ListOfColumnsDaily()
            list_of_columns_number_format=fb_obj.format_cells_daily()
        else:
            list_of_columns = fb_obj.ListOfColumnsAssets()
            list_of_columns_number_format = fb_obj.format_cells()
        # List of Columns of the class
        list_of_calculated_columns = ["CTR","Spent Budget %", "CPC"]
        # List of Columns Which has formulas
        # ws.cell(row=i+1, column=total_columns+1, value=f"=A{str(i+1)}/T2")

        total_columns = len(list_of_columns)
        # Total Columns length
        # print(total_columns)
        total_columns_calculated =len(list_of_calculated_columns)
        # Total Columns calculated length 
        ws.merge_cells(start_column=start_col+1, start_row=start_row-1, end_column = start_col + total_columns+total_columns_calculated + 1, end_row=start_row-1)
        ws.cell(row=start_row-1,column = start_col+1).value = image_size

        i = start_row
        j = start_col
        for x in range(0,total_columns):
            if x==0:
                cell_info = ws.cell(row=i,column=j+1,value="date")
             
                set_border_and_align_weekly(cell_info)

            cell_info = ws.cell(row=i, column=j+2, value=list_of_columns[x])
            
            set_border_and_align_weekly(cell_info)
            j+=1
        # print(j)
        # cell_info = ws.cell(row=i,column=20,value="Total Budget")
        # set_border_and_align(cell_info)
        # # Loop to Print the columns label with index

        counter= 0 
        for j in range(j+1,j+total_columns_calculated+1):
            cell_info = ws.cell(row=i, column=j+1, value=list_of_calculated_columns[counter])
            set_border_and_align_weekly(cell_info)
            counter+=1

        # i=1 
        # starting_row = 1
        # total_cost = "0"
        # for index, row in fb_weekly.iterrows():

        #     for j in range(0,total_columns):
        #         if j == 0:
        #             cell_info = ws.cell(row=starting_row+1,column=j+1,value=str(index))
        #             set_border_and_align(cell_info)
        #         try:  
        #             value = fb_weekly[getattr(fb_obj, list_of_columns[j])][i-1]
                    
                    
        #         except:
        #             value = 0
        #         # value = getattr(fb_obj, list_of_columns[j])

        #         cell_info = ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")
        #         cell_info.number_format = list_of_columns_number_format[j]
        #         set_border_and_align(cell_info)

        #     #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        #     # Formulas here 
        #     cell_info= ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR((D{str(i+1)}/C{str(i+1)}),0)")
        #     cell_info.number_format = '0.00%'
        #     set_border_and_align(cell_info)
        #     cell_info = ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR((({total_cost}+B{str(i+1)})/T2),0)")
        #     cell_info.number_format = '0.00%'
        #     set_border_and_align(cell_info)
        #     CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(ROUND(B{str(i+1)}/D{str(i+1)},2),0)")
        #     CPC.number_format = '$#,##0.00'
        #     set_border_and_align(CPC)
        #     set_border_and_align(CPC)
        #     CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
        #     CPV.number_format = '$#,##0.00'
        #     set_border_and_align(CPV)
        #     set_border_and_align(CPV)
        #     total_cost += f"+B{str(starting_row+1)}"
        #     starting_row+=1
        #     i+=1
        #     # ------------------------

        # # Show total here -----------------
        

        # i=1 
        
        # #  We have to go in each of the attribute and store it 
        # # print(list_of_columns)
        # # for 
        # total_cost = "0"
        
        # # print(fb_total.index)
        # for j in range(0,total_columns):
        #     if j == 0:
        #         cell_info = ws.cell(row=starting_row+1,column=j+1,value=str("Total"))
        #         set_border_and_red_bg(cell_info)
        #     try:  
        #         value = fb_total[getattr(fb_obj, list_of_columns[j])]
        #     except:
        #         value = 0
        #     # value = getattr(fb_obj, list_of_columns[j])
        #     cell_info = ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")
        #     cell_info.number_format = list_of_columns_number_format[j]
        #     set_border_and_red_bg(cell_info)
        # #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        # # Formulas here 
        # cell_info = ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR((D{str(starting_row+1)}/C{str(starting_row+1)}),0)")
        # cell_info.number_format = '0.00%'
        # set_border_and_red_bg(cell_info)
        # cell_info = ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR((({total_cost}+B{str(starting_row+1)})/T2),0)")
        # cell_info.number_format = '0.00%'
        # set_border_and_red_bg(cell_info)
        
        # CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(B{str(starting_row+1)}/D{str(starting_row+1)},0)")
        # CPC.number_format = '$#,##0.00'
        # set_border_and_red_bg(CPC)

        # CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
        # CPV.number_format = '$#,##0.00'
        # set_border_and_red_bg(CPV)
        # total_cost += f"+B{str(starting_row+1)}"
        # starting_row+=1
        # # ---------------------------------

        # i = starting_row+2
        # for j in range(0,total_columns):
        #     if j==0:
        #         cell_info = ws.cell(row=i,column=j+1,value="date")
        #         set_border_and_align_weekly(cell_info)

        #     cell_info = ws.cell(row=i, column=j+2, value=list_of_columns[j])
        #     set_border_and_align_weekly(cell_info)
        # counter= 0 
        # for j in range(total_columns+1,total_columns+total_columns_calculated+1):
        #     cell_info = ws.cell(row=i, column=j+1, value=list_of_calculated_columns[counter])
        #     set_border_and_align_weekly(cell_info)
        #     counter+=1
        
        i = 1
        starting_row = start_row
        total_cost = "0"

        for index, row in fb.iterrows():
            # print(index," rows: ",row)
            starting_col= start_col
            for x in range(0,total_columns):
                if x == 0:
                    cell_info = ws.cell(row=starting_row+1,column=starting_col+1,value=str(index))
                    set_border_and_align(cell_info)
                try:   
                    value = fb[getattr(fb_obj, list_of_columns[x])][i-1]
                    
                except:
                    value = 0
                # value = getattr(fb_obj, list_of_columns[j])
                
                cell_info = ws.cell(row=starting_row+1, column=starting_col+2, value=f"=ROUND({value},2)")
                cell_info.number_format = list_of_columns_number_format[x]
                set_border_and_align(cell_info)
                starting_col+=1



            # list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
            # Formulas here 
            # We have to set D and C dynamic so we can calculate them dynamically without adding in manual....
            # Position of Impression and clicks are 3  , 4
            
            



            cell_info = ws.cell(row = starting_row +1 , column = starting_col)
            clicks_coordinate = cell_info.coordinate

            cell_info = ws.cell(row=starting_row + 1, column = starting_col-1)
            impressions_coordinate = cell_info.coordinate

            cell_info = ws.cell(row=starting_row+1, column = starting_col - 2)
            budget_coordinate = cell_info.coordinate

            cell_info = ws.cell(row=starting_row+1,column=starting_col+2,value=f"=IFERROR(({clicks_coordinate}/{impressions_coordinate}),0)")
            cell_info.number_format = '0.00%'
            set_border_and_align(cell_info)

            cell_info = ws.cell(row=starting_row+1,column=starting_col+3,value=f"=IFERROR((({total_cost}+{budget_coordinate})/T2),0)")
            cell_info.number_format = '0.00%'
            set_border_and_align(cell_info)

            CPC = ws.cell(row=starting_row+1,column=starting_col+4,value=f"=IFERROR(ROUND({budget_coordinate}/{clicks_coordinate},2),0)")
            CPC.number_format = '$#,##0.00'
            set_border_and_align(CPC)
            # CPV = ws.cell(row=starting_row+1,column=starting_col+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
            # CPV.number_format = '$#,##0.00'
            # set_border_and_align(CPV)

            total_cost += f"+{budget_coordinate}"
   
            starting_row+=1
            # ------------------------

            
            i+=1
        
        
        
        
        
        
        wb.save('summary_testing.xlsx')
        j+=2
        return j

# This is for assets

def media_calculation_message(message_,image_size, wb,ws,start_row,start_col,fb_obj,fb,fb_weekly,fb_total,daily = False):
    #...................................................
    # Start working on the summary and show to LJ
    #...................................................
        # print("Start Column:",start_col)
    # Loop for columns 
        # print(fb_total)
        if daily == True:
            list_of_columns = fb_obj.ListOfColumnsDaily()
            list_of_columns_number_format=fb_obj.format_cells_daily()
        else:
            list_of_columns = fb_obj.ListOfColumnsAssets()
            list_of_columns_number_format = fb_obj.format_cells()
        # List of Columns of the class
        list_of_calculated_columns = ["CTR","Spent Budget %", "CPC"]
        # List of Columns Which has formulas
        # ws.cell(row=i+1, column=total_columns+1, value=f"=A{str(i+1)}/T2")

        total_columns = len(list_of_columns)
        # Total Columns length
        # print(total_columns)
        total_columns_calculated =len(list_of_calculated_columns)
        # Total Columns calculated length 

        ws.merge_cells(start_column=start_col+1, start_row=start_row-2, end_column = start_col + total_columns+total_columns_calculated + 1, end_row=start_row-2)
        ws.cell(row=start_row-2,column = start_col+1).value = message_


        ws.merge_cells(start_column=start_col+1, start_row=start_row-1, end_column = start_col + total_columns+total_columns_calculated + 1, end_row=start_row-1)
        ws.cell(row=start_row-1,column = start_col+1).value = image_size

        i = start_row
        j = start_col
        for x in range(0,total_columns):
            if x==0:
                cell_info = ws.cell(row=i,column=j+1,value="date")
             
                set_border_and_align_weekly(cell_info)

            cell_info = ws.cell(row=i, column=j+2, value=list_of_columns[x])
            
            set_border_and_align_weekly(cell_info)
            j+=1
        # print(j)
        # cell_info = ws.cell(row=i,column=20,value="Total Budget")
        # set_border_and_align(cell_info)
        # # Loop to Print the columns label with index

        counter= 0 
        for j in range(j+1,j+total_columns_calculated+1):
            cell_info = ws.cell(row=i, column=j+1, value=list_of_calculated_columns[counter])
            set_border_and_align_weekly(cell_info)
            counter+=1

        # i=1 
        # starting_row = 1
        # total_cost = "0"
        # for index, row in fb_weekly.iterrows():

        #     for j in range(0,total_columns):
        #         if j == 0:
        #             cell_info = ws.cell(row=starting_row+1,column=j+1,value=str(index))
        #             set_border_and_align(cell_info)
        #         try:  
        #             value = fb_weekly[getattr(fb_obj, list_of_columns[j])][i-1]
                    
                    
        #         except:
        #             value = 0
        #         # value = getattr(fb_obj, list_of_columns[j])

        #         cell_info = ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")
        #         cell_info.number_format = list_of_columns_number_format[j]
        #         set_border_and_align(cell_info)

        #     #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        #     # Formulas here 
        #     cell_info= ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR((D{str(i+1)}/C{str(i+1)}),0)")
        #     cell_info.number_format = '0.00%'
        #     set_border_and_align(cell_info)
        #     cell_info = ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR((({total_cost}+B{str(i+1)})/T2),0)")
        #     cell_info.number_format = '0.00%'
        #     set_border_and_align(cell_info)
        #     CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(ROUND(B{str(i+1)}/D{str(i+1)},2),0)")
        #     CPC.number_format = '$#,##0.00'
        #     set_border_and_align(CPC)
        #     set_border_and_align(CPC)
        #     CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
        #     CPV.number_format = '$#,##0.00'
        #     set_border_and_align(CPV)
        #     set_border_and_align(CPV)
        #     total_cost += f"+B{str(starting_row+1)}"
        #     starting_row+=1
        #     i+=1
        #     # ------------------------

        # # Show total here -----------------
        

        # i=1 
        
        # #  We have to go in each of the attribute and store it 
        # # print(list_of_columns)
        # # for 
        # total_cost = "0"
        
        # # print(fb_total.index)
        # for j in range(0,total_columns):
        #     if j == 0:
        #         cell_info = ws.cell(row=starting_row+1,column=j+1,value=str("Total"))
        #         set_border_and_red_bg(cell_info)
        #     try:  
        #         value = fb_total[getattr(fb_obj, list_of_columns[j])]
        #     except:
        #         value = 0
        #     # value = getattr(fb_obj, list_of_columns[j])
        #     cell_info = ws.cell(row=starting_row+1, column=j+2, value=f"=ROUND({value},2)")
        #     cell_info.number_format = list_of_columns_number_format[j]
        #     set_border_and_red_bg(cell_info)
        # #list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
        # # Formulas here 
        # cell_info = ws.cell(row=starting_row+1,column=total_columns+2,value=f"=IFERROR((D{str(starting_row+1)}/C{str(starting_row+1)}),0)")
        # cell_info.number_format = '0.00%'
        # set_border_and_red_bg(cell_info)
        # cell_info = ws.cell(row=starting_row+1,column=total_columns+3,value=f"=IFERROR((({total_cost}+B{str(starting_row+1)})/T2),0)")
        # cell_info.number_format = '0.00%'
        # set_border_and_red_bg(cell_info)
        
        # CPC = ws.cell(row=starting_row+1,column=total_columns+4,value=f"=IFERROR(B{str(starting_row+1)}/D{str(starting_row+1)},0)")
        # CPC.number_format = '$#,##0.00'
        # set_border_and_red_bg(CPC)

        # CPV = ws.cell(row=starting_row+1,column=total_columns+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
        # CPV.number_format = '$#,##0.00'
        # set_border_and_red_bg(CPV)
        # total_cost += f"+B{str(starting_row+1)}"
        # starting_row+=1
        # # ---------------------------------

        # i = starting_row+2
        # for j in range(0,total_columns):
        #     if j==0:
        #         cell_info = ws.cell(row=i,column=j+1,value="date")
        #         set_border_and_align_weekly(cell_info)

        #     cell_info = ws.cell(row=i, column=j+2, value=list_of_columns[j])
        #     set_border_and_align_weekly(cell_info)
        # counter= 0 
        # for j in range(total_columns+1,total_columns+total_columns_calculated+1):
        #     cell_info = ws.cell(row=i, column=j+1, value=list_of_calculated_columns[counter])
        #     set_border_and_align_weekly(cell_info)
        #     counter+=1
        
        i = 1
        starting_row = start_row
        total_cost = "0"

        for index, row in fb.iterrows():
            # print(index," rows: ",row)
            starting_col= start_col
            for x in range(0,total_columns):
                if x == 0:
                    cell_info = ws.cell(row=starting_row+1,column=starting_col+1,value=str(index))
                    set_border_and_align(cell_info)
                try:   
                    value = fb[getattr(fb_obj, list_of_columns[x])][i-1]
                    
                except:
                    value = 0
                # value = getattr(fb_obj, list_of_columns[j])
                
                cell_info = ws.cell(row=starting_row+1, column=starting_col+2, value=f"=ROUND({value},2)")
                cell_info.number_format = list_of_columns_number_format[x]
                set_border_and_align(cell_info)
                starting_col+=1



            # list_of_formulas=[f"=ROUND(D{str(i+1)}/C{str(i+1)},2)",f"=ROUND({total_cost}+B{str(i+1)}/T2,2)",f"=ROUND(B{str(i+1)}/D{str(i+1)},2)"]
            # Formulas here 
            # We have to set D and C dynamic so we can calculate them dynamically without adding in manual....
            # Position of Impression and clicks are 3  , 4
            
            



            cell_info = ws.cell(row = starting_row +1 , column = starting_col)
            clicks_coordinate = cell_info.coordinate

            cell_info = ws.cell(row=starting_row + 1, column = starting_col-1)
            impressions_coordinate = cell_info.coordinate

            cell_info = ws.cell(row=starting_row+1, column = starting_col - 2)
            budget_coordinate = cell_info.coordinate

            cell_info = ws.cell(row=starting_row+1,column=starting_col+2,value=f"=IFERROR(({clicks_coordinate}/{impressions_coordinate}),0)")
            cell_info.number_format = '0.00%'
            set_border_and_align(cell_info)

            cell_info = ws.cell(row=starting_row+1,column=starting_col+3,value=f"=IFERROR((({total_cost}+{budget_coordinate})/T2),0)")
            cell_info.number_format = '0.00%'
            set_border_and_align(cell_info)

            CPC = ws.cell(row=starting_row+1,column=starting_col+4,value=f"=IFERROR(ROUND({budget_coordinate}/{clicks_coordinate},2),0)")
            CPC.number_format = '$#,##0.00'
            set_border_and_align(CPC)
            # CPV = ws.cell(row=starting_row+1,column=starting_col+5,value=f"=IFERROR(ROUND(B{str(starting_row+1)}/E{str(starting_row+1)},2),0)")
            # CPV.number_format = '$#,##0.00'
            # set_border_and_align(CPV)

            total_cost += f"+{budget_coordinate}"
   
            starting_row+=1
            # ------------------------

            
            i+=1
        
        
        
        
        
        
        wb.save(pickle_file_path)
        j+=2
        return j



def set_columns(df):
    df.rename(columns={'Spend /Revenue ': 'Amount spent'}, inplace=True)
    df.rename(columns={'Platform ': 'platform'}, inplace=True)
    df.rename(columns={'Plaform ': 'platform'}, inplace=True)
    df.rename(columns={'Market ': "country"},inplace=True)
    df.rename(columns={'Tget Mket ': "country"},inplace=True)
    return df

# Backup 
# @app.route("/sub",methods=["POST"])
# def submit():
#     # try:
#         try:
#             facebook = request.files["Facebook"]
            
#         except:
#             print("Facebook skipped---")
#         # facebook_budget = int(request.form["Facebook_Budget"])
#         try:
#             instagram = request.files["Instagram"]
#         except:
#             print("instagram skipped---")
#         try:
#             twitter = request.files["Twitter"]
#         except:
#             print("twitter skipped---")

#         try:
#             snapchat = request.files["Snapchat"]
#         except:
#             print("snapchat skipped---")

#         try:
#             dv360 = request.files["DV360"]
#         except:
#             print("dv360 skipped---")
#         try:
#             teads = request.files["Teads"]
#         except:
#             print("teads skipped---")

#         try:
#             hph_raw = request.files["Hph"]
#         except:
#             print("hph skipped---")

#         try:
#             youtube_raw = request.files["Youtube"]
#         except:
#             print("youtube skipped---")
#         # Reading files and taking budget as input 
        
        
#         try:
#             engine.execute('USE media_campaign_dashboard;')

#             # Disable foreign key checks
#             engine.execute('SET FOREIGN_KEY_CHECKS = 0;')

#             # Generate the drop table statements
#             drop_query = '''
#                 SELECT CONCAT('DROP TABLE IF EXISTS ', table_name, ';')
#                 FROM information_schema.tables
#                 WHERE table_schema = 'media_campaign_dashboard';
#                 '''
#             result = engine.execute(drop_query)

#             # Fetch all the drop statements and execute them one by one
#             drop_statements = [row[0] for row in result]
#             for statement in drop_statements:
#                 engine.execute(statement)
#         except:
#             print("No Table exist")
#         # Reading files and storing dataframe of Different Medias 
#         try: 
#             fb = pd.read_excel(facebook) 
#             fb = set_columns(fb)
#             # Fb Assets
#             # fb_asset = pd.read_excel(facebook)
#             # fb_size = pd.read_excel(facebook)
#             # fb_language = pd.read_excel(facebook)
#             # fb_country = pd.read_excel(facebook)
#             # fb_message = pd.read_excel(facebook)
#         except: 
#             pass
#         try:
#             fb.to_sql("facebook", con=engine, if_exists='replace', index=True)
#         except:
#             pass

#         # Conditions = [repr(Date)]
#         # print("-->>>>",Conditions)

#         # select_columns = ', '.join(Columns)
#         # select_tables = ', '.join(Table)
#         # select_where = "WHERE " + ' AND '.join(Where)

#         # print(select_columns)
        
#         # query = f"""
#         #     SELECT {select_columns}  
#         #     FROM {select_tables} 
#         #     WHERE {select_where}
#         #     GROUP BY `Date`;
#         # """
#         # result = engine.execute(query)

#         # # Fetch and print the query results
#         # rows = []
        
#         # for row in result:
#         #     rows.append(row)
        

        
#         # Sum_of_All_Columns_With_Language = pd.DataFrame(rows)
        
#         # print(Sum_of_All_Columns_With_Language)

        
#         # query = f"""
#         #     SELECT `Date`, SUM(`Impression`) AS `Impression`, SUM(`Link clicks`) AS `Link clicks`, SUM(`Amount Spent`) AS `Amount Spent`, 
#         #     SUM(`Conversion`) AS `Conversion`, SUM(`View (Video)`) AS `View (Video)`, SUM(`25%%e Views (Video)`) AS `25%%e Views (Video)`, SUM(`50%% View (Video)`) AS `50%% View (Video)`, 
#         #     SUM(`75%% View (Video)`) AS `75%% View (Video)`, SUM(`Complete Views (Video)`) AS `Complete Views (Video)` 
#         #     FROM `facebook` 
#         #     WHERE `Language` = 1
#         #     GROUP BY `Date`;
#         # """


#         # Sum of all specific columns (Impressions, Clicks, Amount Spent, Conversion, View (Video), 25%e Views (Video), 50% View (Video), 75% View (Video), Complete Views (Video)) with respect to Date and Ad name
        
        
        
        
        
        
#         # Sum of all specific columns (Impressions, Clicks, Amount Spent, Conversion, View (Video), 25%e Views (Video), 50% View (Video), 75% View (Video), Complete Views (Video)) with respect to Date and Size
#         # Sum of all specific columns (Impressions, Clicks, Amount Spent, Conversion, View (Video), 25%e Views (Video), 50% View (Video), 75% View (Video), Complete Views (Video)) with respect to Date and Country
#         # Sum of all specific columns (Impressions, Clicks, Amount Spent, Conversion, View (Video), 25%e Views (Video), 50% View (Video), 75% View (Video), Complete Views (Video)) with respect to Date and Asset type
#         # try:
#         #     ig = pd.read_excel(instagram)
#         # except:
#         #     pass
        
#         try: 
#             ig = pd.read_excel(instagram) 
#             ig = set_columns(ig)
#         except: 
#             pass

#         try:
#             ig.to_sql("instagram", con=engine, if_exists='replace', index=True)
#         except:
#             pass

    

#         try: 
#             tw = pd.read_excel(twitter) 
#             tw = set_columns(tw)
#         except: 
#             pass
#         try:
#             tw.to_sql("twitter", con=engine, if_exists='replace', index=True)
#         except:
#             pass


      
        
#         try: 
#             sc = pd.read_excel(snapchat) 
#             sc = set_columns(sc)

#         except: 
#             pass

#         try:
#             sc.to_sql("snapchat", con=engine, if_exists='replace', index=True)
#         except:
#             pass


#         try:    
#             dv = pd.read_excel(dv360)
#             dv = set_columns(dv)

#             dv_asset = pd.read_excel(dv360)
#         except:
#             pass
#         try:
#             dv.to_sql("dv360", con=engine, if_exists='replace', index=True)
#         except:
#             pass

#         try:
#             td = pd.read_excel(teads)
#             td = set_columns(td)
#         except:
#             pass

#         try:
#             td.to_sql("teads", con=engine, if_exists='replace', index=True)
#         except:
#             pass

#         try:
#             hph = pd.read_excel(hph_raw)
#             hph = set_columns(hph)
#         except:
#             pass

#         try:
#             hph.to_sql("hph", con=engine, if_exists='replace', index=True)
#         except:
#             pass
#         try:
#             youtube = pd.read_excel(youtube_raw)
#             youtube = set_columns(youtube)
#         except:
#             pass

#         try:
#             youtube.to_sql("youtube", con=engine, if_exists='replace', index=True)
#         except:
#             pass
        
#         # Storing different medias in the variables. Its the dataframe 

#         try:
#             fb_obj = MatchingColumn(fb)
#             # fb_asset_obj = MatchingColumnAsset(fb_asset)
#             # fb_size_obj = MatchingColumnAsset(fb_size)
#             # fb_language_obj = MatchingColumnAsset(fb_language)
#             # fb_country_obj = MatchingColumnAsset(fb_country)
#             # fb_message_obj = MatchingColumnAsset(fb_message)
#         except:
#             pass
#         try:
#             ig_obj = MatchingColumn(ig)
#         except:
#             pass
        
#         try:
#             tw_obj = MatchingColumn(tw)
#         except:
#             pass
#         try:
#             sc_obj = MatchingColumn(sc)
#         except:
#             pass
#         try:
#             dv_obj = MatchingColumn(dv)
#             # dv_asset_obj = MatchingColumnAsset(dv_asset)
#         except:
#             pass
#         try:
#             td_obj = MatchingColumn(td)
#         except:
#             pass
#         try:
#             hph_obj = MatchingColumn(hph)
#         except:
#             pass
#         try:
#             youtube_obj = MatchingColumn(youtube)
#         except:
#             pass
#         # Fetching matching columns of the dataframe and storing it  

#         # So now we have the files and the column names in the objects we can use them to fetch data
        
#         # First we have to convert the dates in the data 
#         # print(td[td_obj.date])



#         try:
#             fb[fb_obj.date] = pd.to_datetime(fb[fb_obj.date])
#             rawData = fb
#             fb.sort_values(by=fb_obj.date, inplace=True)
#             fb[fb_obj.date] = fb[fb_obj.date].dt.strftime('%Y-%m-%d')

#             rawData[fb_obj.date] = pd.to_datetime(fb[fb_obj.date])
#         except:
#             print("Fb Dates issues")

#         # try:    
#         #     fb_asset[fb_asset_obj.date] = pd.to_datetime(fb_asset[fb_asset_obj.date])

#         #     fb_asset.sort_values(by=fb_asset_obj.date, inplace=True)
#         #     fb_asset[fb_asset_obj.date] = fb_asset[fb_asset_obj.date].dt.strftime('%Y-%m-%d')
#         # except:
#         #     print("Facebook Assets dates issues")

#         # try:    
#         #     fb_size[fb_size_obj.date] = pd.to_datetime(fb_size[fb_size_obj.date])

#         #     fb_size.sort_values(by=fb_size_obj.date, inplace=True)
#         #     fb_size[fb_size_obj.date] = fb_size[fb_size_obj.date].dt.strftime('%Y-%m-%d')
#         # except:
#         #     print("Facebook Size dates issues")

#         # try:    
#         #     fb_language[fb_language_obj.date] = pd.to_datetime(fb_language[fb_language_obj.date])

#         #     fb_language.sort_values(by=fb_language_obj.date, inplace=True)
#         #     fb_language[fb_language_obj.date] = fb_language[fb_language_obj.date].dt.strftime('%Y-%m-%d')
#         # except:
#         #     print("Facebook Language dates issues")

#         # try:    
#         #     fb_country[fb_country_obj.date] = pd.to_datetime(fb_country[fb_country_obj.date])

#         #     fb_country.sort_values(by=fb_country_obj.date, inplace=True)
#         #     fb_country[fb_country_obj.date] = fb_country[fb_country_obj.date].dt.strftime('%Y-%m-%d')
#         # except:
#         #     print("Facebook Country dates issues")

#         # try:    
#         #     fb_message[fb_message_obj.date] = pd.to_datetime(fb_message[fb_message_obj.date])

#         #     fb_message.sort_values(by=fb_message_obj.date, inplace=True)
#         #     fb_message[fb_message_obj.date] = fb_message[fb_message_obj.date].dt.strftime('%Y-%m-%d')
#         # except:
#         #     print("Facebook Message dates issues")
#         # Dont touch this ----------------------------------------

#         # try:
#         #     rawData.to_sql("facebook", con=engine, if_exists='replace', index=True)
        



        
#             # Queries to show result 
#             # Sum of all specific columns (Impressions, Clicks, Amount Spent, Conversion, View (Video), 25%e Views (Video), 50% View (Video), 75% View (Video), Complete Views (Video)) with respect to Date
#             # SELECT Date,SUM(`Impression`), SUM(`Link clicks`), SUM(`Amount Spent`), SUM(`Conversion`), SUM(`View (Video)`), SUM(`25%e Views (Video)`), SUM(`50% View (Video)`), SUM(`75% View (Video)`), SUM(`Complete Views (Video)`) FROM facebook GROUP BY (Date);
#             # Filter one is completed
#             # query = """
#             #     SELECT `Date`, SUM(`Impression`) AS `Impression`, SUM(`Link clicks`) AS `Link clicks`, SUM(`Amount Spent`) AS `Amount Spent`, 
#             #     SUM(`Conversion`) AS `Conversion`, SUM(`View (Video)`) AS `View (Video)`, SUM(`25%%e Views (Video)`) AS `25%%e Views (Video)`, SUM(`50%% View (Video)`) AS `50%% View (Video)`, 
#             #     SUM(`75%% View (Video)`) AS `75%% View (Video)`, SUM(`Complete Views (Video)`) AS `Complete Views (Video)` 
#             #     FROM `facebook` 
#             #     GROUP BY `Date`;
#             # """
#             # result = engine.execute(query)

#             # # Fetch and print the query results
#             # rows = []
            
#             # for row in result:
#             #     rows.append(row)
            
            
#             # Sum_of_All_Columns = pd.DataFrame(rows)
            
#             # print(Sum_of_All_Columns)




            
#             # Sum of all specific columns (Impressions, Clicks, Amount Spent, Conversion, View (Video), 25%e Views (Video), 50% View (Video), 75% View (Video), Complete Views (Video)) with respect to Date and Platform
#             # SELECT Date,SUM(`Impression`), SUM(`Link clicks`), SUM(`Amount Spent`), SUM(`Conversion`), SUM(`View (Video)`), SUM(`25%e Views (Video)`), SUM(`50% View (Video)`), SUM(`75% View (Video)`), SUM(`Complete Views (Video)`) FROM `facebook` WHERE `Platform` = "facebook" GROUP BY (Date);
#             # Filter one is completed
#             # query = """
#             #     SELECT `Date`, SUM(`Impression`) AS `Impression`, SUM(`Link clicks`) AS `Link clicks`, SUM(`Amount Spent`) AS `Amount Spent`, 
#             #     SUM(`Conversion`) AS `Conversion`, SUM(`View (Video)`) AS `View (Video)`, SUM(`25%%e Views (Video)`) AS `25%%e Views (Video)`, SUM(`50%% View (Video)`) AS `50%% View (Video)`, 
#             #     SUM(`75%% View (Video)`) AS `75%% View (Video)`, SUM(`Complete Views (Video)`) AS `Complete Views (Video)` 
#             #     FROM `facebook` 
#             #     WHERE `Platform` = "facebook"
#             #     GROUP BY `Date`;
#             # """
#             # result = engine.execute(query)

#             # # Fetch and print the query results
#             # rows = []
            
#             # for row in result:
#             #     rows.append(row)
            
            
#             # Sum_of_All_Columns_With_Platform = pd.DataFrame(rows)
            
#             # print(Sum_of_All_Columns_With_Platform)
            
            
            
            
            
#             # Sum of all specific columns (Impressions, Clicks, Amount Spent, Conversion, View (Video), 25%e Views (Video), 50% View (Video), 75% View (Video), Complete Views (Video)) with respect to Date and Objective
#             # SELECT Date,SUM(`Impression`), SUM(`Link clicks`), SUM(`Amount Spent`), SUM(`Conversion`), SUM(`View (Video)`), SUM(`25%e Views (Video)`), SUM(`50% View (Video)`), SUM(`75% View (Video)`), SUM(`Complete Views (Video)`) FROM `facebook` WHERE `objective` = "LINK_CLICKS" GROUP BY (Date);
#             # query = """
#             #     SELECT `Date`, SUM(`Impression`) AS `Impression`, SUM(`Link clicks`) AS `Link clicks`, SUM(`Amount Spent`) AS `Amount Spent`, 
#             #     SUM(`Conversion`) AS `Conversion`, SUM(`View (Video)`) AS `View (Video)`, SUM(`25%%e Views (Video)`) AS `25%%e Views (Video)`, SUM(`50%% View (Video)`) AS `50%% View (Video)`, 
#             #     SUM(`75%% View (Video)`) AS `75%% View (Video)`, SUM(`Complete Views (Video)`) AS `Complete Views (Video)` 
#             #     FROM `facebook` 
#             #     WHERE `objective` = "LINK_CLICKS"
#             #     GROUP BY `Date`;
#             # """
#             # result = engine.execute(query)

#             # # Fetch and print the query results
#             # rows = []
            
#             # for row in result:
#             #     rows.append(row)
            
            
#             # Sum_of_All_Columns_With_Objective = pd.DataFrame(rows)
            
#             # print(Sum_of_All_Columns_With_Objective)






#             # Sum of all specific columns (Impressions, Clicks, Amount Spent, Conversion, View (Video), 25%e Views (Video), 50% View (Video), 75% View (Video), Complete Views (Video)) with respect to Date and Campaign
#             # SELECT Date,SUM(`Impression`), SUM(`Link clicks`), SUM(`Amount Spent`), SUM(`Conversion`), SUM(`View (Video)`), SUM(`25%e Views (Video)`), SUM(`50% View (Video)`), SUM(`75% View (Video)`), SUM(`Complete Views (Video)`) FROM `facebook` WHERE `Campaign / Campaign name / Insertion Order` = "LINK_CLICKS" GROUP BY (Date);
#             # query = """
#             #     SELECT `Date`, SUM(`Impression`) AS `Impression`, SUM(`Link clicks`) AS `Link clicks`, SUM(`Amount Spent`) AS `Amount Spent`, 
#             #     SUM(`Conversion`) AS `Conversion`, SUM(`View (Video)`) AS `View (Video)`, SUM(`25%%e Views (Video)`) AS `25%%e Views (Video)`, SUM(`50%% View (Video)`) AS `50%% View (Video)`, 
#             #     SUM(`75%% View (Video)`) AS `75%% View (Video)`, SUM(`Complete Views (Video)`) AS `Complete Views (Video)` 
#             #     FROM `facebook` 
#             #     WHERE `Campaign / Campaign name   / Insertion Order` = "LGGULF_EMGF-MKT-20230303-0001_AS_rac-ramadan-campaign _2023-03-10_2023-04-14_FB_AE_Other Interests_Traffic"
#             #     GROUP BY `Date`;
#             # """
#             # result = engine.execute(query)

#             # # Fetch and print the query results
#             # rows = []
            
#             # for row in result:
#             #     rows.append(row)
            
            
#             # Sum_of_All_Columns_With_Campaign_name = pd.DataFrame(rows)
            
#             # print(Sum_of_All_Columns_With_Campaign_name)




#             # Sum of all specific columns (Impressions, Clicks, Amount Spent, Conversion, View (Video), 25%e Views (Video), 50% View (Video), 75% View (Video), Complete Views (Video)) with respect to Date and Language
#             # Columns = ['Date',"SUM(`Impression`) AS `Impression`", "SUM(`Link clicks`) AS `Link clicks`", "SUM(`Amount Spent`) AS `Amount Spent`", 
#             #     "SUM(`Conversion`) AS `Conversion`", "SUM(`View (Video)`) AS `View (Video)`", "SUM(`25%%e Views (Video)`) AS `25%% Views (Video)`", "SUM(`50%% View (Video)`) AS `50%% View (Video)`", 
#             #     "SUM(`75%% View (Video)`) AS `75%% View (Video)`", "SUM(`Complete Views (Video)`) AS `Complete Views (Video)`"]
#             # Table = ['facebook']
#             # Where = []
#         # except:
#         #     pass




#         try:
#             ig[ig_obj.date] = pd.to_datetime(ig[ig_obj.date])
#             ig.sort_values(by=ig_obj.date, inplace=True)
#             ig[ig_obj.date] = ig[ig_obj.date].dt.strftime('%Y-%m-%d')
#         except:
#             print("ig Dates issues")

#         try:

#             tw[tw_obj.date] = pd.to_datetime(tw[tw_obj.date])
#             tw.sort_values(by=tw_obj.date, inplace=True)
#             tw[tw_obj.date] = tw[tw_obj.date].dt.strftime('%Y-%m-%d')
#         except:
#             print("twitter dates issues ")

#         try:
#             sc[sc_obj.date] = pd.to_datetime(sc[sc_obj.date])
#             sc.sort_values(by=sc_obj.date, inplace=True)
#             sc[sc_obj.date] = sc[sc_obj.date].dt.strftime('%Y-%m-%d')
#         except:
#             print("Snapchat dates issues")

#         try:

#             td[td_obj.date] = pd.to_datetime(td[td_obj.date])
#             td.sort_values(by=td_obj.date, inplace=True)
#             td[td_obj.date] = td[td_obj.date].dt.strftime('%Y-%m-%d')
#         except:
#             print("Teads dates issues")

#         try:    
#             dv[dv_obj.date] = pd.to_datetime(dv[dv_obj.date])

#             dv.sort_values(by=dv_obj.date, inplace=True)
#             dv[dv_obj.date] = dv[dv_obj.date].dt.strftime('%Y-%m-%d')
#         except:
#             print("Dv360 dates issues")

#         # try:    
#         #     dv_asset[dv_asset_obj.date] = pd.to_datetime(dv_asset[dv_asset_obj.date])

#         #     dv_asset.sort_values(by=dv_asset_obj.date, inplace=True)
#         #     dv_asset[dv_asset_obj.date] = dv_asset[dv_asset_obj.date].dt.strftime('%Y-%m-%d')
#         # except:
#         #     print("Dv360 Assets dates issues")
        
#         try:    
#             hph[hph_obj.date] = pd.to_datetime(hph[hph_obj.date])

#             hph.sort_values(by=hph_obj.date, inplace=True)
#             hph[hph_obj.date] = hph[hph_obj.date].dt.strftime('%Y-%m-%d')
#         except:
#             print("Hph dates issues")

#         try:    
#             youtube[youtube_obj.date] = pd.to_datetime(youtube[youtube_obj.date])

#             youtube.sort_values(by=youtube_obj.date, inplace=True)
#             youtube[youtube_obj.date] = youtube[youtube_obj.date].dt.strftime('%Y-%m-%d')
#         except:
#             print("Youtube dates issues")
#         # Formatting dates and sorting the values according to dates 
        
        
#         my_dir = os.path.dirname(__file__)
#         pickle_file_path = os.path.join(my_dir, 'summary_testing.xlsx')
#         # Creating path for the final output file 


#         # with pd.ExcelWriter(pickle_file_path) as writer:
#         # list_of_summary_medias_with_data =[]
#         try:
#             fb = fb.groupby([fb_obj.date]).sum()
            

#         except:
#             print("fb sum issues")
        
#         # fb.to_sql("facebook", con=engine, if_exists='replace', index=True)
#         try: 
#             ig = ig.groupby([ig_obj.date]).sum()
#         except:
#             print("ig sum issues")
#         try: 
#             tw = tw.groupby([tw_obj.date]).sum()
#         except:
#             print("tw sum issues")
#         try: 
#             td = td.groupby([td_obj.date]).sum()
#         except:
#             print("td sum issues")
#         try: 
#             sc = sc.groupby([sc_obj.date]).sum()
#         except:
#             print("sc sum issues")
#         try: 
#             dv = dv.groupby([dv_obj.date]).sum()
#         except:
#             print("dv sum issues")
#         try: 
#             hph = hph.groupby([hph_obj.date]).sum()
#         except:
#             print("hph sum issues")
        
#         try: 
#             youtube = youtube.groupby([youtube_obj.date]).sum()
#         except:
#             print("youtube sum issues")

#         # ----------------------------------------------
#         try:
#             fb_weekly = CalculateWeekly(fb,fb_obj)
#         except:
#             print("fb_weekly issue")
#         try:
#             ig_weekly = CalculateWeekly(ig,ig_obj)
#         except:
#             print("ig_weekly issue")
#         try:
#             sc_weekly = CalculateWeekly(sc,sc_obj)
#         except:
#             print("sc_weekly issue")
#         try:
#             dv_weekly = CalculateWeekly(dv,dv_obj)
#         except:
#             print("dv_weekly issue")
        
#         try:
#             hph_weekly = CalculateWeekly(hph,hph_obj)
#         except:
#             print("hph_weekly issue")
#         try:
#             youtube_weekly = CalculateWeekly(youtube,youtube_obj)
#         except:
#             print("youtube_weekly issue")
#         # =================================================
#         try:
#             fb_total = fb_weekly.sum()
#         except:
#             print("fb_total issue")
#         try:
#             ig_total = ig_weekly.sum()
#         except:
#             print("ig_total issue")
#         try:
#             sc_total = sc_weekly.sum()
#         except:
#             print("sc_total issue")
#         try:
#             dv_total = dv_weekly.sum()
#         except:
#             print("dv_total issue")
#         try:
#             hph_total = hph_weekly.sum()
#         except:
#             print("hph_total issue")
        
#         try:
#             youtube_total = youtube_weekly.sum()
#         except:
#             print("youtube_total issue")
        
#         # Summing the data according to the date so we have one data for each data
        
#         # Until here everything is fine 
#         # Instead of setting columns using dataframe we will set them in excel anad update the excel sheet according to row and column 
#         wb = Workbook()
#         wb.remove(wb.active)
#         try:
#             ws = wb.create_sheet('Facebook')
#             wb.active = ws
#             media_calculation(wb,ws,1,fb_obj,fb,fb_weekly,fb_total)
#         except:
#             wb.remove(wb.active)
#             print("Facebook sheet skipped----")

#         try:
#             ws = wb.create_sheet('Instagram')
#             wb.active = ws
#             media_calculation(wb,ws,1,ig_obj,ig,ig_weekly,ig_total)
#         except:
#             wb.remove(wb.active)
#             print("Ig sheet skipped----")

#         try:
#             ws = wb.create_sheet('Snapchat')
#             wb.active = ws
#             media_calculation(wb,ws,1,sc_obj,sc,sc_weekly,sc_total)
#         except:
#             wb.remove(wb.active)
#             print("Snapchat sheet skipped---")
        
#         try:
#             ws = wb.create_sheet('DV360')
#             wb.active = ws
#             media_calculation(wb,ws,1,dv_obj,dv,dv_weekly,dv_total)
#         except:
#             wb.remove(wb.active)
#             print("Dv360 sheet skipped---")

#         try:
#             ws = wb.create_sheet('HPH')
#             wb.active = ws
#             media_calculation(wb,ws,1,hph_obj,hph,hph_weekly,hph_total)
#         except:
#             wb.remove(wb.active)
#             print("hph sheet skipped---")

#         try:
#             ws = wb.create_sheet('Youtube')
#             wb.active = ws
#             media_calculation(wb,ws,1,youtube_obj,youtube,youtube_weekly,youtube_total)
#         except:
#             wb.remove(wb.active)
#             print("youtube sheet skipped---")

#         # Sheet is active ====
#         daily_total = "="
#         concat_arr = []
        
#         try:
#             universal_fb = SettingUniversalDataframe(fb,fb_obj)
#             if daily_total == "=":
#                 daily_total+="'Facebook'!T2"
#             else:
#                 daily_total+="+'Facebook'!T2"
#             concat_arr.append(universal_fb)
#         except:
#             print("Universal Fb skipped----")
        
#         try:
#         # print(universal_fb["view"])
#             universal_ig = SettingUniversalDataframe(ig,ig_obj)
#             if daily_total == "=":
#                 daily_total+="'Instagram'!T2"
#             else:
#                 daily_total+="+'Instagram'!T2"
            
#             concat_arr.append(universal_ig)
#         except:
#             print("Universal ig skipped -----")

#         try:
#             universal_sc = SettingUniversalDataframe(sc,sc_obj)
#             if daily_total == "=":
#                 daily_total+="'Snapchat'!T2"
#             else:
#                 daily_total+="+'Snapchat'!T2"
#             concat_arr.append(universal_sc)
#         except:
#             print("Universal snapchat skipped ------")
#         # universal_tw = SettingUniversalDataframe(tw,tw_obj)
#         # universal_td = SettingUniversalDataframe(td,td_obj)
        
#         try:
#             universal_dv = SettingUniversalDataframe(dv,dv_obj)
#             if daily_total == "=":
#                 daily_total+="'DV360'!T2"
#             else:
#                 daily_total+="+'DV360'!T2"
#             concat_arr.append(universal_dv)
#         except:
#             print("Universal dv360 skipped----")

#         try:
#             universal_hph = SettingUniversalDataframe(hph,hph_obj)
#             if daily_total == "=":
#                 daily_total+="'hph'!T2"
#             else:
#                 daily_total+="+'hph'!T2"
#             concat_arr.append(universal_hph)
#         except:
#             print("Universal hph skipped----")

#         try:
#             universal_youtube = SettingUniversalDataframe(youtube,youtube_obj)
#             if daily_total == "=":
#                 daily_total+="'youtube'!T2"
#             else:
#                 daily_total+="+'youtube'!T2"
#             concat_arr.append(universal_youtube)
#         except:
#             print("Universal youtube skipped----")
#         # print(daily_total)

#         universal_dataframe = pd.concat(concat_arr).groupby(["date"]).sum()
        
#         # print(universal_dataframe["view"])
#         # print(universal_dataframe)
#         universal_obj = MatchingColumn(universal_dataframe)
#         # print(universal_dataframe)
#         # universal_dataframe = universal_dataframe.groupby([universal_obj.date]).sum()
#         universal_dataframe_weekly = CalculateWeekly(universal_dataframe,universal_obj)
#         universal_dataframe_total = universal_dataframe_weekly.sum()
#         ws = wb.create_sheet('Daily')
#         wb.active = ws
#         media_calculation(wb,ws,1,universal_obj,universal_dataframe,universal_dataframe_weekly,universal_dataframe_total,daily=True)
#         ws.cell(row=2,column=20,value=daily_total)
#         wb.save('summary_testing.xlsx')

#         # Generating summary ----------------------

#         have_column = True
#         ws = wb.create_sheet('Summary')
#         wb.active = ws
#         daily_media_index=2
#         try:
#             summary_calculation(wb,ws,daily_media_index,"Facebook",fb_obj,fb_total,have_column=have_column,media_label="Facebook")
#             have_column = False
#             daily_media_index+=1
#         except:
#             print("Facebook summary is missing")
#         try:
#             summary_calculation(wb,ws,daily_media_index,"Instagram",ig_obj,ig_total,have_column=have_column,media_label="Instagram")
#             have_column = False
#             daily_media_index+=1
#         except:
#             print("Instagram summary is missing")
#         try:
#             summary_calculation(wb,ws,daily_media_index,"Snapchat",sc_obj,sc_total,have_column=have_column,media_label="Snapchat")
#             have_column = False
#             daily_media_index+=1
#         except:
#             print("Snapchat summary is missing")
#         try:
#             summary_calculation(wb,ws,daily_media_index,"DV360",dv_obj,dv_total,have_column=have_column,media_label="DV360")
#             have_column = False
#             daily_media_index+=1
#         except:
#             print("DV360 summary is missing")
#         try:
#             summary_calculation(wb,ws,daily_media_index,"HPH",hph_obj,hph_total,have_column=have_column,media_label="HPH")
#             have_column = False
#             daily_media_index+=1
#         except:
#             print("HPH summary is missing")

#         try:
#             summary_calculation(wb,ws,daily_media_index,"Youtube",youtube_obj,youtube_total,have_column=have_column,media_label="Youtube")
#             have_column = False
#             daily_media_index+=1
#         except:
#             print("Youtube summary is missing")
#         try:
#             summary_calculation(wb,ws,daily_media_index,"Daily",universal_obj,universal_dataframe_total,have_column=have_column,media_label="Total")
#             have_column = False
#             daily_media_index+=1
#         except:
#             print("Daily summary is missing")
#         engine.dispose()
#         wb.save(pickle_file_path)

#         # Calculating DV 360 Asset part



#     # --------------------------------------------------------------------
#     # Asset Type
#         # end_col = LookupsGenerator(wb,"Facebook",fb_asset,fb_asset_obj,"asset_type",13)
        
#     # Assets End here----------------------------------------------------------

#     # Size 
#         # end_col = LookupsGenerator(wb,"Facebook",fb_size,fb_size_obj,"size",end_col)
#     # Size End here----------------------------------------------------------
#     # -------------------------------------------------------------------------
#     # Language 
#         # end_col = LookupsGenerator(wb,"Facebook",fb_language,fb_language_obj,"language",end_col)
#     # Language End here----------------------------------------------------------
    
#     # Country
#         # end_col = LookupsGenerator(wb,"Facebook",fb_country,fb_country_obj,"country",end_col)
#     # Country End here----------------------------------------------------------
    
#     # Message
#         # end_col = LookupsGenerator(wb,"Facebook",fb_message,fb_message_obj,"message",end_col)
#     # Message End here----------------------------------------------------------
    








#             # Daily total is coming for asset 

#             # Saving is an issue here 
#             # We have to store all assets in dv360 tab only 
#             # 
#             # Now save each asset in the separate sheet of the summary test 
            
            







#         # ---- Assets Ends ---------------------------------

        
#         # media_calculation(wb,ws,1,universal_obj,universal_dataframe,universal_dataframe_weekly,universal_dataframe_total)
#         # summary_calculation(wb,ws,1,universal_dataframe_total,fb_total,ig_total,sc_total,dv_total)
#     # From here to end -------------------------------------------------------
#             # We have to store the dataframe 

#         #     print("==FB==")
#         #     fb = SettingOtherColumns(fb,fb_obj,facebook_budget)
#         #     print("==IG==")
#         #     ig = SettingOtherColumns(ig,ig_obj,instagram_budget)
#         #     print("==SC==")
#         #     sc = SettingOtherColumns(sc,sc_obj,snapchat_budget)
#         #     print("==TW==")
#         #     tw = SettingOtherColumns(tw,tw_obj,twitter_budget)
#         #     print("==TD==")
#         #     td = SettingOtherColumns(td,td_obj,teads_budget)
#         #     print("==DV==")
#         #     dv = SettingOtherColumns(dv,dv_obj,dv360_budget)
#         #     # Setting other columns of the each media ....

#             # fb_weekly = CalculateWeekly(fb,fb_obj)
#         #     ig_weekly = CalculateWeekly(ig,ig_obj)
#         #     tw_weekly = CalculateWeekly(tw,tw_obj)
#         #     td_weekly = CalculateWeekly(td,td_obj)
#         #     sc_weekly = CalculateWeekly(sc,sc_obj)
#         #     dv_weekly = CalculateWeekly(dv,dv_obj)
#         #     # Calculating weekly records for each media ...
        

#         #     total_weekly_data_fb = len(fb_weekly.index)
#         #     total_weekly_data_ig = len(ig_weekly.index)
#         #     total_weekly_data_tw = len(tw_weekly.index)
#         #     total_weekly_data_td = len(td_weekly.index)
#         #     total_weekly_data_sc = len(sc_weekly.index)
#         #     total_weekly_data_dv = len(dv_weekly.index)
#         #     # Calculating the len of the total weekly records so we can add spaces between the tables 


#         #     fb_weekly.to_excel(writer,sheet_name="fb",index=True)
#         #     # Summary of FB --------------
#         #     fb_summary = CalculateEachMedia(fb_weekly,fb_obj,facebook_budget)
#         #     fb_summary.to_excel(writer, sheet_name="fb",startrow=total_weekly_data_fb+1, index=True,header=False)

#         #     # --------------------------------------
#         #     fb.to_excel(writer, sheet_name="fb",startrow=total_weekly_data_fb+3, index=True)
#         #     # Saving the weekly record in the sheet

#         #     # Dividing each Total budget with the value of column 

#         #     ig_weekly.to_excel(writer,sheet_name="ig",index=True)
#         #     # Summary of IG
#         #     ig_summary = CalculateEachMedia(ig_weekly,ig_obj,facebook_budget)
#         #     ig_summary.to_excel(writer, sheet_name="ig",startrow=total_weekly_data_ig+1, index=True,header=False)
#         #     # -------------------------------------- 
#         #     ig.to_excel(writer, sheet_name="ig",startrow=total_weekly_data_ig+3, index=True)
            
            
#         #     tw_weekly.to_excel(writer,sheet_name="tw",index=True)
#         #     # Summary of TW
#         #     tw_summary = CalculateEachMedia(tw_weekly,tw_obj,facebook_budget)
#         #     tw_summary.to_excel(writer, sheet_name="tw",startrow=total_weekly_data_tw+1, index=True,header=False)
#         #     # -------------------------------------- 
#         #     tw.to_excel(writer, sheet_name="tw",startrow=total_weekly_data_tw+3, index=True)
            
#         #     td_weekly.to_excel(writer,sheet_name="td",index=True)
#         #     # Summary of TD
#         #     td_summary = CalculateEachMedia(td_weekly,td_obj,facebook_budget)
#         #     td_summary.to_excel(writer, sheet_name="td",startrow=total_weekly_data_td+1, index=True,header=False)
#         #     # -------------------------------------- 
#         #     td.to_excel(writer, sheet_name="td",startrow=total_weekly_data_td+3, index=True)
            
#         #     sc_weekly.to_excel(writer,sheet_name="sc",index=True)
#         #     # Summary of SC
#         #     sc_summary = CalculateEachMedia(sc_weekly,sc_obj,facebook_budget)
#         #     sc_summary.to_excel(writer, sheet_name="sc",startrow=total_weekly_data_sc+1, index=True,header=False)
#         #     # -------------------------------------- 
#         #     sc.to_excel(writer, sheet_name="sc",startrow=total_weekly_data_sc+3, index=True)
            
#         #     dv_weekly.to_excel(writer,sheet_name="dv",index=True)
#         #     # Summary of DV
#         #     dv_summary = CalculateEachMedia(dv_weekly,dv_obj,facebook_budget)
#         #     dv_summary.to_excel(writer, sheet_name="dv",startrow=total_weekly_data_dv+1, index=True,header=False)
#         #     # -------------------------------------- 
#         #     dv.to_excel(writer, sheet_name="dv",startrow=total_weekly_data_dv+3, index=True)


#         #     # Calculating Daily Total 
#         #     # Get all the columns of the Dataframe using the class objects and add them together for Daily Total 
#             # universal_fb = SettingUniversalDataframe(fb,fb_obj)
#             # universal_ig = SettingUniversalDataframe(ig,ig_obj)
#             # universal_sc = SettingUniversalDataframe(sc,sc_obj)
#             # # universal_tw = SettingUniversalDataframe(tw,tw_obj)
#             # # universal_td = SettingUniversalDataframe(td,td_obj)
#             # universal_dv = SettingUniversalDataframe(dv,dv_obj)

#         #     # print(universal_fb)
#         #     universal_dataframe = pd.concat([
#         #         universal_fb,
#         #         universal_ig,
#         #         universal_sc,
#         #         universal_tw,
#         #         universal_td,
#         #         universal_dv
#         #     ]).groupby(["date"]).sum()

#         #     # Show only those columns which are necessary not all
#         #     # Meaningful Columns
            
        
#         #     universal_obj = MatchingColumn(universal_dataframe)
#         #     cols_to_keep = universal_obj.ListOfColumns()
#         #     print(universal_dataframe)
#         #     print(cols_to_keep)
#         #     universal_dataframe = universal_dataframe.loc[:, cols_to_keep]

#         #     universal_budget = facebook_budget + instagram_budget + snapchat_budget + twitter_budget + teads_budget + dv360_budget 
#         #     universal_dataframe = SettingOtherColumns(universal_dataframe,universal_obj,universal_budget)
#         #     print(universal_dataframe)
            
#         #     universal_dataframe_weekly = CalculateWeekly(universal_dataframe,universal_obj)
#         #     total_weekly_data_universal_dataframe = len(universal_dataframe_weekly.index)

            
#         #     universal_dataframe_weekly.to_excel(writer,sheet_name="Daily",index=True)
#         #     universal_dataframe.to_excel(writer, sheet_name="Daily",startrow=total_weekly_data_universal_dataframe+2, index=True)

#         #     # Sum all for summary

#         #     summary = universal_dataframe_weekly.copy()
                            
                        
#         #     summary.drop(summary.index, inplace=True)
#         #         # print(gdn_df_daily)
#         #     for i in range(0, 1):
#         #             row = universal_dataframe_weekly.iloc[i:len(universal_dataframe_weekly)].select_dtypes(include=['int64','double']).sum()
#         #             row['CTR'] = row['clicks'] / row['impressions']
#         #             # print(totalBudget)
#         #             row['budget'] = round(row['budget'],2)
#         #             row['spent_budget'] = round((row['budget']/universal_budget)*100)
#         #             row['CTR'] = round(row['CTR']*100,2)
                
#         #             summary = summary.append(row, ignore_index=True)
                
                
#         #     # summary = summary.assign(MVKPI_Total = TotalMonthlyViewKPI)
#         #     summary = summary.assign(CPC = round((summary["budget"] / summary["clicks"]),2))
#         #     # summary.to_sql("summary_total",con=db.engine,if_exists="replace",index=False)
#         #     list_of_summary_medias_with_data.append({
#         #         "label": "Facebook",
#         #         "data":fb_summary,
#         #         "Period": "",
#         #         "columns_object": fb_obj, 
#         #         "Visit_from_Media": "",
#         #         "budget":int(facebook_budget),
#         #     })
#         #     list_of_summary_medias_with_data.append({
#         #         "label": "Instagram",
#         #         "data":ig_summary,
#         #         "Period": "",
#         #         "columns_object": ig_obj, 
#         #         "Visit_from_Media": "",
#         #         "budget":int(instagram_budget),
#         #     })

#         #     list_of_summary_medias_with_data.append({
#         #         "label": "Snapchat",
#         #         "data":sc_summary,
#         #         "Period": "",
#         #         "columns_object": sc_obj, 
#         #         "Visit_from_Media": "",
#         #         "budget":int(snapchat_budget),
#         #     })
#         #     list_of_summary_medias_with_data.append({
#         #         "label": "Teads",
#         #         "data":td_summary,
#         #         "Period": "",
#         #         "columns_object": td_obj, 
#         #         "Visit_from_Media": "",
#         #         "budget":int(teads_budget),
#         #     })
#         #     list_of_summary_medias_with_data.append({
#         #         "label": "Twitter",
#         #         "data":tw_summary,
#         #         "Period": "",
#         #         "columns_object": tw_obj, 
#         #         "Visit_from_Media": "",
#         #         "budget":int(twitter_budget),
#         #     })
#         #     list_of_summary_medias_with_data.append({
#         #         "label": "DV360",
#         #         "data":dv_summary,
#         #         "Period": "",
#         #         "columns_object": dv_obj, 
#         #         "Visit_from_Media": "",
#         #         "budget":int(dv360_budget),
#         #     })
#         #     list_of_summary_medias_with_data.append({
#         #         "label": "Total",
#         #         "data":summary,
#         #         "budget": int(universal_budget),
#         #     })
            
        
#         #     print(summary)
#         #     summary.to_excel(writer, sheet_name="Daily",startrow=total_weekly_data_universal_dataframe+1, index=True,header=False)
            
            
            
#         # #=============== Summary logic ends ============================================================
#         #     count_start = 0
            
#         #     df2 = pd.DataFrame() 
#         #     for dt in list_of_summary_medias_with_data:
#         #             # print(dt["data"]["budget"].values[0])
#         #             Media_Column = dt["label"]
#         #             Budget_Column = dt["budget"]
#         #             Period_Column = dt["Period"] if ("columns_object" in dt) else ""
#         #             Spent_Column = dt["data"][dt["columns_object"].budget].values[0] if ("columns_object" in dt) else dt["data"]["budget"].values[0]
#         #             SPENT_BUDGET_Column = dt["data"]["SPENT_BUDGET"].values[0] if ("columns_object" in dt) else dt["data"]["spent_budget"].values[0]
#         #             Impressions_Column = dt["data"][dt["columns_object"].impressions].values[0] if ("columns_object" in dt) else dt["data"]["impressions"].values[0]
#         #             Clicks_Column = dt["data"][dt["columns_object"].clicks].values[0] if ("columns_object" in dt) else dt["data"]["clicks"].values[0]
#         #             CTR_Column = dt["data"]["CTR"].values[0]
#         #             CPC = dt["data"]["CPC"].values[0] 
#         #             Visit_from_Media = dt["Visit_from_Media"] if ("columns_object" in dt) else ""
#         #             Visit_KPI = ""
#         #             KPI_Percent = ""
#         #             CPT = ""
#         #             CVR = ""
#         #             Bounce_Rate = "=SUM(A1:C1)"
#         #             df = pd.DataFrame([[Media_Column,Period_Column,Budget_Column,Spent_Column,SPENT_BUDGET_Column,Impressions_Column,Clicks_Column,CTR_Column,CPC,Visit_from_Media,Visit_KPI,KPI_Percent,CPT,CVR,Bounce_Rate]],columns=['Media','Period','Budget','Budget Spent','Budget Spent %','Impressions','Clicks','CTR','CPC','Visit from Media','Visit KPI','KPI %','CPT','CVR (Media to LG.COM)','Bounce Rate'])
#         #             df2 = df2.append(df)

#         #     # print(df2)
#         #     df2.to_excel(writer, sheet_name="Summary",index=False)

#         # # Now next is weekly _total so we can move on to summary 

#         # # daily_total_weekly_total =  pd.pivot_table(daily_total_weekly_dataframe,values=column_values,aggfunc = 'sum')


#         # # print(list_of_dataframes)
        
#         # return send_file(writer),200
#     # -----------------------------------------------------





#     # except:
#     #     return jsonify({
#     #         "message": "File not uploaded",
#     #         "status":500
#     #     }),500
#         return jsonify({
#                 "message": "Successfully uploaded",
#                 "status": 200
#         }),200
#     # return send_file(pickle_file_path, as_attachment=True)
    
        
      
        
#         # Create a unique dataframe with all the column names are matching 
#         # Then sum all the dataframes and generate in excel

#         # Function to change all the columns of the dataframe according to the universal columns 
        
        
#         # print(universal_dataframe)

#     # -------------------------------------------------------------------


# # Product wise data (No Language / Mixed Images and videos) ----
@app.route("/mc_product_wise",methods=["POST"])
def mc_product_wise():
    
    # Fetch the raw data file (having all the platforms separated in each tab) 
    try:
        raw_data = request.files["raw_data"]
        start_date = request.form['start_date']
        end_date = request.form['end_date']

        facebook_instagram = pd.read_excel(raw_data,sheet_name="META") 
        # snapchat = pd.read_excel(raw_data,sheet_name="Snapchat") 
        gdn = pd.read_excel(raw_data,sheet_name="GDN - Google Ads") 
        # criteo = pd.read_excel(raw_data,sheet_name="Criteo") 
        youtube = pd.read_excel(raw_data,sheet_name="YouTube") 
        
    except:
        print("Raw Data skipped---")
    # Go to each tab, fetch the columns and merge them together 

    # fb_ig = set_columns(facebook_instagram)

    # Separate them market wise
    # Then in each market (Country), separate them using platform wise
    # Save the weeklydata, dailytotal and summary (for each column in a object for each platform using date)

    # Then separate platform wise data using product wise 


   


    try:
        fb_ig_obj = MatchingColumn(facebook_instagram)      # Compare each column and if the data is matching with any of the list attribute in the class it will save it there so we don't have to call facebook dataframe according to the static name 
        #Storing column names in the class attributes so we dont memorize the names everytime. 
        # snapchat_obj = MatchingColumn(snapchat)
        gdn_obj = MatchingColumn(gdn)
        # criteo_obj = MatchingColumn(criteo)
        youtube_obj = MatchingColumn(youtube)

    except:
        print("Issue with the matching column")
    
    

    


    
    try:
     
        facebook_instagram[fb_ig_obj.date] = pd.to_datetime(facebook_instagram[fb_ig_obj.date])
        facebook_instagram.sort_values(by=fb_ig_obj.date, inplace=True)
        # Sorting the data according to date 
        facebook_instagram[fb_ig_obj.date] = facebook_instagram[fb_ig_obj.date].dt.strftime('%Y-%m-%d')
        # Converting the data back to string to store in the dataframe
        
        # snapchat[snapchat_obj.date] = pd.to_datetime(snapchat[snapchat_obj.date])
        # snapchat.sort_values(by=snapchat_obj.date, inplace=True)
        # # Sorting the data according to date 
        # snapchat[snapchat_obj.date] = snapchat[snapchat_obj.date].dt.strftime('%Y-%m-%d')
        
        gdn[gdn_obj.date] = pd.to_datetime(gdn[gdn_obj.date])
        gdn.sort_values(by=gdn_obj.date, inplace=True)
        # Sorting the data according to date 
        gdn[gdn_obj.date] = gdn[gdn_obj.date].dt.strftime('%Y-%m-%d')
        # criteo[criteo_obj.date] = pd.to_datetime(criteo[criteo_obj.date])
        # criteo.sort_values(by=criteo_obj.date, inplace=True)
        # # Sorting the data according to date 
        # criteo[criteo_obj.date] = criteo[criteo_obj.date].dt.strftime('%Y-%m-%d')
        youtube[youtube_obj.date] = pd.to_datetime(youtube[youtube_obj.date])
        youtube.sort_values(by=youtube_obj.date, inplace=True)
        # Sorting the data according to date 
        youtube[youtube_obj.date] = youtube[youtube_obj.date].dt.strftime('%Y-%m-%d')

        # print(snapchat_obj)
    except:
        print("Issue in Date")
    # try:


    # Create a dataframe 
    universal_df_fb_ig = pd.DataFrame()
    list_of_columns = fb_ig_obj.ListOfColumns()
    for col in list_of_columns:
        # print(col)
        universal_df_fb_ig[col] = facebook_instagram[getattr(fb_ig_obj, col)]
        # Store those columns which are needed and save it in meaningful variable
    # print("Countries:",universal_df_fb_ig['country'].unique())
    # universal_df_sc = pd.DataFrame()
    # # list_of_columns = snapchat_obj.ListOfColumns()
    # for col in list_of_columns:
    #     # print(col)
    #     universal_df_sc[col] = snapchat[getattr(snapchat_obj, col)]
        # Store those columns which are needed and save it in meaningful variable
    # print(universal_df_sc)
    universal_df_gdn = pd.DataFrame()
    # list_of_columns = gdn_obj.ListOfColumns()
    for col in list_of_columns:
        if col in ["view","percent_25","percent_50","percent_75","percent_100"]:
            universal_df_gdn[col] = 0
        else:
            universal_df_gdn[col] = gdn[getattr(gdn_obj, col)]
        # Store those columns which are needed and save it in meaningful variable
    # print(universal_df_gdn)
    # universal_df_criteo = pd.DataFrame()
    # # list_of_columns = criteo_obj.ListOfColumns()
    # for col in list_of_columns:
    #     if col in ["view","percent_25","percent_50","percent_75","percent_100"]:
    #         universal_df_criteo[col] = 0
    #     else:
    #         universal_df_criteo[col] = criteo[getattr(criteo_obj, col)]
        # Store those columns which are needed and save it in meaningful variable
    # print(universal_df_criteo)
    universal_df_youtube = pd.DataFrame()
    # list_of_columns = youtube_obj.ListOfColumns()
    for col in list_of_columns:
        
        universal_df_youtube[col] = youtube[getattr(youtube_obj, col)]
        # Store those columns which are needed and save it in meaningful variable
    # print(universal_df_youtube)
    # Add all datafram together
    # universal_df = pd.concat([universal_df_fb_ig, universal_df_sc, universal_df_gdn, universal_df_criteo,universal_df_youtube], ignore_index=True)
    universal_df = pd.concat([universal_df_fb_ig, 
                            #   universal_df_sc, 
                              universal_df_gdn, 
                            #   universal_df_criteo,
                              universal_df_youtube], ignore_index=True)
    # print(universal_df)
    # except:
    #     print("Issue in universal")
        # # testDf = pd.DataFrame(columns=fb_ig_obj.l)
    
    
    universal_df # Dataframe with data
    list_of_columns # List of columns for dataframe 
    print(universal_df)
    print("---------Till here---------")
    print("---------------------------")
    # Separate them country wise ---------------------------------------------------------------------------

    # Country names 
    # list_of_countries = universal_df['country'].unique()
    # list_of_dataframes_by_country = []

    # for country in list_of_countries:
    #     filtered_condition = universal_df['country'].str.contains(country, case=False)    
    #     list_of_dataframes_by_country.append({"country":country,"data": universal_df[filtered_condition]})
        # print(f"country: {country}")
        # print(universal_df[filtered_condition])

    # print(list_of_dataframes_by_country)
    # --------------------------------------------------------------------------------------------------------
    list_of_columns_for_platform = fb_ig_obj.ListOfColumnsForPlatform()
    # print(list_of_columns_for_platform)
    list_of_dataframes_eachcountry_by_platform = []
    list_of_dataframes_eachcountry_by_platform_by_product=[]

    # for country_info in list_of_dataframes_by_country:
    #     countryName= country_info["country"]    # Country name
    #     eachCountryDf = country_info["data"]       # Country dataframe

        # Create a new dataframe with only platform columns 

        
        # For platform only ------------------------------------------------------
    list_of_platforms = universal_df['platform'].unique()
    # Get unique list of platforms 
    youtube_check = True
    # Youtube check True if the record has multiple youtube 
    for platform in list_of_platforms:
        # getting each platform 
        if "YouTube".lower() in (platform).lower() and youtube_check == True:
            # If the platform name has youtube then just count one and skip all other names of youtube so we can unite all youtube data in one
            filtered_condition = universal_df['platform'].str.contains("YouTube", case=False)
            # Filter condition if the platform value has youtube then get the data
            new_df = universal_df[filtered_condition]
            # new dataframe has all the records of the youtube 
            eachPlatformDf = pd.DataFrame(columns=list_of_columns_for_platform)
            # creating a new dataframe and passing list of columns for platform 
            for col in list_of_columns_for_platform:
                # going to each column and storing it to new dataframe and the new dataframe has only specific columns and not country language and platform records
                eachPlatformDf[col] = new_df[col]

            list_of_dataframes_eachcountry_by_platform.append({"platform":"YouTube","data": eachPlatformDf })
            youtube_check = False  
        elif "YouTube".lower() in (platform).lower() and youtube_check == False:
            continue
        else:
            filtered_condition = universal_df['platform'].str.contains(platform, case=False)
            new_df = universal_df[filtered_condition]
            
            
            # print(new_df.columns)
            eachPlatformDf = pd.DataFrame(columns=list_of_columns_for_platform)
            for col in list_of_columns_for_platform:
                # print(col)
                eachPlatformDf[col] = new_df[col]
            eachPlatformDf      # This is the dataframe for each platform 
           
            list_of_dataframes_eachcountry_by_platform.append({"platform":platform,"data": eachPlatformDf})
            
    # For platform only -------------------------------------------------------
    
    # For assets only 
    youtube_check = True
    for platform in list_of_platforms:
        # getting each platform 
        if "YouTube".lower() in (platform).lower() and youtube_check == True:
            # If the platform name has youtube then just count one and skip all other names of youtube so we can unite all youtube data in one
            filtered_condition = universal_df['platform'].str.contains("YouTube", case=False)
            # Filter condition if the platform value has youtube then get the data
            new_df_platform = universal_df[filtered_condition]
            # List of assets of that particular dataframe platform wise 
            list_of_assets = new_df_platform['asset_type'].unique()
            for asset_type in list_of_assets:
                filtered_condition_assets = new_df_platform['asset_type'].str.contains(asset_type, case=False)
                new_df_asset = new_df_platform[filtered_condition_assets]

                list_of_products = new_df_asset['product'].unique()
                for prod in list_of_products:
                    filtered_condition_products = new_df_asset['product'].str.contains(prod, case=False)
                    new_df_prod = new_df_asset[filtered_condition_products]

                    # New dataframe for getting the data only for that particular product 
                    eachPlatformProductDf = pd.DataFrame(columns=list_of_columns_for_platform)
                    for col in list_of_columns_for_platform:
                        eachPlatformProductDf[col] = new_df_prod[col]
                    # print({"platform":"YouTube","product": prod,"data": eachPlatformProductDf })
                    list_of_dataframes_eachcountry_by_platform_by_product.append({"platform":"YouTube","asset_type":asset_type,"product": prod,"data": eachPlatformProductDf })
            
            youtube_check = False  
                # --------------------------------------------------------------------
         
        elif "YouTube".lower() in (platform).lower() and youtube_check == False:
            pass
        else:
            filtered_condition = universal_df['platform'].str.contains(platform, case=False)
            # Filter condition if the platform value has youtube then get the data
            new_df_platform = universal_df[filtered_condition]
            # List of assets of that particular dataframe platform wise 
            list_of_assets = new_df_platform['asset_type'].unique()
            for asset_type in list_of_assets:
                filtered_condition_assets = new_df_platform['asset_type'].str.contains(asset_type, case=False)
                new_df_asset = new_df_platform[filtered_condition_assets]

                list_of_products = new_df_asset['product'].unique()
                for prod in list_of_products:
                    filtered_condition_products = new_df_asset['product'].str.contains(prod, case=False)
                    new_df_prod = new_df_asset[filtered_condition_products]

                    # New dataframe for getting the data only for that particular product 
                    eachPlatformProductDf = pd.DataFrame(columns=list_of_columns_for_platform)
                    for col in list_of_columns_for_platform:
                        eachPlatformProductDf[col] = new_df_prod[col]
                    # print({"platform":"YouTube","product": prod,"data": eachPlatformProductDf })
                    list_of_dataframes_eachcountry_by_platform_by_product.append({"platform":platform,"asset_type":asset_type,"product": prod,"data": eachPlatformProductDf })
    # print("Total Dataframes:",len(list_of_dataframes_eachcountry_by_platform_by_product))
        # -----------------------------------------------------------------------
    
    # Then in each market (Country), separate them using platform wise
    
    # print(list_of_dataframes_eachcountry_by_platform)
    
    # Loop through list of countries 
    
    for entry in list_of_dataframes_eachcountry_by_platform:
      
        platform = entry['platform']
        data_frame = entry['data']
        

        # print(platform)


        # Create or open the Excel file for the country
        file_name = f'Morocco_data.xlsx'
        start_date=pd.to_datetime(start_date)


        end_date=pd.to_datetime(end_date)
   
        
       
        temp_start_date = start_date
        temp_end_date = end_date
        
        data_values = []
        while temp_start_date <= temp_end_date:
            # Your logic ....
            # Add row with date 
            
            new_data_temp_start_date = [temp_start_date.strftime('%Y-%m-%d')]
            # Created new array with end date which will be incremented one by one  
            for column in list_of_columns_for_platform:
                # Going to each column and if the column is not date then add 0 to the array
                if column != 'date':
                    new_data_temp_start_date.append(0)
      
            
            data_values.append(new_data_temp_start_date)
  
            temp_start_date = temp_start_date + timedelta(days=1)
        
        newDF= pd.DataFrame(data_values, columns=list_of_columns_for_platform)
        data_frame = pd.concat([data_frame,newDF ], ignore_index=True)

        data_frame.sort_values(by='date', inplace=True)
        data_frame = data_frame.groupby(["date"]).sum()
        
        


        # Add extra columns here and show the data
        list_of_formula_columns = ['Budget Spent', 'CTR','CPC','CPV'] 
        
        i=1
        for _, row in data_frame.iterrows():
            row['CTR'] = f'=(D{i}/C{i})'

        # ----------------------------------------
        df_weekly = CalculateWeekly(data_frame,list_of_columns_for_platform)


        # printing_daily_total = data_frame.copy()  # For printing daily + weekly records

        # print(['date'] + data_frame.columns.append())
        columns_ = data_frame.columns.tolist()

        # Add a new column name to the beginning of the list
        new_column_name = 'date'
        columns_.insert(0, new_column_name)
        # print(columns_)
        print_daily_total_df = pd.DataFrame(columns=columns_)
        # for _, row in df_weekly.iterrows():
        #     values_list = row.values.tolist()
            
        #     print(values_list)
            # This is row of week


        records_added = 0

        # Define the number of records after which to insert a values_list
        insert_every = 7

        weeklyTable = []
        # Iterate through rows and create DataFrames for each values_list and separator row
        for _, row in df_weekly.iterrows():
            eachWeek = row.values.tolist() # This is week 1 
            weeklyTable.append([f"{int(_)+1} Weekly Total"]+eachWeek)

        records_added = 0
        weekly_ind=0
        for _, row in data_frame.iterrows():
            eachEntry = row.values.tolist() # This is dailyEntry
            print_daily_total_df.loc[len(print_daily_total_df)] = [_]+eachEntry
            records_added +=1
            if records_added % insert_every == 0:
                print_daily_total_df.loc[len(print_daily_total_df)] = weeklyTable[weekly_ind] 
                weekly_ind+=1
               
           


        # print(printing_daily_total)
        summary_total = df_weekly.sum().to_frame().T
        
        # print(summary_total)
        excel_file_path = f'final_output.xlsx'
        if os.path.exists(excel_file_path):
            # If the file exists, load it and add a new sheet with the DataFrame data
            # print("Check if it is running")
            wb = load_workbook(excel_file_path)
            

        else:
            wb = Workbook()
      
        ws = wb.create_sheet(platform)
        wb.active = ws
        wb,fb_start_row,start_row = media_calculation(wb,ws,1,fb_ig_obj,print_daily_total_df,df_weekly,summary_total)
        
        wb.save(excel_file_path)
        
    # {
    #     country: UAE,
    #     data: {"facebook": 13,"youtube": 13}
    # }
    # countryListWithAssets = []
    # for countryName in list_of_countries:
    temp = {}
    youtube_ = False
    # temp["country"]= countryName
    for plat in list_of_platforms:
        if youtube_ == False and "YouTube" in plat:
            temp["YouTube"] = 13
            youtube_ = True
        elif youtube_ == True:
            pass
        else:        
            temp[plat] = 13
    # countryListWithAssets.append(temp)
    
    # print(len(list_of_dataframes_eachcountry_by_platform_by_product))
    # Assets tables 
    # print(len(list_of_dataframes_eachcountry_by_platform_by_product))
    
    for entry in list_of_dataframes_eachcountry_by_platform_by_product:
        # country = entry['country']
        platform = entry['platform']
        product = entry['product']
        asset_type = entry['asset_type']
        data_frame_product = entry['data']
        new_column = temp[platform]
        
        print(f"{platform}:",product)
        # print(platform)


        # Create or open the Excel file for the country
        file_name = f'final_output.xlsx'
        # start_date=pd.to_datetime(start_date)


        # end_date=pd.to_datetime(end_date)
   
        
       
        temp_start_date = start_date
        temp_end_date = end_date
        
        data_values = []
        while temp_start_date <= temp_end_date:
            # Your logic ....
            # Add row with date 
            
            new_data_temp_start_date = [temp_start_date.strftime('%Y-%m-%d')]
            # Created new array with end date which will be incremented one by one  
            for column in list_of_columns_for_platform:
                # Going to each column and if the column is not date then add 0 to the array
                if column != 'date':
                    new_data_temp_start_date.append(0)
      
            
            data_values.append(new_data_temp_start_date)
  
            temp_start_date = temp_start_date + timedelta(days=1)
        
        newDF= pd.DataFrame(data_values, columns=list_of_columns_for_platform)
        data_frame_product = pd.concat([data_frame_product,newDF ], ignore_index=True)

        data_frame_product.sort_values(by='date', inplace=True)
        data_frame_product = data_frame_product.groupby(["date"]).sum()
        
        # print(data_frame_product)


        # Add extra columns here and show the data
        list_of_formula_columns = ['Budget Spent', 'CTR','CPC','CPV'] 
        
        i=1
        for _, row in data_frame_product.iterrows():
            row['CTR'] = f'=(D{i}/C{i})'

        # ----------------------------------------
        df_weekly_assets = CalculateWeekly(data_frame_product,list_of_columns_for_platform)


        # printing_daily_total = data_frame_product.copy()  # For printing daily + weekly records

        # print(['date'] + data_frame_product.columns.append())
        columns_ = data_frame_product.columns.tolist()

        # Add a new column name to the beginning of the list
        new_column_name = 'date'
        columns_.insert(0, new_column_name)
        # print(columns_)
        print_daily_total_df_assets = pd.DataFrame(columns=columns_)
      


        records_added = 0

        # Define the number of records after which to insert a values_list
        insert_every = 7

        weeklyTable = []
        # Iterate through rows and create DataFrames for each values_list and separator row
        for _, row in df_weekly_assets.iterrows():
            eachWeek = row.values.tolist() # This is week 1 
            weeklyTable.append([f"{int(_)+1} Weekly Total"]+eachWeek)

        records_added = 0
        weekly_ind=0
        for _, row in data_frame_product.iterrows():
            eachEntry = row.values.tolist() # This is dailyEntry
            print_daily_total_df_assets.loc[len(print_daily_total_df_assets)] = [_]+eachEntry
            records_added +=1
            if records_added % insert_every == 0:
                print_daily_total_df_assets.loc[len(print_daily_total_df_assets)] = weeklyTable[weekly_ind] 
                weekly_ind+=1
               
           


        # print(print_daily_total_df_assets)
        summary_total_assets = df_weekly_assets.sum().to_frame().T
        
        # print(summary_total_assets)
        excel_file_path = f'final_output.xlsx'
        # if os.path.exists(excel_file_path):
            # If the file exists, load it and add a new sheet with the DataFrame data
            # print("Check if it is running")
        wb = load_workbook(excel_file_path)
        ws = wb[platform]

        # else:
        #     wb = Workbook()
        #     ws = wb.create_sheet(platform)
        wb.active = ws
        # print(countryListWithAssets)
        

        # print(f"{country}: {platform} : {new_column}")
        wb,new_record_column = media_calculation_assets(wb,ws,1,fb_ig_obj,print_daily_total_df_assets,df_weekly_assets,summary_total_assets,start_row,new_column,product,asset_type)
        temp[platform] = new_record_column
        # print(countryListWithAssets)
        # list_of_platforms_for_assets[platform] = new_column
        # print(country)
        wb.save(excel_file_path)



    
    # Loop here for the dataframe eachcountry_by_platform_by asset
    # print(list_of_dataframes_eachcountry_by_platform_by_product)            

    # ============================================================


    return jsonify({
        "message": "completed"
    }),200
        # try:    
        #     fb_asset[fb_asset_obj.date] = pd.to_datetime(fb_asset[fb_asset_obj.date])

        #     fb_asset.sort_values(by=fb_asset_obj.date, inplace=True)
        #     fb_asset[fb_asset_obj.date] = fb_asset[fb_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Assets dates issues")

        # try:    
        #     fb_size[fb_size_obj.date] = pd.to_datetime(fb_size[fb_size_obj.date])

        #     fb_size.sort_values(by=fb_size_obj.date, inplace=True)
        #     fb_size[fb_size_obj.date] = fb_size[fb_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Size dates issues")

        # try:    
        #     fb_language[fb_language_obj.date] = pd.to_datetime(fb_language[fb_language_obj.date])

        #     fb_language.sort_values(by=fb_language_obj.date, inplace=True)
        #     fb_language[fb_language_obj.date] = fb_language[fb_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Language dates issues")

        # try:    
        #     fb_country[fb_country_obj.date] = pd.to_datetime(fb_country[fb_country_obj.date])

        #     fb_country.sort_values(by=fb_country_obj.date, inplace=True)
        #     fb_country[fb_country_obj.date] = fb_country[fb_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Country dates issues")

        # try:    
        #     fb_message[fb_message_obj.date] = pd.to_datetime(fb_message[fb_message_obj.date])

        #     fb_message.sort_values(by=fb_message_obj.date, inplace=True)
        #     fb_message[fb_message_obj.date] = fb_message[fb_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Message dates issues")
       

        # try:
        #     ig[ig_obj.date] = pd.to_datetime(ig[ig_obj.date])
        #     ig.sort_values(by=ig_obj.date, inplace=True)
        #     ig[ig_obj.date] = ig[ig_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("ig Dates issues")

        # try:    
        #     ig_asset[ig_asset_obj.date] = pd.to_datetime(ig_asset[ig_asset_obj.date])

        #     ig_asset.sort_values(by=ig_asset_obj.date, inplace=True)
        #     ig_asset[ig_asset_obj.date] = ig_asset[ig_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Assets dates issues")

        # try:    
        #     ig_size[ig_size_obj.date] = pd.to_datetime(ig_size[ig_size_obj.date])

        #     ig_size.sort_values(by=ig_size_obj.date, inplace=True)
        #     ig_size[ig_size_obj.date] = ig_size[ig_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Size dates issues")

        # try:    
        #     ig_language[ig_language_obj.date] = pd.to_datetime(ig_language[ig_language_obj.date])

        #     ig_language.sort_values(by=ig_language_obj.date, inplace=True)
        #     ig_language[ig_language_obj.date] = ig_language[ig_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Language dates issues")

        # try:    
        #     ig_country[ig_country_obj.date] = pd.to_datetime(ig_country[ig_country_obj.date])

        #     ig_country.sort_values(by=ig_country_obj.date, inplace=True)
        #     ig_country[ig_country_obj.date] = ig_country[ig_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Country dates issues")

        # try:    
        #     ig_message[ig_message_obj.date] = pd.to_datetime(ig_message[ig_message_obj.date])

        #     ig_message.sort_values(by=ig_message_obj.date, inplace=True)
        #     ig_message[ig_message_obj.date] = ig_message[ig_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Message dates issues")




        # try:
        #     sc[sc_obj.date] = pd.to_datetime(sc[sc_obj.date])
        #     rawData = sc
        #     sc.sort_values(by=sc_obj.date, inplace=True)
        #     sc[sc_obj.date] = sc[sc_obj.date].dt.strftime('%Y-%m-%d')

        #     rawData[sc_obj.date] = pd.to_datetime(sc[sc_obj.date])
        # except:
        #     print("sc Dates issues")

        # try:    
        #     sc_asset[sc_asset_obj.date] = pd.to_datetime(sc_asset[sc_asset_obj.date])

        #     sc_asset.sort_values(by=sc_asset_obj.date, inplace=True)
        #     sc_asset[sc_asset_obj.date] = sc_asset[sc_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Assets dates issues")

        # try:    
        #     sc_size[sc_size_obj.date] = pd.to_datetime(sc_size[sc_size_obj.date])

        #     sc_size.sort_values(by=sc_size_obj.date, inplace=True)
        #     sc_size[sc_size_obj.date] = sc_size[sc_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Size dates issues")

        # try:    
        #     sc_language[sc_language_obj.date] = pd.to_datetime(sc_language[sc_language_obj.date])

        #     sc_language.sort_values(by=sc_language_obj.date, inplace=True)
        #     sc_language[sc_language_obj.date] = sc_language[sc_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Language dates issues")

        # try:    
        #     sc_country[sc_country_obj.date] = pd.to_datetime(sc_country[sc_country_obj.date])

        #     sc_country.sort_values(by=sc_country_obj.date, inplace=True)
        #     sc_country[sc_country_obj.date] = sc_country[sc_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Country dates issues")

        # try:    
        #     sc_message[sc_message_obj.date] = pd.to_datetime(sc_message[sc_message_obj.date])

        #     sc_message.sort_values(by=sc_message_obj.date, inplace=True)
        #     sc_message[sc_message_obj.date] = sc_message[sc_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Message dates issues")

      
        

        # # Dv 360 -------------------------------------------
        # try:
        #     dv[dv_obj.date] = pd.to_datetime(dv[dv_obj.date])
        #     dv.sort_values(by=dv_obj.date, inplace=True)
        #     dv[dv_obj.date] = dv[ig_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("dv Dates issues")

        # try:    
        #     dv_asset[dv_asset_obj.date] = pd.to_datetime(dv_asset[dv_asset_obj.date])

        #     dv_asset.sort_values(by=dv_asset_obj.date, inplace=True)
        #     dv_asset[dv_asset_obj.date] = dv_asset[dv_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Assets dates issues")

        # try:    
        #     dv_size[dv_size_obj.date] = pd.to_datetime(dv_size[dv_size_obj.date])

        #     dv_size.sort_values(by=dv_size_obj.date, inplace=True)
        #     dv_size[dv_size_obj.date] = dv_size[dv_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Size dates issues")

        # try:    
        #     dv_language[dv_language_obj.date] = pd.to_datetime(dv_language[dv_language_obj.date])

        #     dv_language.sort_values(by=dv_language_obj.date, inplace=True)
        #     dv_language[dv_language_obj.date] = dv_language[dv_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Language dates issues")

        # try:    
        #     dv_country[dv_country_obj.date] = pd.to_datetime(dv_country[dv_country_obj.date])

        #     dv_country.sort_values(by=dv_country_obj.date, inplace=True)
        #     dv_country[dv_country_obj.date] = dv_country[dv_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Country dates issues")

        # try:    
        #     dv_message[dv_message_obj.date] = pd.to_datetime(dv_message[dv_message_obj.date])

        #     dv_message.sort_values(by=dv_message_obj.date, inplace=True)
        #     dv_message[dv_message_obj.date] = dv_message[dv_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Message dates issues")

        # --------------------------------------------------------
        
        

       
        # Formatting dates and sorting the values according to dates 
        
        
        
        # Creating path for the final output file 

        
        # ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
    #     try:
    #         fb = facebook_instagram.groupby([fb_obj.date]).sum()

    #     except:
    #         print("fb sum issues")
        
    #     try: 
    #         ig = ig.groupby([ig_obj.date]).sum()
    #     except:
    #         print("ig sum issues")
        
    #     try: 
    #         sc = sc.groupby([sc_obj.date]).sum()
    #     except:
    #         print("sc sum issues")
    #     try: 
    #         dv = dv.groupby([dv_obj.date]).sum()
    #     except:
    #         print("dv sum issues")
        

    #     # ----------------------------------------------
    #     try:
    #         fb_weekly = CalculateWeekly(fb,fb_obj)
    #     except:
    #         print("fb_weekly issue")
    #     try:
    #         ig_weekly = CalculateWeekly(ig,ig_obj)
    #     except:
    #         print("ig_weekly issue")
    #     try:
    #         sc_weekly = CalculateWeekly(sc,sc_obj)
    #     except:
    #         print("sc_weekly issue")
    #     try:
    #         dv_weekly = CalculateWeekly(dv,dv_obj)
    #     except:
    #         print("dv_weekly issue")
        
       
    #     # =================================================
    #     try:
    #         fb_total = fb_weekly.sum()
    #     except:
    #         print("fb_total issue")
    #     try:
    #         ig_total = ig_weekly.sum()
    #     except:
    #         print("ig_total issue")
    #     try:
    #         sc_total = sc_weekly.sum()
    #     except:
    #         print("sc_total issue")
    #     try:
    #         dv_total = dv_weekly.sum()
    #     except:
    #         print("dv_total issue")
   
        
    #     # Summing the data according to the date so we have one data for each data
        
    #     # Until here everything is fine 
    #     # Instead of setting columns using dataframe we will set them in excel anad update the excel sheet according to row and column 
    #     wb = Workbook()
    #     wb.remove(wb.active)
    #     try:
    #         ws = wb.create_sheet('Facebook')
    #         wb.active = ws
    #         fb_start_row = media_calculation(wb,ws,1,fb_obj,fb,fb_weekly,fb_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Facebook sheet skipped----")

    #     try:
    #         ws = wb.create_sheet('Instagram')
    #         wb.active = ws
    #         ig_start_row = media_calculation(wb,ws,1,ig_obj,ig,ig_weekly,ig_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Ig sheet skipped----")

    #     try:
    #         ws = wb.create_sheet('Youtube')
    #         wb.active = ws
    #         youtube_start_row = media_calculation(wb,ws,1,sc_obj,sc,sc_weekly,sc_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Youtube sheet skipped---")
        
    #     try:
    #         ws = wb.create_sheet('GDN')
    #         wb.active = ws
    #         gdn_start_row = media_calculation(wb,ws,1,dv_obj,dv,dv_weekly,dv_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("GDN sheet skipped---")

       

    #     # Sheet is active ====
    #     daily_total = "="
    #     concat_arr = []
        
    #     try:
    #         universal_fb = SettingUniversalDataframe(fb,fb_obj)
    #         if daily_total == "=":
    #             daily_total+="'Facebook'!T2"
    #         else:
    #             daily_total+="+'Facebook'!T2"
    #         concat_arr.append(universal_fb)
    #     except:
    #         print("Universal Fb skipped----")
        
    #     try:
    #     # print(universal_fb["view"])
    #         universal_ig = SettingUniversalDataframe(ig,ig_obj)
    #         if daily_total == "=":
    #             daily_total+="'Instagram'!T2"
    #         else:
    #             daily_total+="+'Instagram'!T2"
            
    #         concat_arr.append(universal_ig)
    #     except:
    #         print("Universal ig skipped -----")

    #     try:
    #         universal_sc = SettingUniversalDataframe(sc,sc_obj)
    #         if daily_total == "=":
    #             daily_total+="'Youtube'!T2"
    #         else:
    #             daily_total+="+'Youtube'!T2"
    #         concat_arr.append(universal_sc)
    #     except:
    #         print("Universal Youtube skipped ------")
    #     # universal_tw = SettingUniversalDataframe(tw,tw_obj)
    #     # universal_td = SettingUniversalDataframe(td,td_obj)
        
    #     try:
    #         universal_dv = SettingUniversalDataframe(dv,dv_obj)
    #         if daily_total == "=":
    #             daily_total+="'GDN'!T2"
    #         else:
    #             daily_total+="+'GDN'!T2"
    #         concat_arr.append(universal_dv)
    #     except:
    #         print("Universal GDN skipped----")

       

    #     universal_dataframe = pd.concat(concat_arr).groupby(["date"]).sum()
        
        
    #     universal_obj = MatchingColumn(universal_dataframe)
        
    #     universal_dataframe_weekly = CalculateWeekly(universal_dataframe,universal_obj)
    #     universal_dataframe_total = universal_dataframe_weekly.sum()
    #     ws = wb.create_sheet('Daily')
    #     wb.active = ws
    #     media_calculation(wb,ws,1,universal_obj,universal_dataframe,universal_dataframe_weekly,universal_dataframe_total,daily=True)
    #     ws.cell(row=2,column=20,value=daily_total)
    #     wb.save('summary_testing.xlsx')

    #     # Generating summary ----------------------

    #     have_column = True
    #     ws = wb.create_sheet('Summary')
    #     wb.active = ws
    #     daily_media_index=2
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Facebook",fb_obj,fb_total,have_column=have_column,media_label="Facebook")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Facebook summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Instagram",ig_obj,ig_total,have_column=have_column,media_label="Instagram")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Instagram summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Youtube",sc_obj,sc_total,have_column=have_column,media_label="Youtube")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Snapchat summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"GDN",dv_obj,dv_total,have_column=have_column,media_label="GDN")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("GDN summary is missing")
        
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Daily",universal_obj,universal_dataframe_total,have_column=have_column,media_label="Total")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Daily summary is missing")
    #     # engine.dispose()
    #     wb.save(pickle_file_path)

    #     # Calculating Asset part



    # # --------------------------------------------------------------------
    #     try:
    #         end_col = LookupsGenerator(wb,"Facebook",fb_asset,fb_asset_obj,"asset_type",13,fb_start_row)
    #         end_col = LookupsGenerator(wb,"Facebook",fb_language,fb_language_obj,"language",end_col,fb_start_row)
    #         # end_col = LookupsGenerator(wb,"Facebook",fb_country,fb_country_obj,"country",end_col,fb_start_row)
    #         end_col = LookupsGenerator(wb,"Facebook",fb_message,fb_message_obj,"message",end_col,fb_start_row)
    #     except:
    #         print("Facebook Lookup failed")
    #     try:
    #         end_col = LookupsGenerator(wb,"Instagram",ig_asset,ig_asset_obj,"asset_type",13,ig_start_row)
    #         end_col = LookupsGenerator(wb,"Instagram",ig_language,ig_language_obj,"language",end_col,ig_start_row)
    #         # end_col = LookupsGenerator(wb,"Instagram",ig_country,ig_country_obj,"country",end_col,ig_start_row)
    #         end_col = LookupsGenerator(wb,"Instagram",ig_message,ig_message_obj,"message",end_col,ig_start_row)
    #     except:
    #         print("Ig lookup failed")
        
    #     try:
    #         end_col = LookupsGenerator(wb,"GDN",dv_asset,dv_asset_obj,"asset_type",13,gdn_start_row)
    #         end_col = LookupsGenerator(wb,"GDN",dv_language,dv_language_obj,"language",end_col,gdn_start_row)
    #         # end_col = LookupsGenerator(wb,"GDN",dv_country,dv_country_obj,"country",end_col,gdn_start_row)
    #         end_col = LookupsGenerator(wb,"GDN",dv_message,dv_message_obj,"message",end_col,gdn_start_row)
    #     except:
    #         print("GDN Lookup failed")

    #     try:
    #         end_col = LookupsGenerator(wb,"Youtube",sc_asset,sc_asset_obj,"asset_type",13,youtube_start_row)
    #         end_col = LookupsGenerator(wb,"Youtube",sc_language,sc_language_obj,"language",end_col,youtube_start_row)
    #         # end_col = LookupsGenerator(wb,"Youtube",sc_country,sc_country_obj,"country",end_col,youtube_start_row)
    #         end_col = LookupsGenerator(wb,"Youtube",sc_message,sc_message_obj,"message",end_col,youtube_start_row)
    #     except:
    #         print("Youtube Lookup failed")

    #     # ---- Assets Ends ---------------------------------

    #     # return jsonify({
    #     #         "message": "Successfully uploaded",
    #     #         "status": 200
    #     # }),200
    #     return send_file(pickle_file_path, as_attachment=True)
        # ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
        
      
        
        # Create a unique dataframe with all the column names are matching 
        # Then sum all the dataframes and generate in excel

        # Function to change all the columns of the dataframe according to the universal columns 
        
        
        # print(universal_dataframe)

    # -------------------------------------------------------------------






# # Product wise data (No Language / Mixed Images and videos) ----
@app.route("/product_wise",methods=["POST"])
def product_wise():
    
    # Fetch the raw data file (having all the platforms separated in each tab) 
    try:
        raw_data = request.files["raw_data"]
        start_date = request.form['start_date']
        end_date = request.form['end_date']

        facebook_instagram = pd.read_excel(raw_data,sheet_name="FB-IG") 
        snapchat = pd.read_excel(raw_data,sheet_name="Snapchat") 
        gdn = pd.read_excel(raw_data,sheet_name="GDN - Google Ads") 
        criteo = pd.read_excel(raw_data,sheet_name="Criteo") 
        youtube = pd.read_excel(raw_data,sheet_name="YouTube") 
        
    except:
        print("Raw Data skipped---")
    # Go to each tab, fetch the columns and merge them together 

    # fb_ig = set_columns(facebook_instagram)

    # Separate them market wise
    # Then in each market (Country), separate them using platform wise
    # Save the weeklydata, dailytotal and summary (for each column in a object for each platform using date)

    # Then separate platform wise data using product wise 


   


    try:
        fb_ig_obj = MatchingColumn(facebook_instagram)      # Compare each column and if the data is matching with any of the list attribute in the class it will save it there so we don't have to call facebook dataframe according to the static name 
        #Storing column names in the class attributes so we dont memorize the names everytime. 
        snapchat_obj = MatchingColumn(snapchat)
        gdn_obj = MatchingColumn(gdn)
        criteo_obj = MatchingColumn(criteo)
        youtube_obj = MatchingColumn(youtube)

    except:
        print("Issue with the matching column")
    
    

    


    
    try:
     
        facebook_instagram[fb_ig_obj.date] = pd.to_datetime(facebook_instagram[fb_ig_obj.date])
        facebook_instagram.sort_values(by=fb_ig_obj.date, inplace=True)
        # Sorting the data according to date 
        facebook_instagram[fb_ig_obj.date] = facebook_instagram[fb_ig_obj.date].dt.strftime('%Y-%m-%d')
        # Converting the data back to string to store in the dataframe
        
        snapchat[snapchat_obj.date] = pd.to_datetime(snapchat[snapchat_obj.date])
        snapchat.sort_values(by=snapchat_obj.date, inplace=True)
        # Sorting the data according to date 
        snapchat[snapchat_obj.date] = snapchat[snapchat_obj.date].dt.strftime('%Y-%m-%d')
        
        gdn[gdn_obj.date] = pd.to_datetime(gdn[gdn_obj.date])
        gdn.sort_values(by=gdn_obj.date, inplace=True)
        # Sorting the data according to date 
        gdn[gdn_obj.date] = gdn[gdn_obj.date].dt.strftime('%Y-%m-%d')
        criteo[criteo_obj.date] = pd.to_datetime(criteo[criteo_obj.date])
        criteo.sort_values(by=criteo_obj.date, inplace=True)
        # Sorting the data according to date 
        criteo[criteo_obj.date] = criteo[criteo_obj.date].dt.strftime('%Y-%m-%d')
        youtube[youtube_obj.date] = pd.to_datetime(youtube[youtube_obj.date])
        youtube.sort_values(by=youtube_obj.date, inplace=True)
        # Sorting the data according to date 
        youtube[youtube_obj.date] = youtube[youtube_obj.date].dt.strftime('%Y-%m-%d')

        # print(snapchat_obj)
    except:
        print("Issue in Date")
    # try:


    # Create a dataframe 
    universal_df_fb_ig = pd.DataFrame()
    list_of_columns = fb_ig_obj.ListOfColumns()
    for col in list_of_columns:
        # print(col)
        universal_df_fb_ig[col] = facebook_instagram[getattr(fb_ig_obj, col)]
        # Store those columns which are needed and save it in meaningful variable
    # print("Countries:",universal_df_fb_ig['country'].unique())
    universal_df_sc = pd.DataFrame()
    # list_of_columns = snapchat_obj.ListOfColumns()
    for col in list_of_columns:
        # print(col)
        universal_df_sc[col] = snapchat[getattr(snapchat_obj, col)]
        # Store those columns which are needed and save it in meaningful variable
    # print(universal_df_sc)
    universal_df_gdn = pd.DataFrame()
    # list_of_columns = gdn_obj.ListOfColumns()
    for col in list_of_columns:
        if col in ["view","percent_25","percent_50","percent_75","percent_100"]:
            universal_df_gdn[col] = 0
        else:
            universal_df_gdn[col] = gdn[getattr(gdn_obj, col)]
        # Store those columns which are needed and save it in meaningful variable
    # print(universal_df_gdn)
    universal_df_criteo = pd.DataFrame()
    # list_of_columns = criteo_obj.ListOfColumns()
    for col in list_of_columns:
        if col in ["view","percent_25","percent_50","percent_75","percent_100"]:
            universal_df_criteo[col] = 0
        else:
            universal_df_criteo[col] = criteo[getattr(criteo_obj, col)]
        # Store those columns which are needed and save it in meaningful variable
    # print(universal_df_criteo)
    universal_df_youtube = pd.DataFrame()
    # list_of_columns = youtube_obj.ListOfColumns()
    for col in list_of_columns:
        
        universal_df_youtube[col] = youtube[getattr(youtube_obj, col)]
        # Store those columns which are needed and save it in meaningful variable
    # print(universal_df_youtube)
    # Add all datafram together
    universal_df = pd.concat([universal_df_fb_ig, universal_df_sc, universal_df_gdn, universal_df_criteo,universal_df_youtube], ignore_index=True)
    # print(universal_df)
    # except:
    #     print("Issue in universal")
        # # testDf = pd.DataFrame(columns=fb_ig_obj.l)
    
    
    universal_df # Dataframe with data
    list_of_columns # List of columns for dataframe 

    # Separate them country wise ---------------------------------------------------------------------------

    # Country names 
    list_of_countries = universal_df['country'].unique()
    list_of_dataframes_by_country = []

    for country in list_of_countries:
        filtered_condition = universal_df['country'].str.contains(country, case=False)    
        list_of_dataframes_by_country.append({"country":country,"data": universal_df[filtered_condition]})
        # print(f"country: {country}")
        # print(universal_df[filtered_condition])

    # print(list_of_dataframes_by_country)
    # --------------------------------------------------------------------------------------------------------
    list_of_columns_for_platform = fb_ig_obj.ListOfColumnsForPlatform()
    # print(list_of_columns_for_platform)
    list_of_dataframes_eachcountry_by_platform = []
    list_of_dataframes_eachcountry_by_platform_by_product=[]

    for country_info in list_of_dataframes_by_country:
        countryName= country_info["country"]    # Country name
        eachCountryDf = country_info["data"]       # Country dataframe

        # Create a new dataframe with only platform columns 

        
        # For platform only ------------------------------------------------------
        list_of_platforms = eachCountryDf['platform'].unique()
        # Get unique list of platforms 
        youtube_check = True
        # Youtube check True if the record has multiple youtube 
        for platform in list_of_platforms:
            # getting each platform 
            if "YouTube".lower() in (platform).lower() and youtube_check == True:
                # If the platform name has youtube then just count one and skip all other names of youtube so we can unite all youtube data in one
                filtered_condition = eachCountryDf['platform'].str.contains("YouTube", case=False)
                # Filter condition if the platform value has youtube then get the data
                new_df = eachCountryDf[filtered_condition]
                # new dataframe has all the records of the youtube 
                eachPlatformDf = pd.DataFrame(columns=list_of_columns_for_platform)
                # creating a new dataframe and passing list of columns for platform 
                for col in list_of_columns_for_platform:
                    # going to each column and storing it to new dataframe and the new dataframe has only specific columns and not country language and platform records
                    eachPlatformDf[col] = new_df[col]
  
                list_of_dataframes_eachcountry_by_platform.append({"country":countryName ,"platform":"YouTube","data": eachPlatformDf })
                youtube_check = False  
            elif "YouTube".lower() in (platform).lower() and youtube_check == False:
                continue
            else:
                filtered_condition = eachCountryDf['platform'].str.contains(platform, case=False)
                new_df = eachCountryDf[filtered_condition]
                
                

                # print(new_df.columns)
                eachPlatformDf = pd.DataFrame(columns=list_of_columns_for_platform)
                for col in list_of_columns_for_platform:
                    # print(col)
                    eachPlatformDf[col] = new_df[col]

                eachPlatformDf      # This is the dataframe for each platform 


               

                list_of_dataframes_eachcountry_by_platform.append({"country":countryName ,"platform":platform,"data": eachPlatformDf})
                
        # For platform only -------------------------------------------------------
        
        # For assets only 
        youtube_check = True

        for platform in list_of_platforms:
            # getting each platform 
            if "YouTube".lower() in (platform).lower() and youtube_check == True:
                # If the platform name has youtube then just count one and skip all other names of youtube so we can unite all youtube data in one
                filtered_condition = eachCountryDf['platform'].str.contains("YouTube", case=False)
                # Filter condition if the platform value has youtube then get the data
                new_df_product = eachCountryDf[filtered_condition]
                # List of assets of that particular dataframe platform wise 
                list_of_products = new_df_product['product'].unique()

                for prod in list_of_products:
                    filtered_condition_products = new_df_product['product'].str.contains(prod, case=False)
                    new_df_prod = new_df_product[filtered_condition_products]

                    # New dataframe for getting the data only for that particular product 
                    eachPlatformProductDf = pd.DataFrame(columns=list_of_columns_for_platform)
                    for col in list_of_columns_for_platform:
                        eachPlatformProductDf[col] = new_df_prod[col]
                    
                    list_of_dataframes_eachcountry_by_platform_by_product.append({"country":countryName ,"platform":"YouTube","product": prod,"data": eachPlatformProductDf })
                
                youtube_check = False  

                    # --------------------------------------------------------------------
             
            elif "YouTube".lower() in (platform).lower() and youtube_check == False:
                pass
            else:
                filtered_condition = eachCountryDf['platform'].str.contains(platform, case=False)
                new_df_product = eachCountryDf[filtered_condition]
                # List of assets of that particular dataframe platform wise 
                list_of_products = new_df_product['product'].unique()

                for prod in list_of_products:
                    filtered_condition_products = new_df_product['product'].str.contains(prod, case=False)
                    new_df_prod = new_df_product[filtered_condition_products]

                    # New dataframe for getting the data only for that particular product 
                    eachPlatformProductDf = pd.DataFrame(columns=list_of_columns_for_platform)
                    for col in list_of_columns_for_platform:
                        eachPlatformProductDf[col] = new_df_prod[col]
                    # print(countryName)
                    list_of_dataframes_eachcountry_by_platform_by_product.append({"country":countryName ,"platform":platform,"product": prod,"data": eachPlatformProductDf })
    # print("Total Dataframes:",len(list_of_dataframes_eachcountry_by_platform_by_product))
        # -----------------------------------------------------------------------
    
    # Then in each market (Country), separate them using platform wise
    
    # print(list_of_dataframes_eachcountry_by_platform)
    
    # Loop through list of countries 
    
    for entry in list_of_dataframes_eachcountry_by_platform:
        country = entry['country']
        platform = entry['platform']
        data_frame = entry['data']
        

        # print(platform)


        # Create or open the Excel file for the country
        file_name = f'{country}_data.xlsx'
        start_date=pd.to_datetime(start_date)


        end_date=pd.to_datetime(end_date)
   
        
       
        temp_start_date = start_date
        temp_end_date = end_date
        
        data_values = []
        while temp_start_date <= temp_end_date:
            # Your logic ....
            # Add row with date 
            
            new_data_temp_start_date = [temp_start_date.strftime('%Y-%m-%d')]
            # Created new array with end date which will be incremented one by one  
            for column in list_of_columns_for_platform:
                # Going to each column and if the column is not date then add 0 to the array
                if column != 'date':
                    new_data_temp_start_date.append(0)
      
            
            data_values.append(new_data_temp_start_date)
  
            temp_start_date = temp_start_date + timedelta(days=1)
        
        newDF= pd.DataFrame(data_values, columns=list_of_columns_for_platform)
        data_frame = pd.concat([data_frame,newDF ], ignore_index=True)

        data_frame.sort_values(by='date', inplace=True)
        data_frame = data_frame.groupby(["date"]).sum()
        
        


        # Add extra columns here and show the data
        list_of_formula_columns = ['Budget Spent', 'CTR','CPC','CPV'] 
        
        i=1
        for _, row in data_frame.iterrows():
            row['CTR'] = f'=(D{i}/C{i})'

        # ----------------------------------------
        df_weekly = CalculateWeekly(data_frame,list_of_columns_for_platform)


        # printing_daily_total = data_frame.copy()  # For printing daily + weekly records

        # print(['date'] + data_frame.columns.append())
        columns_ = data_frame.columns.tolist()

        # Add a new column name to the beginning of the list
        new_column_name = 'date'
        columns_.insert(0, new_column_name)
        # print(columns_)
        print_daily_total_df = pd.DataFrame(columns=columns_)
        # for _, row in df_weekly.iterrows():
        #     values_list = row.values.tolist()
            
        #     print(values_list)
            # This is row of week


        records_added = 0

        # Define the number of records after which to insert a values_list
        insert_every = 7

        weeklyTable = []
        # Iterate through rows and create DataFrames for each values_list and separator row
        for _, row in df_weekly.iterrows():
            eachWeek = row.values.tolist() # This is week 1 
            weeklyTable.append([f"{int(_)+1} Weekly Total"]+eachWeek)

        records_added = 0
        weekly_ind=0
        for _, row in data_frame.iterrows():
            eachEntry = row.values.tolist() # This is dailyEntry
            print_daily_total_df.loc[len(print_daily_total_df)] = [_]+eachEntry
            records_added +=1
            if records_added % insert_every == 0:
                print_daily_total_df.loc[len(print_daily_total_df)] = weeklyTable[weekly_ind] 
                weekly_ind+=1
               
           


        # print(printing_daily_total)
        summary_total = df_weekly.sum().to_frame().T
        
        # print(summary_total)
        excel_file_path = f'{country}.xlsx'
        if os.path.exists(excel_file_path):
            # If the file exists, load it and add a new sheet with the DataFrame data
            # print("Check if it is running")
            wb = load_workbook(excel_file_path)
            

        else:
            wb = Workbook()
      
        ws = wb.create_sheet(platform)
        wb.active = ws
        wb,fb_start_row,start_row = media_calculation(wb,ws,1,fb_ig_obj,print_daily_total_df,df_weekly,summary_total)
        
        wb.save(excel_file_path)
        
    # {
    #     country: UAE,
    #     data: {"facebook": 13,"youtube": 13}
    # }
    countryListWithAssets = []
    for countryName in list_of_countries:
        temp = {}
        youtube_ = False
        temp["country"]= countryName
        for plat in list_of_platforms:
            if youtube_ == False and "YouTube" in plat:
                temp["YouTube"] = 13
                youtube_ = True
            elif youtube_ == True:
                pass
            else:        
                temp[plat] = 13
        countryListWithAssets.append(temp)
    
    # print(len(list_of_dataframes_eachcountry_by_platform_by_product))
    # Assets tables 
    # print(len(list_of_dataframes_eachcountry_by_platform_by_product))
    for entry in list_of_dataframes_eachcountry_by_platform_by_product:
        country = entry['country']
        platform = entry['platform']
        product = entry['product']
        data_frame_product = entry['data']
        
        print(f"{country}:",product)
        # print(platform)


        # Create or open the Excel file for the country
        file_name = f'{country}_data.xlsx'
        # start_date=pd.to_datetime(start_date)


        # end_date=pd.to_datetime(end_date)
   
        
       
        temp_start_date = start_date
        temp_end_date = end_date
        
        data_values = []
        while temp_start_date <= temp_end_date:
            # Your logic ....
            # Add row with date 
            
            new_data_temp_start_date = [temp_start_date.strftime('%Y-%m-%d')]
            # Created new array with end date which will be incremented one by one  
            for column in list_of_columns_for_platform:
                # Going to each column and if the column is not date then add 0 to the array
                if column != 'date':
                    new_data_temp_start_date.append(0)
      
            
            data_values.append(new_data_temp_start_date)
  
            temp_start_date = temp_start_date + timedelta(days=1)
        
        newDF= pd.DataFrame(data_values, columns=list_of_columns_for_platform)
        data_frame_product = pd.concat([data_frame_product,newDF ], ignore_index=True)

        data_frame_product.sort_values(by='date', inplace=True)
        data_frame_product = data_frame_product.groupby(["date"]).sum()
        
        # print(data_frame_product)


        # Add extra columns here and show the data
        list_of_formula_columns = ['Budget Spent', 'CTR','CPC','CPV'] 
        
        i=1
        for _, row in data_frame_product.iterrows():
            row['CTR'] = f'=(D{i}/C{i})'

        # ----------------------------------------
        df_weekly_assets = CalculateWeekly(data_frame_product,list_of_columns_for_platform)


        # printing_daily_total = data_frame_product.copy()  # For printing daily + weekly records

        # print(['date'] + data_frame_product.columns.append())
        columns_ = data_frame_product.columns.tolist()

        # Add a new column name to the beginning of the list
        new_column_name = 'date'
        columns_.insert(0, new_column_name)
        # print(columns_)
        print_daily_total_df_assets = pd.DataFrame(columns=columns_)
      


        records_added = 0

        # Define the number of records after which to insert a values_list
        insert_every = 7

        weeklyTable = []
        # Iterate through rows and create DataFrames for each values_list and separator row
        for _, row in df_weekly_assets.iterrows():
            eachWeek = row.values.tolist() # This is week 1 
            weeklyTable.append([f"{int(_)+1} Weekly Total"]+eachWeek)

        records_added = 0
        weekly_ind=0
        for _, row in data_frame_product.iterrows():
            eachEntry = row.values.tolist() # This is dailyEntry
            print_daily_total_df_assets.loc[len(print_daily_total_df_assets)] = [_]+eachEntry
            records_added +=1
            if records_added % insert_every == 0:
                print_daily_total_df_assets.loc[len(print_daily_total_df_assets)] = weeklyTable[weekly_ind] 
                weekly_ind+=1
               
           


        # print(print_daily_total_df_assets)
        summary_total_assets = df_weekly_assets.sum().to_frame().T
        
        # print(summary_total_assets)
        excel_file_path = f'{country}.xlsx'
        # if os.path.exists(excel_file_path):
            # If the file exists, load it and add a new sheet with the DataFrame data
            # print("Check if it is running")
        wb = load_workbook(excel_file_path)
        ws = wb[platform]

        # else:
        #     wb = Workbook()
        #     ws = wb.create_sheet(platform)
        wb.active = ws
        # print(countryListWithAssets)
        for list_of_coun in countryListWithAssets:
            if country == list_of_coun["country"]:
                new_column = list_of_coun[platform]
                break

        print(f"{country}: {platform} : {new_column}")
        wb,new_record_column = media_calculation_assets(wb,ws,1,fb_ig_obj,print_daily_total_df_assets,df_weekly_assets,summary_total_assets,start_row,new_column)
        for list_of_coun in countryListWithAssets:
            if country == list_of_coun["country"]:
                list_of_coun[platform] = new_record_column
                break
        # print(countryListWithAssets)
        # list_of_platforms_for_assets[platform] = new_column
        # print(country)
        wb.save(excel_file_path)



    
    # Loop here for the dataframe eachcountry_by_platform_by asset
    # print(list_of_dataframes_eachcountry_by_platform_by_product)            

    # ============================================================


    return jsonify({
        "message": "completed"
    }),200
        # try:    
        #     fb_asset[fb_asset_obj.date] = pd.to_datetime(fb_asset[fb_asset_obj.date])

        #     fb_asset.sort_values(by=fb_asset_obj.date, inplace=True)
        #     fb_asset[fb_asset_obj.date] = fb_asset[fb_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Assets dates issues")

        # try:    
        #     fb_size[fb_size_obj.date] = pd.to_datetime(fb_size[fb_size_obj.date])

        #     fb_size.sort_values(by=fb_size_obj.date, inplace=True)
        #     fb_size[fb_size_obj.date] = fb_size[fb_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Size dates issues")

        # try:    
        #     fb_language[fb_language_obj.date] = pd.to_datetime(fb_language[fb_language_obj.date])

        #     fb_language.sort_values(by=fb_language_obj.date, inplace=True)
        #     fb_language[fb_language_obj.date] = fb_language[fb_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Language dates issues")

        # try:    
        #     fb_country[fb_country_obj.date] = pd.to_datetime(fb_country[fb_country_obj.date])

        #     fb_country.sort_values(by=fb_country_obj.date, inplace=True)
        #     fb_country[fb_country_obj.date] = fb_country[fb_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Country dates issues")

        # try:    
        #     fb_message[fb_message_obj.date] = pd.to_datetime(fb_message[fb_message_obj.date])

        #     fb_message.sort_values(by=fb_message_obj.date, inplace=True)
        #     fb_message[fb_message_obj.date] = fb_message[fb_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Message dates issues")
       

        # try:
        #     ig[ig_obj.date] = pd.to_datetime(ig[ig_obj.date])
        #     ig.sort_values(by=ig_obj.date, inplace=True)
        #     ig[ig_obj.date] = ig[ig_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("ig Dates issues")

        # try:    
        #     ig_asset[ig_asset_obj.date] = pd.to_datetime(ig_asset[ig_asset_obj.date])

        #     ig_asset.sort_values(by=ig_asset_obj.date, inplace=True)
        #     ig_asset[ig_asset_obj.date] = ig_asset[ig_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Assets dates issues")

        # try:    
        #     ig_size[ig_size_obj.date] = pd.to_datetime(ig_size[ig_size_obj.date])

        #     ig_size.sort_values(by=ig_size_obj.date, inplace=True)
        #     ig_size[ig_size_obj.date] = ig_size[ig_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Size dates issues")

        # try:    
        #     ig_language[ig_language_obj.date] = pd.to_datetime(ig_language[ig_language_obj.date])

        #     ig_language.sort_values(by=ig_language_obj.date, inplace=True)
        #     ig_language[ig_language_obj.date] = ig_language[ig_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Language dates issues")

        # try:    
        #     ig_country[ig_country_obj.date] = pd.to_datetime(ig_country[ig_country_obj.date])

        #     ig_country.sort_values(by=ig_country_obj.date, inplace=True)
        #     ig_country[ig_country_obj.date] = ig_country[ig_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Country dates issues")

        # try:    
        #     ig_message[ig_message_obj.date] = pd.to_datetime(ig_message[ig_message_obj.date])

        #     ig_message.sort_values(by=ig_message_obj.date, inplace=True)
        #     ig_message[ig_message_obj.date] = ig_message[ig_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Message dates issues")




        # try:
        #     sc[sc_obj.date] = pd.to_datetime(sc[sc_obj.date])
        #     rawData = sc
        #     sc.sort_values(by=sc_obj.date, inplace=True)
        #     sc[sc_obj.date] = sc[sc_obj.date].dt.strftime('%Y-%m-%d')

        #     rawData[sc_obj.date] = pd.to_datetime(sc[sc_obj.date])
        # except:
        #     print("sc Dates issues")

        # try:    
        #     sc_asset[sc_asset_obj.date] = pd.to_datetime(sc_asset[sc_asset_obj.date])

        #     sc_asset.sort_values(by=sc_asset_obj.date, inplace=True)
        #     sc_asset[sc_asset_obj.date] = sc_asset[sc_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Assets dates issues")

        # try:    
        #     sc_size[sc_size_obj.date] = pd.to_datetime(sc_size[sc_size_obj.date])

        #     sc_size.sort_values(by=sc_size_obj.date, inplace=True)
        #     sc_size[sc_size_obj.date] = sc_size[sc_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Size dates issues")

        # try:    
        #     sc_language[sc_language_obj.date] = pd.to_datetime(sc_language[sc_language_obj.date])

        #     sc_language.sort_values(by=sc_language_obj.date, inplace=True)
        #     sc_language[sc_language_obj.date] = sc_language[sc_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Language dates issues")

        # try:    
        #     sc_country[sc_country_obj.date] = pd.to_datetime(sc_country[sc_country_obj.date])

        #     sc_country.sort_values(by=sc_country_obj.date, inplace=True)
        #     sc_country[sc_country_obj.date] = sc_country[sc_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Country dates issues")

        # try:    
        #     sc_message[sc_message_obj.date] = pd.to_datetime(sc_message[sc_message_obj.date])

        #     sc_message.sort_values(by=sc_message_obj.date, inplace=True)
        #     sc_message[sc_message_obj.date] = sc_message[sc_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Message dates issues")

      
        

        # # Dv 360 -------------------------------------------
        # try:
        #     dv[dv_obj.date] = pd.to_datetime(dv[dv_obj.date])
        #     dv.sort_values(by=dv_obj.date, inplace=True)
        #     dv[dv_obj.date] = dv[ig_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("dv Dates issues")

        # try:    
        #     dv_asset[dv_asset_obj.date] = pd.to_datetime(dv_asset[dv_asset_obj.date])

        #     dv_asset.sort_values(by=dv_asset_obj.date, inplace=True)
        #     dv_asset[dv_asset_obj.date] = dv_asset[dv_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Assets dates issues")

        # try:    
        #     dv_size[dv_size_obj.date] = pd.to_datetime(dv_size[dv_size_obj.date])

        #     dv_size.sort_values(by=dv_size_obj.date, inplace=True)
        #     dv_size[dv_size_obj.date] = dv_size[dv_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Size dates issues")

        # try:    
        #     dv_language[dv_language_obj.date] = pd.to_datetime(dv_language[dv_language_obj.date])

        #     dv_language.sort_values(by=dv_language_obj.date, inplace=True)
        #     dv_language[dv_language_obj.date] = dv_language[dv_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Language dates issues")

        # try:    
        #     dv_country[dv_country_obj.date] = pd.to_datetime(dv_country[dv_country_obj.date])

        #     dv_country.sort_values(by=dv_country_obj.date, inplace=True)
        #     dv_country[dv_country_obj.date] = dv_country[dv_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Country dates issues")

        # try:    
        #     dv_message[dv_message_obj.date] = pd.to_datetime(dv_message[dv_message_obj.date])

        #     dv_message.sort_values(by=dv_message_obj.date, inplace=True)
        #     dv_message[dv_message_obj.date] = dv_message[dv_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Message dates issues")

        # --------------------------------------------------------
        
        

       
        # Formatting dates and sorting the values according to dates 
        
        
        
        # Creating path for the final output file 

        
        # ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
    #     try:
    #         fb = facebook_instagram.groupby([fb_obj.date]).sum()

    #     except:
    #         print("fb sum issues")
        
    #     try: 
    #         ig = ig.groupby([ig_obj.date]).sum()
    #     except:
    #         print("ig sum issues")
        
    #     try: 
    #         sc = sc.groupby([sc_obj.date]).sum()
    #     except:
    #         print("sc sum issues")
    #     try: 
    #         dv = dv.groupby([dv_obj.date]).sum()
    #     except:
    #         print("dv sum issues")
        

    #     # ----------------------------------------------
    #     try:
    #         fb_weekly = CalculateWeekly(fb,fb_obj)
    #     except:
    #         print("fb_weekly issue")
    #     try:
    #         ig_weekly = CalculateWeekly(ig,ig_obj)
    #     except:
    #         print("ig_weekly issue")
    #     try:
    #         sc_weekly = CalculateWeekly(sc,sc_obj)
    #     except:
    #         print("sc_weekly issue")
    #     try:
    #         dv_weekly = CalculateWeekly(dv,dv_obj)
    #     except:
    #         print("dv_weekly issue")
        
       
    #     # =================================================
    #     try:
    #         fb_total = fb_weekly.sum()
    #     except:
    #         print("fb_total issue")
    #     try:
    #         ig_total = ig_weekly.sum()
    #     except:
    #         print("ig_total issue")
    #     try:
    #         sc_total = sc_weekly.sum()
    #     except:
    #         print("sc_total issue")
    #     try:
    #         dv_total = dv_weekly.sum()
    #     except:
    #         print("dv_total issue")
   
        
    #     # Summing the data according to the date so we have one data for each data
        
    #     # Until here everything is fine 
    #     # Instead of setting columns using dataframe we will set them in excel anad update the excel sheet according to row and column 
    #     wb = Workbook()
    #     wb.remove(wb.active)
    #     try:
    #         ws = wb.create_sheet('Facebook')
    #         wb.active = ws
    #         fb_start_row = media_calculation(wb,ws,1,fb_obj,fb,fb_weekly,fb_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Facebook sheet skipped----")

    #     try:
    #         ws = wb.create_sheet('Instagram')
    #         wb.active = ws
    #         ig_start_row = media_calculation(wb,ws,1,ig_obj,ig,ig_weekly,ig_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Ig sheet skipped----")

    #     try:
    #         ws = wb.create_sheet('Youtube')
    #         wb.active = ws
    #         youtube_start_row = media_calculation(wb,ws,1,sc_obj,sc,sc_weekly,sc_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Youtube sheet skipped---")
        
    #     try:
    #         ws = wb.create_sheet('GDN')
    #         wb.active = ws
    #         gdn_start_row = media_calculation(wb,ws,1,dv_obj,dv,dv_weekly,dv_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("GDN sheet skipped---")

       

    #     # Sheet is active ====
    #     daily_total = "="
    #     concat_arr = []
        
    #     try:
    #         universal_fb = SettingUniversalDataframe(fb,fb_obj)
    #         if daily_total == "=":
    #             daily_total+="'Facebook'!T2"
    #         else:
    #             daily_total+="+'Facebook'!T2"
    #         concat_arr.append(universal_fb)
    #     except:
    #         print("Universal Fb skipped----")
        
    #     try:
    #     # print(universal_fb["view"])
    #         universal_ig = SettingUniversalDataframe(ig,ig_obj)
    #         if daily_total == "=":
    #             daily_total+="'Instagram'!T2"
    #         else:
    #             daily_total+="+'Instagram'!T2"
            
    #         concat_arr.append(universal_ig)
    #     except:
    #         print("Universal ig skipped -----")

    #     try:
    #         universal_sc = SettingUniversalDataframe(sc,sc_obj)
    #         if daily_total == "=":
    #             daily_total+="'Youtube'!T2"
    #         else:
    #             daily_total+="+'Youtube'!T2"
    #         concat_arr.append(universal_sc)
    #     except:
    #         print("Universal Youtube skipped ------")
    #     # universal_tw = SettingUniversalDataframe(tw,tw_obj)
    #     # universal_td = SettingUniversalDataframe(td,td_obj)
        
    #     try:
    #         universal_dv = SettingUniversalDataframe(dv,dv_obj)
    #         if daily_total == "=":
    #             daily_total+="'GDN'!T2"
    #         else:
    #             daily_total+="+'GDN'!T2"
    #         concat_arr.append(universal_dv)
    #     except:
    #         print("Universal GDN skipped----")

       

    #     universal_dataframe = pd.concat(concat_arr).groupby(["date"]).sum()
        
        
    #     universal_obj = MatchingColumn(universal_dataframe)
        
    #     universal_dataframe_weekly = CalculateWeekly(universal_dataframe,universal_obj)
    #     universal_dataframe_total = universal_dataframe_weekly.sum()
    #     ws = wb.create_sheet('Daily')
    #     wb.active = ws
    #     media_calculation(wb,ws,1,universal_obj,universal_dataframe,universal_dataframe_weekly,universal_dataframe_total,daily=True)
    #     ws.cell(row=2,column=20,value=daily_total)
    #     wb.save('summary_testing.xlsx')

    #     # Generating summary ----------------------

    #     have_column = True
    #     ws = wb.create_sheet('Summary')
    #     wb.active = ws
    #     daily_media_index=2
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Facebook",fb_obj,fb_total,have_column=have_column,media_label="Facebook")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Facebook summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Instagram",ig_obj,ig_total,have_column=have_column,media_label="Instagram")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Instagram summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Youtube",sc_obj,sc_total,have_column=have_column,media_label="Youtube")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Snapchat summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"GDN",dv_obj,dv_total,have_column=have_column,media_label="GDN")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("GDN summary is missing")
        
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Daily",universal_obj,universal_dataframe_total,have_column=have_column,media_label="Total")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Daily summary is missing")
    #     # engine.dispose()
    #     wb.save(pickle_file_path)

    #     # Calculating Asset part



    # # --------------------------------------------------------------------
    #     try:
    #         end_col = LookupsGenerator(wb,"Facebook",fb_asset,fb_asset_obj,"asset_type",13,fb_start_row)
    #         end_col = LookupsGenerator(wb,"Facebook",fb_language,fb_language_obj,"language",end_col,fb_start_row)
    #         # end_col = LookupsGenerator(wb,"Facebook",fb_country,fb_country_obj,"country",end_col,fb_start_row)
    #         end_col = LookupsGenerator(wb,"Facebook",fb_message,fb_message_obj,"message",end_col,fb_start_row)
    #     except:
    #         print("Facebook Lookup failed")
    #     try:
    #         end_col = LookupsGenerator(wb,"Instagram",ig_asset,ig_asset_obj,"asset_type",13,ig_start_row)
    #         end_col = LookupsGenerator(wb,"Instagram",ig_language,ig_language_obj,"language",end_col,ig_start_row)
    #         # end_col = LookupsGenerator(wb,"Instagram",ig_country,ig_country_obj,"country",end_col,ig_start_row)
    #         end_col = LookupsGenerator(wb,"Instagram",ig_message,ig_message_obj,"message",end_col,ig_start_row)
    #     except:
    #         print("Ig lookup failed")
        
    #     try:
    #         end_col = LookupsGenerator(wb,"GDN",dv_asset,dv_asset_obj,"asset_type",13,gdn_start_row)
    #         end_col = LookupsGenerator(wb,"GDN",dv_language,dv_language_obj,"language",end_col,gdn_start_row)
    #         # end_col = LookupsGenerator(wb,"GDN",dv_country,dv_country_obj,"country",end_col,gdn_start_row)
    #         end_col = LookupsGenerator(wb,"GDN",dv_message,dv_message_obj,"message",end_col,gdn_start_row)
    #     except:
    #         print("GDN Lookup failed")

    #     try:
    #         end_col = LookupsGenerator(wb,"Youtube",sc_asset,sc_asset_obj,"asset_type",13,youtube_start_row)
    #         end_col = LookupsGenerator(wb,"Youtube",sc_language,sc_language_obj,"language",end_col,youtube_start_row)
    #         # end_col = LookupsGenerator(wb,"Youtube",sc_country,sc_country_obj,"country",end_col,youtube_start_row)
    #         end_col = LookupsGenerator(wb,"Youtube",sc_message,sc_message_obj,"message",end_col,youtube_start_row)
    #     except:
    #         print("Youtube Lookup failed")

    #     # ---- Assets Ends ---------------------------------

    #     # return jsonify({
    #     #         "message": "Successfully uploaded",
    #     #         "status": 200
    #     # }),200
    #     return send_file(pickle_file_path, as_attachment=True)
        # ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
        
      
        
        # Create a unique dataframe with all the column names are matching 
        # Then sum all the dataframes and generate in excel

        # Function to change all the columns of the dataframe according to the universal columns 
        
        
        # print(universal_dataframe)

    # -------------------------------------------------------------------



# # Language wise data (Mixed Images and videos, separate product for each language)
@app.route("/language_wise",methods=["POST"])
def language_wise():
    
    # Fetch the raw data file (having all the platforms separated in each tab) 
    try:
        raw_data = request.files["raw_data"]
        start_date = request.form['start_date']
        end_date = request.form['end_date']
    except:
        print("Raw Data skipped---")

    try:
        facebook_instagram = pd.read_excel(raw_data,sheet_name="FB-IG") 
    except:
        print("facebook skipped----")

    try:
        snapchat = pd.read_excel(raw_data,sheet_name="Snapchat") 
    except:
        print("snapchat skipped--")
    try:
        gdn = pd.read_excel(raw_data,sheet_name="GDN - Google Ads") 
    except:
        print("snapchat skipped--")
    try:
        criteo = pd.read_excel(raw_data,sheet_name="Criteo")
    except:
        print("criteo skipped--")
        # print(criteo) 
    try:
        youtube = pd.read_excel(raw_data,sheet_name="YouTube") 
    except:
        print("youtube skipped--")
    
    # Go to each tab, fetch the columns and merge them together 

    # fb_ig = set_columns(facebook_instagram)

    # Separate them market wise
    # Then in each market (Country), separate them using platform wise
    # Save the weeklydata, dailytotal and summary (for each column in a object for each platform using date)

    # Then separate platform wise data using product wise 


   


    try:
        fb_ig_obj = MatchingColumn(facebook_instagram)      # Compare each column and if the data is matching with any of the list attribute in the class it will save it there so we don't have to call facebook dataframe according to the static name 
    except:
        print("Issue with the fb_ig_obj")    #Storing column names in the class attributes so we dont memorize the names everytime. 
    try:
        snapchat_obj = MatchingColumn(snapchat)
    except:
        print("Issue with the snapchat_obj")    
    try:
        gdn_obj = MatchingColumn(gdn)
    except:
        print("Issue with the gdn_obj")    
    try:
        criteo_obj = MatchingColumn(criteo)
    except:
        print("Issue with the criteo_obj")    
    try:
        youtube_obj = MatchingColumn(youtube)
    except:
        print("Issue with the youtube_obj")
        
        # print(youtube_obj.ListOfColumns())

    
    
    

    


    
    try:
     
        facebook_instagram[fb_ig_obj.date] = pd.to_datetime(facebook_instagram[fb_ig_obj.date])
        facebook_instagram.sort_values(by=fb_ig_obj.date, inplace=True)
        # Sorting the data according to date 
        facebook_instagram[fb_ig_obj.date] = facebook_instagram[fb_ig_obj.date].dt.strftime('%Y-%m-%d')
        # Converting the data back to string to store in the dataframe
    except:
        print("Facebook date skipped--")    
    try:
        snapchat[snapchat_obj.date] = pd.to_datetime(snapchat[snapchat_obj.date])
        snapchat.sort_values(by=snapchat_obj.date, inplace=True)
        # Sorting the data according to date 
        snapchat[snapchat_obj.date] = snapchat[snapchat_obj.date].dt.strftime('%Y-%m-%d')
    except:
        print("Snapchat date skipped--") 

    try:
        gdn[gdn_obj.date] = pd.to_datetime(gdn[gdn_obj.date])
        gdn.sort_values(by=gdn_obj.date, inplace=True)
        # Sorting the data according to date 
        gdn[gdn_obj.date] = gdn[gdn_obj.date].dt.strftime('%Y-%m-%d')
    except:
        print("GDN date skipped--") 
    try:
        criteo[criteo_obj.date] = pd.to_datetime(criteo[criteo_obj.date])
        criteo.sort_values(by=criteo_obj.date, inplace=True)
        # Sorting the data according to date 
        criteo[criteo_obj.date] = criteo[criteo_obj.date].dt.strftime('%Y-%m-%d')
    except:
        print("Criteo date skipped--")
    try:
        youtube[youtube_obj.date] = pd.to_datetime(youtube[youtube_obj.date])
        youtube.sort_values(by=youtube_obj.date, inplace=True)
        # Sorting the data according to date 
        youtube[youtube_obj.date] = youtube[youtube_obj.date].dt.strftime('%Y-%m-%d')

        # print(snapchat_obj)
    except:
        print("Youtube date skipped--")
    universal_df=pd.DataFrame()
    try:
        # print(youtube)

        # Create a dataframe 
        universal_df_fb_ig = pd.DataFrame()
        list_of_columns = fb_ig_obj.ListOfColumns()
        for col in list_of_columns:
            # print(col)
            universal_df_fb_ig[col] = facebook_instagram[getattr(fb_ig_obj, col)]
        universal_df = pd.concat([universal_df,universal_df_fb_ig], ignore_index=True)
    except:
        print("Facebook universal skipped--")

            # Store those columns which are needed and save it in meaningful variable
        # print(universal_df_fb_ig['country'].unique())
    try:    
        universal_df_sc = pd.DataFrame()
        # list_of_columns = snapchat_obj.ListOfColumns()
        for col in list_of_columns:
            # print(col)
            universal_df_sc[col] = snapchat[getattr(snapchat_obj, col)]
        universal_df = pd.concat([universal_df,universal_df_sc], ignore_index=True)
        
    except:
        print("Snapchat universal skipped--")
            # Store those columns which are needed and save it in meaningful variable
        # print(universal_df_sc)
        # print(universal_df_sc['country'].unique())
    try:
        universal_df_gdn = pd.DataFrame()
        # list_of_columns = gdn_obj.ListOfColumns()
        for col in list_of_columns:
            if col in ["view","percent_25","percent_50","percent_75","percent_100"]:
                universal_df_gdn[col] = 0
            else:
                universal_df_gdn[col] = gdn[getattr(gdn_obj, col)]
        universal_df = pd.concat([universal_df,universal_df_gdn], ignore_index=True)
    except:
        print("Gdn universal skipped--")

    try:
            # Store those columns which are needed and save it in meaningful variable
        # print(universal_df_gdn)
        # print(universal_df_gdn['country'].unique())
        universal_df_criteo = pd.DataFrame()
        # list_of_columns = criteo_obj.ListOfColumns()
        for col in list_of_columns:
            if col in ["view","percent_25","percent_50","percent_75","percent_100"]:
                universal_df_criteo[col] = 0
            else:
                universal_df_criteo[col] = criteo[getattr(criteo_obj, col)]
        universal_df = pd.concat([universal_df,universal_df_criteo], ignore_index=True)
    except:
        print("criteo universal skipped")
            # Store those columns which are needed and save it in meaningful variable
        # print(universal_df_criteo)
        # print(universal_df_criteo['country'].unique())
    try:
        universal_df_youtube = pd.DataFrame()
        # list_of_columns = youtube_obj.ListOfColumns()
        for col in list_of_columns:
            print(getattr(youtube_obj,col))
            universal_df_youtube[col] = youtube[getattr(youtube_obj, col)]
            # Store those columns which are needed and save it in meaningful variable
        # print(universal_df_youtube)
        # Add all datafram together
        universal_df = pd.concat([universal_df,universal_df_youtube], ignore_index=True)

        # print(universal_df)
    except:
        print("Youtube universal skipped--")
        # testDf = pd.DataFrame(columns=fb_ig_obj.l)
    
    
    universal_df # Dataframe with data
    list_of_columns # List of columns for dataframe 

    # Separate them country wise ---------------------------------------------------------------------------

    # Country names 
    list_of_countries = universal_df['country'].unique()
    print(list_of_countries)
    list_of_dataframes_by_country = []

    for country in list_of_countries:
        filtered_condition = universal_df['country'].str.contains(country, case=False)    
        list_of_dataframes_by_country.append({"country":country,"data": universal_df[filtered_condition]})
        # print(f"country: {country}")
        # print(universal_df[filtered_condition])

    # print(list_of_dataframes_by_country)
    # --------------------------------------------------------------------------------------------------------
    list_of_columns_for_platform = fb_ig_obj.ListOfColumnsForPlatform()
    # print(list_of_columns_for_platform)
    list_of_dataframes_eachcountry_by_platform = []
    list_of_dataframes_eachcountry_by_platform_by_language=[]
    list_of_dataframes_eachcountry_by_platform_by_language_by_product = []
    for country_info in list_of_dataframes_by_country:
        countryName= country_info["country"]    # Country name
        eachCountryDf = country_info["data"]       # Country dataframe

        # Create a new dataframe with only platform columns 

        
        # For platform only ------------------------------------------------------
        list_of_platforms = eachCountryDf['platform'].unique()
        # Get unique list of platforms 
        youtube_check = True
        # Youtube check True if the record has multiple youtube 
        for platform in list_of_platforms:
            # getting each platform 
            if "YouTube".lower() in (platform).lower() and youtube_check == True:
                # If the platform name has youtube then just count one and skip all other names of youtube so we can unite all youtube data in one
                filtered_condition = eachCountryDf['platform'].str.contains("YouTube", case=False)
                # Filter condition if the platform value has youtube then get the data
                new_df = eachCountryDf[filtered_condition]
                # new dataframe has all the records of the youtube 
                eachPlatformDf = pd.DataFrame(columns=list_of_columns_for_platform)
                # creating a new dataframe and passing list of columns for platform 
                for col in list_of_columns_for_platform:
                    # going to each column and storing it to new dataframe and the new dataframe has only specific columns and not country language and platform records
                    eachPlatformDf[col] = new_df[col]
  
                list_of_dataframes_eachcountry_by_platform.append({"country":countryName ,"platform":"YouTube","data": eachPlatformDf })
                youtube_check = False  
            elif "YouTube".lower() in (platform).lower() and youtube_check == False:
                continue
            else:
                filtered_condition = eachCountryDf['platform'].str.contains(platform, case=False)
                new_df = eachCountryDf[filtered_condition]
                
                

                # print(new_df.columns)
                eachPlatformDf = pd.DataFrame(columns=list_of_columns_for_platform)
                for col in list_of_columns_for_platform:
                    # print(col)
                    eachPlatformDf[col] = new_df[col]

                eachPlatformDf      # This is the dataframe for each platform 


               

                list_of_dataframes_eachcountry_by_platform.append({"country":countryName ,"platform":platform,"data": eachPlatformDf})
                
        # For platform only -------------------------------------------------------
        
        # For assets only 
        youtube_check = True

        for platform in list_of_platforms:
            # getting each platform 
            if "YouTube".lower() in (platform).lower() and youtube_check == True:
                # If the platform name has youtube then just count one and skip all other names of youtube so we can unite all youtube data in one
                filtered_condition = eachCountryDf['platform'].str.contains("YouTube", case=False)
                # Filter condition if the platform value has youtube then get the data
                new_df_language = eachCountryDf[filtered_condition]
                # List of assets of that particular dataframe platform wise 
                list_of_languages = new_df_language['language'].unique()

                for prod in list_of_languages:
                    filtered_condition_languages = new_df_language['language'].str.contains(prod, case=False)
                    new_df_prod = new_df_language[filtered_condition_languages]
                    
                    # =================
                    new_df_prod # Language wise dataframe 
                    # =============
                    
                    # print(countryName ," : ",platform," : ",prod," : ", new_df_prod['product'].unique())
                    list_of_products_language_wise = new_df_prod['product'].unique() 



                    for prod_language_wise in list_of_products_language_wise:
                        filtered_condition_language_product = new_df_prod['product'].str.contains(prod_language_wise, case=False)
                        new_df_prod_language = new_df_prod[filtered_condition_language_product]

                        new_df_prod_language # Language wise dataframe 
                    # New dataframe for getting the data only for that particular language 




                        eachPlatformlanguageProductDf = pd.DataFrame(columns=list_of_columns_for_platform)
                        for col in list_of_columns_for_platform:
                            eachPlatformlanguageProductDf[col] = new_df_prod_language[col]
                    
                        list_of_dataframes_eachcountry_by_platform_by_language_by_product.append({"country":countryName ,"platform":"YouTube","language": prod, "product":prod_language_wise,"data": eachPlatformlanguageProductDf })
                
                youtube_check = False  

                    # --------------------------------------------------------------------
             
            elif "YouTube".lower() in (platform).lower() and youtube_check == False:
                pass
            else:
                filtered_condition = eachCountryDf['platform'].str.contains(platform, case=False)
                new_df_language = eachCountryDf[filtered_condition]
                # List of assets of that particular dataframe platform wise 
                list_of_languages = new_df_language['language'].unique()

                for prod in list_of_languages:
                    filtered_condition_languages = new_df_language['language'].str.contains(prod, case=False)
                    new_df_prod = new_df_language[filtered_condition_languages]
                    
                    # =================
                    new_df_prod # Language wise dataframe 
                    # =============
                    
                    # print(countryName ," : ",platform," : ",prod," : ", new_df_prod['product'].unique())
                    list_of_products_language_wise = new_df_prod['product'].unique() 



                    for prod_language_wise in list_of_products_language_wise:
                        filtered_condition_language_product = new_df_prod['product'].str.contains(prod_language_wise, case=False)
                        
                        new_df_prod_language = new_df_prod[filtered_condition_language_product]

                        new_df_prod_language # Language wise dataframe 
                    # New dataframe for getting the data only for that particular language 




                        eachPlatformlanguageProductDf = pd.DataFrame(columns=list_of_columns_for_platform)
                        for col in list_of_columns_for_platform:
                            eachPlatformlanguageProductDf[col] = new_df_prod_language[col]
                        # print({"country":countryName ,"platform":platform,"language": prod, "product":prod_language_wise,"data": eachPlatformlanguageProductDf})
                        list_of_dataframes_eachcountry_by_platform_by_language_by_product.append({"country":countryName ,"platform":platform,"language": prod, "product":prod_language_wise,"data": eachPlatformlanguageProductDf })

    # print("Total Dataframes:",len(list_of_dataframes_eachcountry_by_platform_by_language))
        # -----------------------------------------------------------------------
    
    # Then in each market (Country), separate them using platform wise
    
    # print(list_of_dataframes_eachcountry_by_platform)
    
    # Loop through list of countries 
    
    for entry in list_of_dataframes_eachcountry_by_platform:
        country = entry['country']
        platform = entry['platform']
        data_frame = entry['data']
        

        # print(platform)


        # Create or open the Excel file for the country
        file_name = f'{country}_data.xlsx'
        start_date=pd.to_datetime(start_date)


        end_date=pd.to_datetime(end_date)
   
        
       
        temp_start_date = start_date
        temp_end_date = end_date
        
        data_values = []
        while temp_start_date <= temp_end_date:
            # Your logic ....
            # Add row with date 
            
            new_data_temp_start_date = [temp_start_date.strftime('%Y-%m-%d')]
            # Created new array with end date which will be incremented one by one  
            for column in list_of_columns_for_platform:
                # Going to each column and if the column is not date then add 0 to the array
                if column != 'date':
                    new_data_temp_start_date.append(0)
      
            
            data_values.append(new_data_temp_start_date)
  
            temp_start_date = temp_start_date + timedelta(days=1)
        
        newDF= pd.DataFrame(data_values, columns=list_of_columns_for_platform)
        data_frame = pd.concat([data_frame,newDF ], ignore_index=True)

        data_frame.sort_values(by='date', inplace=True)
        data_frame = data_frame.groupby(["date"]).sum()
        
        


        # Add extra columns here and show the data
        list_of_formula_columns = ['Budget Spent', 'CTR','CPC','CPV'] 
        
        i=1
        for _, row in data_frame.iterrows():
            row['CTR'] = f'=(D{i}/C{i})'

        # ----------------------------------------
        df_weekly = CalculateWeekly(data_frame,list_of_columns_for_platform)


        # printing_daily_total = data_frame.copy()  # For printing daily + weekly records

        # print(['date'] + data_frame.columns.append())
        columns_ = data_frame.columns.tolist()

        # Add a new column name to the beginning of the list
        new_column_name = 'date'
        columns_.insert(0, new_column_name)
        # print(columns_)
        print_daily_total_df = pd.DataFrame(columns=columns_)
        # for _, row in df_weekly.iterrows():
        #     values_list = row.values.tolist()
            
        #     print(values_list)
            # This is row of week


        records_added = 0

        # Define the number of records after which to insert a values_list
        insert_every = 7

        weeklyTable = []
        # Iterate through rows and create DataFrames for each values_list and separator row
        for _, row in df_weekly.iterrows():
            eachWeek = row.values.tolist() # This is week 1 
            weeklyTable.append([f"{int(_)+1} Weekly Total"]+eachWeek)

        records_added = 0
        weekly_ind=0
        for _, row in data_frame.iterrows():
            eachEntry = row.values.tolist() # This is dailyEntry
            print_daily_total_df.loc[len(print_daily_total_df)] = [_]+eachEntry
            records_added +=1
            if records_added % insert_every == 0:
                print_daily_total_df.loc[len(print_daily_total_df)] = weeklyTable[weekly_ind] 
                weekly_ind+=1
               
           


        # print(printing_daily_total)
        summary_total = df_weekly.sum().to_frame().T
        
        # print(summary_total)
        excel_file_path = f'{country}.xlsx'
        if os.path.exists(excel_file_path):
            # If the file exists, load it and add a new sheet with the DataFrame data
            # print("Check if it is running")
            wb = load_workbook(excel_file_path)
            

        else:
            wb = Workbook()
      
        ws = wb.create_sheet(platform)
        wb.active = ws
        wb,fb_start_row,start_row = media_calculation(wb,ws,1,fb_ig_obj,print_daily_total_df,df_weekly,summary_total)
        
        wb.save(excel_file_path)
        
    # {
    #     country: UAE,
    #     data: {"facebook": 13,"youtube": 13}
    # }
    countryListWithAssets = []
    for countryName in list_of_countries:
        temp = {}
        youtube_ = False
        temp["country"]= countryName
        for plat in list_of_platforms:
            if youtube_ == False and "YouTube" in plat:
                temp["YouTube"] = 13
                youtube_ = True
            elif youtube_ == True:
                pass
            else:        
                temp[plat] = 13
        countryListWithAssets.append(temp)
    
    # print(len(list_of_dataframes_eachcountry_by_platform_by_language))
    # Assets tables 
    # print(len(list_of_dataframes_eachcountry_by_platform_by_language))
    for entry in list_of_dataframes_eachcountry_by_platform_by_language_by_product:
        country = entry['country']
        platform = entry['platform']
        language = entry['language']
        product = entry['product']
        data_frame_language = entry['data']
        
        # print(f"{country}:",language)
        # print(platform)


        # Create or open the Excel file for the country
        file_name = f'{country}_data.xlsx'
        # start_date=pd.to_datetime(start_date)


        # end_date=pd.to_datetime(end_date)
   
        
       
        temp_start_date = start_date
        temp_end_date = end_date
        
        data_values = []
        while temp_start_date <= temp_end_date:
            # Your logic ....
            # Add row with date 
            
            new_data_temp_start_date = [temp_start_date.strftime('%Y-%m-%d')]
            # Created new array with end date which will be incremented one by one  
            for column in list_of_columns_for_platform:
                # Going to each column and if the column is not date then add 0 to the array
                if column != 'date':
                    new_data_temp_start_date.append(0)
      
            
            data_values.append(new_data_temp_start_date)
  
            temp_start_date = temp_start_date + timedelta(days=1)
        
        newDF= pd.DataFrame(data_values, columns=list_of_columns_for_platform)
        data_frame_language = pd.concat([data_frame_language,newDF ], ignore_index=True)

        data_frame_language.sort_values(by='date', inplace=True)
        data_frame_language = data_frame_language.groupby(["date"]).sum()
        
        # print(data_frame_language)


        # Add extra columns here and show the data
        list_of_formula_columns = ['Budget Spent', 'CTR','CPC','CPV'] 
        
        i=1
        for _, row in data_frame_language.iterrows():
            row['CTR'] = f'=(D{i}/C{i})'

        # ----------------------------------------
        df_weekly_assets = CalculateWeekly(data_frame_language,list_of_columns_for_platform)


        # printing_daily_total = data_frame_language.copy()  # For printing daily + weekly records

        # print(['date'] + data_frame_language.columns.append())
        columns_ = data_frame_language.columns.tolist()

        # Add a new column name to the beginning of the list
        new_column_name = 'date'
        columns_.insert(0, new_column_name)
        # print(columns_)
        print_daily_total_df_assets = pd.DataFrame(columns=columns_)
      


        records_added = 0

        # Define the number of records after which to insert a values_list
        insert_every = 7

        weeklyTable = []
        # Iterate through rows and create DataFrames for each values_list and separator row
        for _, row in df_weekly_assets.iterrows():
            eachWeek = row.values.tolist() # This is week 1 
            weeklyTable.append([f"{int(_)+1} Weekly Total"]+eachWeek)

        records_added = 0
        weekly_ind=0
        for _, row in data_frame_language.iterrows():
            eachEntry = row.values.tolist() # This is dailyEntry
            print_daily_total_df_assets.loc[len(print_daily_total_df_assets)] = [_]+eachEntry
            records_added +=1
            if records_added % insert_every == 0:
                print_daily_total_df_assets.loc[len(print_daily_total_df_assets)] = weeklyTable[weekly_ind] 
                weekly_ind+=1
               
           


        # print(print_daily_total_df_assets)
        summary_total_assets = df_weekly_assets.sum().to_frame().T
        
        # print(summary_total_assets)
        excel_file_path = f'{country}.xlsx'
        # if os.path.exists(excel_file_path):
            # If the file exists, load it and add a new sheet with the DataFrame data
            # print("Check if it is running")
        wb = load_workbook(excel_file_path)
        ws = wb[platform]

        # else:
        #     wb = Workbook()
        #     ws = wb.create_sheet(platform)
        wb.active = ws
        # print(countryListWithAssets)
        for list_of_coun in countryListWithAssets:
            if country == list_of_coun["country"]:
                new_column = list_of_coun[platform]
                break

        # print(f"{country}: {platform} : {new_column}")
        wb,new_record_column = media_calculation_assets(wb,ws,1,fb_ig_obj,print_daily_total_df_assets,df_weekly_assets,summary_total_assets,start_row,new_column,product,language)
        for list_of_coun in countryListWithAssets:
            if country == list_of_coun["country"]:
                list_of_coun[platform] = new_record_column
                break
        # print(countryListWithAssets)
        # list_of_platforms_for_assets[platform] = new_column
        # print(country)
        wb.save(excel_file_path)



    
    # Loop here for the dataframe eachcountry_by_platform_by asset
    # print(list_of_dataframes_eachcountry_by_platform_by_language)            

    # ============================================================


    return jsonify({
        "message": "completed"
    }),200
        # try:    
        #     fb_asset[fb_asset_obj.date] = pd.to_datetime(fb_asset[fb_asset_obj.date])

        #     fb_asset.sort_values(by=fb_asset_obj.date, inplace=True)
        #     fb_asset[fb_asset_obj.date] = fb_asset[fb_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Assets dates issues")

        # try:    
        #     fb_size[fb_size_obj.date] = pd.to_datetime(fb_size[fb_size_obj.date])

        #     fb_size.sort_values(by=fb_size_obj.date, inplace=True)
        #     fb_size[fb_size_obj.date] = fb_size[fb_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Size dates issues")

        # try:    
        #     fb_language[fb_language_obj.date] = pd.to_datetime(fb_language[fb_language_obj.date])

        #     fb_language.sort_values(by=fb_language_obj.date, inplace=True)
        #     fb_language[fb_language_obj.date] = fb_language[fb_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Language dates issues")

        # try:    
        #     fb_country[fb_country_obj.date] = pd.to_datetime(fb_country[fb_country_obj.date])

        #     fb_country.sort_values(by=fb_country_obj.date, inplace=True)
        #     fb_country[fb_country_obj.date] = fb_country[fb_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Country dates issues")

        # try:    
        #     fb_message[fb_message_obj.date] = pd.to_datetime(fb_message[fb_message_obj.date])

        #     fb_message.sort_values(by=fb_message_obj.date, inplace=True)
        #     fb_message[fb_message_obj.date] = fb_message[fb_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Message dates issues")
       

        # try:
        #     ig[ig_obj.date] = pd.to_datetime(ig[ig_obj.date])
        #     ig.sort_values(by=ig_obj.date, inplace=True)
        #     ig[ig_obj.date] = ig[ig_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("ig Dates issues")

        # try:    
        #     ig_asset[ig_asset_obj.date] = pd.to_datetime(ig_asset[ig_asset_obj.date])

        #     ig_asset.sort_values(by=ig_asset_obj.date, inplace=True)
        #     ig_asset[ig_asset_obj.date] = ig_asset[ig_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Assets dates issues")

        # try:    
        #     ig_size[ig_size_obj.date] = pd.to_datetime(ig_size[ig_size_obj.date])

        #     ig_size.sort_values(by=ig_size_obj.date, inplace=True)
        #     ig_size[ig_size_obj.date] = ig_size[ig_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Size dates issues")

        # try:    
        #     ig_language[ig_language_obj.date] = pd.to_datetime(ig_language[ig_language_obj.date])

        #     ig_language.sort_values(by=ig_language_obj.date, inplace=True)
        #     ig_language[ig_language_obj.date] = ig_language[ig_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Language dates issues")

        # try:    
        #     ig_country[ig_country_obj.date] = pd.to_datetime(ig_country[ig_country_obj.date])

        #     ig_country.sort_values(by=ig_country_obj.date, inplace=True)
        #     ig_country[ig_country_obj.date] = ig_country[ig_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Country dates issues")

        # try:    
        #     ig_message[ig_message_obj.date] = pd.to_datetime(ig_message[ig_message_obj.date])

        #     ig_message.sort_values(by=ig_message_obj.date, inplace=True)
        #     ig_message[ig_message_obj.date] = ig_message[ig_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Message dates issues")




        # try:
        #     sc[sc_obj.date] = pd.to_datetime(sc[sc_obj.date])
        #     rawData = sc
        #     sc.sort_values(by=sc_obj.date, inplace=True)
        #     sc[sc_obj.date] = sc[sc_obj.date].dt.strftime('%Y-%m-%d')

        #     rawData[sc_obj.date] = pd.to_datetime(sc[sc_obj.date])
        # except:
        #     print("sc Dates issues")

        # try:    
        #     sc_asset[sc_asset_obj.date] = pd.to_datetime(sc_asset[sc_asset_obj.date])

        #     sc_asset.sort_values(by=sc_asset_obj.date, inplace=True)
        #     sc_asset[sc_asset_obj.date] = sc_asset[sc_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Assets dates issues")

        # try:    
        #     sc_size[sc_size_obj.date] = pd.to_datetime(sc_size[sc_size_obj.date])

        #     sc_size.sort_values(by=sc_size_obj.date, inplace=True)
        #     sc_size[sc_size_obj.date] = sc_size[sc_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Size dates issues")

        # try:    
        #     sc_language[sc_language_obj.date] = pd.to_datetime(sc_language[sc_language_obj.date])

        #     sc_language.sort_values(by=sc_language_obj.date, inplace=True)
        #     sc_language[sc_language_obj.date] = sc_language[sc_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Language dates issues")

        # try:    
        #     sc_country[sc_country_obj.date] = pd.to_datetime(sc_country[sc_country_obj.date])

        #     sc_country.sort_values(by=sc_country_obj.date, inplace=True)
        #     sc_country[sc_country_obj.date] = sc_country[sc_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Country dates issues")

        # try:    
        #     sc_message[sc_message_obj.date] = pd.to_datetime(sc_message[sc_message_obj.date])

        #     sc_message.sort_values(by=sc_message_obj.date, inplace=True)
        #     sc_message[sc_message_obj.date] = sc_message[sc_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Message dates issues")

      
        

        # # Dv 360 -------------------------------------------
        # try:
        #     dv[dv_obj.date] = pd.to_datetime(dv[dv_obj.date])
        #     dv.sort_values(by=dv_obj.date, inplace=True)
        #     dv[dv_obj.date] = dv[ig_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("dv Dates issues")

        # try:    
        #     dv_asset[dv_asset_obj.date] = pd.to_datetime(dv_asset[dv_asset_obj.date])

        #     dv_asset.sort_values(by=dv_asset_obj.date, inplace=True)
        #     dv_asset[dv_asset_obj.date] = dv_asset[dv_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Assets dates issues")

        # try:    
        #     dv_size[dv_size_obj.date] = pd.to_datetime(dv_size[dv_size_obj.date])

        #     dv_size.sort_values(by=dv_size_obj.date, inplace=True)
        #     dv_size[dv_size_obj.date] = dv_size[dv_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Size dates issues")

        # try:    
        #     dv_language[dv_language_obj.date] = pd.to_datetime(dv_language[dv_language_obj.date])

        #     dv_language.sort_values(by=dv_language_obj.date, inplace=True)
        #     dv_language[dv_language_obj.date] = dv_language[dv_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Language dates issues")

        # try:    
        #     dv_country[dv_country_obj.date] = pd.to_datetime(dv_country[dv_country_obj.date])

        #     dv_country.sort_values(by=dv_country_obj.date, inplace=True)
        #     dv_country[dv_country_obj.date] = dv_country[dv_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Country dates issues")

        # try:    
        #     dv_message[dv_message_obj.date] = pd.to_datetime(dv_message[dv_message_obj.date])

        #     dv_message.sort_values(by=dv_message_obj.date, inplace=True)
        #     dv_message[dv_message_obj.date] = dv_message[dv_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Message dates issues")

        # --------------------------------------------------------
        
        

       
        # Formatting dates and sorting the values according to dates 
        
        
        
        # Creating path for the final output file 

        
        # ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
    #     try:
    #         fb = facebook_instagram.groupby([fb_obj.date]).sum()

    #     except:
    #         print("fb sum issues")
        
    #     try: 
    #         ig = ig.groupby([ig_obj.date]).sum()
    #     except:
    #         print("ig sum issues")
        
    #     try: 
    #         sc = sc.groupby([sc_obj.date]).sum()
    #     except:
    #         print("sc sum issues")
    #     try: 
    #         dv = dv.groupby([dv_obj.date]).sum()
    #     except:
    #         print("dv sum issues")
        

    #     # ----------------------------------------------
    #     try:
    #         fb_weekly = CalculateWeekly(fb,fb_obj)
    #     except:
    #         print("fb_weekly issue")
    #     try:
    #         ig_weekly = CalculateWeekly(ig,ig_obj)
    #     except:
    #         print("ig_weekly issue")
    #     try:
    #         sc_weekly = CalculateWeekly(sc,sc_obj)
    #     except:
    #         print("sc_weekly issue")
    #     try:
    #         dv_weekly = CalculateWeekly(dv,dv_obj)
    #     except:
    #         print("dv_weekly issue")
        
       
    #     # =================================================
    #     try:
    #         fb_total = fb_weekly.sum()
    #     except:
    #         print("fb_total issue")
    #     try:
    #         ig_total = ig_weekly.sum()
    #     except:
    #         print("ig_total issue")
    #     try:
    #         sc_total = sc_weekly.sum()
    #     except:
    #         print("sc_total issue")
    #     try:
    #         dv_total = dv_weekly.sum()
    #     except:
    #         print("dv_total issue")
   
        
    #     # Summing the data according to the date so we have one data for each data
        
    #     # Until here everything is fine 
    #     # Instead of setting columns using dataframe we will set them in excel anad update the excel sheet according to row and column 
    #     wb = Workbook()
    #     wb.remove(wb.active)
    #     try:
    #         ws = wb.create_sheet('Facebook')
    #         wb.active = ws
    #         fb_start_row = media_calculation(wb,ws,1,fb_obj,fb,fb_weekly,fb_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Facebook sheet skipped----")

    #     try:
    #         ws = wb.create_sheet('Instagram')
    #         wb.active = ws
    #         ig_start_row = media_calculation(wb,ws,1,ig_obj,ig,ig_weekly,ig_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Ig sheet skipped----")

    #     try:
    #         ws = wb.create_sheet('Youtube')
    #         wb.active = ws
    #         youtube_start_row = media_calculation(wb,ws,1,sc_obj,sc,sc_weekly,sc_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Youtube sheet skipped---")
        
    #     try:
    #         ws = wb.create_sheet('GDN')
    #         wb.active = ws
    #         gdn_start_row = media_calculation(wb,ws,1,dv_obj,dv,dv_weekly,dv_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("GDN sheet skipped---")

       

    #     # Sheet is active ====
    #     daily_total = "="
    #     concat_arr = []
        
    #     try:
    #         universal_fb = SettingUniversalDataframe(fb,fb_obj)
    #         if daily_total == "=":
    #             daily_total+="'Facebook'!T2"
    #         else:
    #             daily_total+="+'Facebook'!T2"
    #         concat_arr.append(universal_fb)
    #     except:
    #         print("Universal Fb skipped----")
        
    #     try:
    #     # print(universal_fb["view"])
    #         universal_ig = SettingUniversalDataframe(ig,ig_obj)
    #         if daily_total == "=":
    #             daily_total+="'Instagram'!T2"
    #         else:
    #             daily_total+="+'Instagram'!T2"
            
    #         concat_arr.append(universal_ig)
    #     except:
    #         print("Universal ig skipped -----")

    #     try:
    #         universal_sc = SettingUniversalDataframe(sc,sc_obj)
    #         if daily_total == "=":
    #             daily_total+="'Youtube'!T2"
    #         else:
    #             daily_total+="+'Youtube'!T2"
    #         concat_arr.append(universal_sc)
    #     except:
    #         print("Universal Youtube skipped ------")
    #     # universal_tw = SettingUniversalDataframe(tw,tw_obj)
    #     # universal_td = SettingUniversalDataframe(td,td_obj)
        
    #     try:
    #         universal_dv = SettingUniversalDataframe(dv,dv_obj)
    #         if daily_total == "=":
    #             daily_total+="'GDN'!T2"
    #         else:
    #             daily_total+="+'GDN'!T2"
    #         concat_arr.append(universal_dv)
    #     except:
    #         print("Universal GDN skipped----")

       

    #     universal_dataframe = pd.concat(concat_arr).groupby(["date"]).sum()
        
        
    #     universal_obj = MatchingColumn(universal_dataframe)
        
    #     universal_dataframe_weekly = CalculateWeekly(universal_dataframe,universal_obj)
    #     universal_dataframe_total = universal_dataframe_weekly.sum()
    #     ws = wb.create_sheet('Daily')
    #     wb.active = ws
    #     media_calculation(wb,ws,1,universal_obj,universal_dataframe,universal_dataframe_weekly,universal_dataframe_total,daily=True)
    #     ws.cell(row=2,column=20,value=daily_total)
    #     wb.save('summary_testing.xlsx')

    #     # Generating summary ----------------------

    #     have_column = True
    #     ws = wb.create_sheet('Summary')
    #     wb.active = ws
    #     daily_media_index=2
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Facebook",fb_obj,fb_total,have_column=have_column,media_label="Facebook")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Facebook summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Instagram",ig_obj,ig_total,have_column=have_column,media_label="Instagram")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Instagram summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Youtube",sc_obj,sc_total,have_column=have_column,media_label="Youtube")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Snapchat summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"GDN",dv_obj,dv_total,have_column=have_column,media_label="GDN")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("GDN summary is missing")
        
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Daily",universal_obj,universal_dataframe_total,have_column=have_column,media_label="Total")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Daily summary is missing")
    #     # engine.dispose()
    #     wb.save(pickle_file_path)

    #     # Calculating Asset part



    # # --------------------------------------------------------------------
    #     try:
    #         end_col = LookupsGenerator(wb,"Facebook",fb_asset,fb_asset_obj,"asset_type",13,fb_start_row)
    #         end_col = LookupsGenerator(wb,"Facebook",fb_language,fb_language_obj,"language",end_col,fb_start_row)
    #         # end_col = LookupsGenerator(wb,"Facebook",fb_country,fb_country_obj,"country",end_col,fb_start_row)
    #         end_col = LookupsGenerator(wb,"Facebook",fb_message,fb_message_obj,"message",end_col,fb_start_row)
    #     except:
    #         print("Facebook Lookup failed")
    #     try:
    #         end_col = LookupsGenerator(wb,"Instagram",ig_asset,ig_asset_obj,"asset_type",13,ig_start_row)
    #         end_col = LookupsGenerator(wb,"Instagram",ig_language,ig_language_obj,"language",end_col,ig_start_row)
    #         # end_col = LookupsGenerator(wb,"Instagram",ig_country,ig_country_obj,"country",end_col,ig_start_row)
    #         end_col = LookupsGenerator(wb,"Instagram",ig_message,ig_message_obj,"message",end_col,ig_start_row)
    #     except:
    #         print("Ig lookup failed")
        
    #     try:
    #         end_col = LookupsGenerator(wb,"GDN",dv_asset,dv_asset_obj,"asset_type",13,gdn_start_row)
    #         end_col = LookupsGenerator(wb,"GDN",dv_language,dv_language_obj,"language",end_col,gdn_start_row)
    #         # end_col = LookupsGenerator(wb,"GDN",dv_country,dv_country_obj,"country",end_col,gdn_start_row)
    #         end_col = LookupsGenerator(wb,"GDN",dv_message,dv_message_obj,"message",end_col,gdn_start_row)
    #     except:
    #         print("GDN Lookup failed")

    #     try:
    #         end_col = LookupsGenerator(wb,"Youtube",sc_asset,sc_asset_obj,"asset_type",13,youtube_start_row)
    #         end_col = LookupsGenerator(wb,"Youtube",sc_language,sc_language_obj,"language",end_col,youtube_start_row)
    #         # end_col = LookupsGenerator(wb,"Youtube",sc_country,sc_country_obj,"country",end_col,youtube_start_row)
    #         end_col = LookupsGenerator(wb,"Youtube",sc_message,sc_message_obj,"message",end_col,youtube_start_row)
    #     except:
    #         print("Youtube Lookup failed")

    #     # ---- Assets Ends ---------------------------------

    #     # return jsonify({
    #     #         "message": "Successfully uploaded",
    #     #         "status": 200
    #     # }),200
    #     return send_file(pickle_file_path, as_attachment=True)
        # ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
        
      
        
        # Create a unique dataframe with all the column names are matching 
        # Then sum all the dataframes and generate in excel

        # Function to change all the columns of the dataframe according to the universal columns 
        
        
        # print(universal_dataframe)

    # -------------------------------------------------------------------



# # Language wise data (Mixed Images and videos, separate product for each language)
@app.route("/video_images",methods=["POST"])
def video_images():
    
    # Fetch the raw data file (having all the platforms separated in each tab) 
    try:
        raw_data = request.files["raw_data"]
        start_date = request.form['start_date']
        end_date = request.form['end_date']
    except:
        print("Raw Data skipped---")

    try:
        facebook_instagram = pd.read_excel(raw_data,sheet_name="FB-IG") 
    except:
        print("facebook skipped----")

    try:
        snapchat = pd.read_excel(raw_data,sheet_name="Snapchat") 
    except:
        print("snapchat skipped--")
    try:
        gdn = pd.read_excel(raw_data,sheet_name="GDN - Google Ads") 
    except:
        print("snapchat skipped--")
    try:
        criteo = pd.read_excel(raw_data,sheet_name="Criteo")
    except:
        print("criteo skipped--")
        # print(criteo) 
    try:
        youtube = pd.read_excel(raw_data,sheet_name="YouTube") 
    except:
        print("youtube skipped--")
    
    # Go to each tab, fetch the columns and merge them together 

    # fb_ig = set_columns(facebook_instagram)

    # Separate them market wise
    # Then in each market (Country), separate them using platform wise
    # Save the weeklydata, dailytotal and summary (for each column in a object for each platform using date)

    # Then separate platform wise data using product wise 


   


    try:
        fb_ig_obj = MatchingColumn(facebook_instagram)      # Compare each column and if the data is matching with any of the list attribute in the class it will save it there so we don't have to call facebook dataframe according to the static name 
    except:
        print("Issue with the fb_ig_obj")    #Storing column names in the class attributes so we dont memorize the names everytime. 
    try:
        snapchat_obj = MatchingColumn(snapchat)
    except:
        print("Issue with the snapchat_obj")    
    try:
        gdn_obj = MatchingColumn(gdn)
    except:
        print("Issue with the gdn_obj")    
    try:
        criteo_obj = MatchingColumn(criteo)
    except:
        print("Issue with the criteo_obj")    
    try:
        youtube_obj = MatchingColumn(youtube)
    except:
        print("Issue with the youtube_obj")
        
        # print(youtube_obj.ListOfColumns())

    
    
    

    


    
    try:
     
        facebook_instagram[fb_ig_obj.date] = pd.to_datetime(facebook_instagram[fb_ig_obj.date])
        facebook_instagram.sort_values(by=fb_ig_obj.date, inplace=True)
        # Sorting the data according to date 
        facebook_instagram[fb_ig_obj.date] = facebook_instagram[fb_ig_obj.date].dt.strftime('%Y-%m-%d')
        # Converting the data back to string to store in the dataframe
    except:
        print("Facebook date skipped--")    
    try:
        snapchat[snapchat_obj.date] = pd.to_datetime(snapchat[snapchat_obj.date])
        snapchat.sort_values(by=snapchat_obj.date, inplace=True)
        # Sorting the data according to date 
        snapchat[snapchat_obj.date] = snapchat[snapchat_obj.date].dt.strftime('%Y-%m-%d')
    except:
        print("Snapchat date skipped--") 

    try:
        gdn[gdn_obj.date] = pd.to_datetime(gdn[gdn_obj.date])
        gdn.sort_values(by=gdn_obj.date, inplace=True)
        # Sorting the data according to date 
        gdn[gdn_obj.date] = gdn[gdn_obj.date].dt.strftime('%Y-%m-%d')
    except:
        print("GDN date skipped--") 
    try:
        criteo[criteo_obj.date] = pd.to_datetime(criteo[criteo_obj.date])
        criteo.sort_values(by=criteo_obj.date, inplace=True)
        # Sorting the data according to date 
        criteo[criteo_obj.date] = criteo[criteo_obj.date].dt.strftime('%Y-%m-%d')
    except:
        print("Criteo date skipped--")
    try:
        youtube[youtube_obj.date] = pd.to_datetime(youtube[youtube_obj.date])
        youtube.sort_values(by=youtube_obj.date, inplace=True)
        # Sorting the data according to date 
        youtube[youtube_obj.date] = youtube[youtube_obj.date].dt.strftime('%Y-%m-%d')

        # print(snapchat_obj)
    except:
        print("Youtube date skipped--")
    universal_df=pd.DataFrame()
    try:
        # print(youtube)

        # Create a dataframe 
        universal_df_fb_ig = pd.DataFrame()
        list_of_columns = fb_ig_obj.ListOfColumns()
        for col in list_of_columns:
            # print(col)
            universal_df_fb_ig[col] = facebook_instagram[getattr(fb_ig_obj, col)]
        universal_df = pd.concat([universal_df,universal_df_fb_ig], ignore_index=True)
    except:
        print("Facebook universal skipped--")

            # Store those columns which are needed and save it in meaningful variable
        # print(universal_df_fb_ig['country'].unique())
    try:    
        universal_df_sc = pd.DataFrame()
        list_of_columns = snapchat_obj.ListOfColumns()
        for col in list_of_columns:
            # print(col)
            universal_df_sc[col] = snapchat[getattr(snapchat_obj, col)]
        universal_df = pd.concat([universal_df,universal_df_sc], ignore_index=True)
        
    except:
        print("Snapchat universal skipped--")
        #     Store those columns which are needed and save it in meaningful variable
        # print(universal_df_sc)
        # print(universal_df_sc['country'].unique())
    try:
        universal_df_gdn = pd.DataFrame()
        list_of_columns = gdn_obj.ListOfColumns()
        for col in list_of_columns:
            if col in ["view","percent_25","percent_50","percent_75","percent_100"]:
                universal_df_gdn[col] = 0
            else:
                universal_df_gdn[col] = gdn[getattr(gdn_obj, col)]
        universal_df = pd.concat([universal_df,universal_df_gdn], ignore_index=True)
    except:
        print("Gdn universal skipped--")

    try:
            # Store those columns which are needed and save it in meaningful variable
        # print(universal_df_gdn)
        # print(universal_df_gdn['country'].unique())
        universal_df_criteo = pd.DataFrame()
        list_of_columns = criteo_obj.ListOfColumns()
        for col in list_of_columns:
            if col in ["view","percent_25","percent_50","percent_75","percent_100"]:
                universal_df_criteo[col] = 0
            else:
                universal_df_criteo[col] = criteo[getattr(criteo_obj, col)]
        universal_df = pd.concat([universal_df,universal_df_criteo], ignore_index=True)
    except:
        print("criteo universal skipped")
            # Store those columns which are needed and save it in meaningful variable
        # print(universal_df_criteo)
        # print(universal_df_criteo['country'].unique())
    try:
        universal_df_youtube = pd.DataFrame()
        list_of_columns = youtube_obj.ListOfColumns()
        for col in list_of_columns:
            # print(getattr(youtube_obj,col))
            universal_df_youtube[col] = youtube[getattr(youtube_obj, col)]
            # Store those columns which are needed and save it in meaningful variable
        # print(universal_df_youtube)
        # Add all datafram together
        universal_df = pd.concat([universal_df,universal_df_youtube], ignore_index=True)

        # print(universal_df)
    except:
        print("Youtube universal skipped--")
        # testDf = pd.DataFrame(columns=fb_ig_obj.l)
    

    list_of_columns = fb_ig_obj.ListOfColumns()
    universal_df # Dataframe with data
    list_of_columns # List of columns for dataframe 

    # Separate them country wise ---------------------------------------------------------------------------

    # Country names 
    list_of_countries = universal_df['country'].unique()
    # print(list_of_countries)
    list_of_dataframes_by_country = []

    for country in list_of_countries:
        filtered_condition = universal_df['country'].str.contains(country, case=False)    
        list_of_dataframes_by_country.append({"country":country,"data": universal_df[filtered_condition]})
        # print(f"country: {country}")
        # print(universal_df[filtered_condition])

    # print(list_of_dataframes_by_country)
    # --------------------------------------------------------------------------------------------------------
    list_of_columns_for_platform = fb_ig_obj.ListOfColumnsForPlatform()
    # print(list_of_columns_for_platform)
    list_of_dataframes_eachcountry_by_platform = []
    list_of_dataframes_eachcountry_by_platform_by_language=[]
    list_of_dataframes_eachcountry_by_platform_by_language_by_product = []
    for country_info in list_of_dataframes_by_country:
        countryName= country_info["country"]    # Country name
        eachCountryDf = country_info["data"]       # Country dataframe

        # Create a new dataframe with only platform columns 

        
        # For platform only ------------------------------------------------------
        list_of_platforms = eachCountryDf['platform'].unique()
        
        # Get unique list of platforms 
        youtube_check = True
        # Youtube check True if the record has multiple youtube 
        for platform in list_of_platforms:
            # getting each platform 
            if "YouTube".lower() in (platform).lower() and youtube_check == True:
                # If the platform name has youtube then just count one and skip all other names of youtube so we can unite all youtube data in one
                filtered_condition = eachCountryDf['platform'].str.contains("YouTube", case=False)
                # Filter condition if the platform value has youtube then get the data
                new_df = eachCountryDf[filtered_condition]
                # new dataframe has all the records of the youtube 
                eachPlatformDf = pd.DataFrame(columns=list_of_columns_for_platform)
                # creating a new dataframe and passing list of columns for platform 
                for col in list_of_columns_for_platform:
                    # going to each column and storing it to new dataframe and the new dataframe has only specific columns and not country language and platform records
                    eachPlatformDf[col] = new_df[col]
  
                list_of_dataframes_eachcountry_by_platform.append({"country":countryName ,"platform":"YouTube","data": eachPlatformDf })
                youtube_check = False  
            elif "YouTube".lower() in (platform).lower() and youtube_check == False:
                continue
            else:
                if platform.lower() == "facebook":
                   
                    filtered_condition = eachCountryDf['platform'].str.lower() == platform.lower()
                else:
                    
                    filtered_condition = eachCountryDf['platform'].str.contains(platform, case=False)
                new_df = eachCountryDf[filtered_condition]
                
                

                # print(new_df.columns)
                eachPlatformDf = pd.DataFrame(columns=list_of_columns_for_platform)
                for col in list_of_columns_for_platform:
                    # print(col)
                    eachPlatformDf[col] = new_df[col]

                eachPlatformDf      # This is the dataframe for each platform 


               

                list_of_dataframes_eachcountry_by_platform.append({"country":countryName ,"platform":platform,"data": eachPlatformDf})
                
        # For platform only -------------------------------------------------------
        
        # For assets only 
        youtube_check = True
       
        for platform in list_of_platforms:
            # getting each platform 
            if "YouTube".lower() in (platform).lower() and youtube_check == True:
                # If the platform name has youtube then just count one and skip all other names of youtube so we can unite all youtube data in one
                filtered_condition = eachCountryDf['platform'].str.contains("YouTube", case=False)
                # Filter condition if the platform value has youtube then get the data
                new_df_assets = eachCountryDf[filtered_condition]

                list_of_assets = new_df_assets['asset_type'].unique()

                for prod_asset in list_of_assets:
                    
                    filtered_condition_assets = new_df_assets['asset_type'].str.contains(prod_asset, case=False)
                    new_df_assets_prod = new_df_assets[filtered_condition_assets]
                    
                    # List of assets of that particular dataframe platform wise 
                    list_of_languages = new_df_assets_prod['language'].unique()

                    for prod in list_of_languages:
                        filtered_condition_languages = new_df_assets_prod['language'].str.contains(prod, case=False)
                        new_df_prod = new_df_assets_prod[filtered_condition_languages]
                    
                        # =================
                        new_df_prod # Language wise dataframe 
                        # =============
                        
                        # print(countryName ," : ",platform," : ",prod," : ", new_df_prod['product'].unique())
                        list_of_products_language_wise = new_df_prod['product'].unique() 



                        for prod_language_wise in list_of_products_language_wise:
                            filtered_condition_language_product = new_df_prod['product'].str.contains(prod_language_wise, case=False)
                            new_df_prod_language = new_df_prod[filtered_condition_language_product]

                            new_df_prod_language # Language wise dataframe 
                        # New dataframe for getting the data only for that particular language 




                            eachPlatformlanguageProductDf = pd.DataFrame(columns=list_of_columns_for_platform)
                            for col in list_of_columns_for_platform:
                                eachPlatformlanguageProductDf[col] = new_df_prod_language[col]
                        
                            list_of_dataframes_eachcountry_by_platform_by_language_by_product.append({"country":countryName ,"platform":"YouTube","asset":prod_asset,"language": prod, "product":prod_language_wise,"data": eachPlatformlanguageProductDf })
                    
                    youtube_check = False  

                    # --------------------------------------------------------------------
             
            elif "YouTube".lower() in (platform).lower() and youtube_check == False:
                pass
            else:
                if platform.lower() == "facebook":
                   
                    filtered_condition = eachCountryDf['platform'].str.lower() == platform.lower()
                else:
                    
                    filtered_condition = eachCountryDf['platform'].str.contains(platform, case=False)

                # new_df_language = eachCountryDf[filtered_condition]

                new_df_assets = eachCountryDf[filtered_condition]

                list_of_assets = new_df_assets['asset_type'].unique()

                for prod_asset in list_of_assets:
                    
                    filtered_condition_assets = new_df_assets['asset_type'].str.contains(prod_asset, case=False)
                    new_df_assets_prod = new_df_assets[filtered_condition_assets]
                    
                    # List of assets of that particular dataframe platform wise 
                    list_of_languages = new_df_assets_prod['language'].unique()

                    for prod in list_of_languages:
                        filtered_condition_languages = new_df_assets_prod['language'].str.contains(prod, case=False)
                        new_df_prod = new_df_assets_prod[filtered_condition_languages]
                        
                        # =================
                        new_df_prod # Language wise dataframe 
                        # =============
                        
                        # print(countryName ," : ",platform," : ",prod," : ", new_df_prod['product'].unique())
                        list_of_products_language_wise = new_df_prod['product'].unique() 



                        for prod_language_wise in list_of_products_language_wise:
                            filtered_condition_language_product = new_df_prod['product'].str.contains(prod_language_wise, case=False)
                            # print(platform,prod,prod_language_wise,new_df_prod)
                            new_df_prod_language = new_df_prod[filtered_condition_language_product]
                            new_df_prod_language # Language wise dataframe 
                        # New dataframe for getting the data only for that particular language 




                            eachPlatformlanguageProductDf = pd.DataFrame(columns=list_of_columns_for_platform)
                            for col in list_of_columns_for_platform:
                                eachPlatformlanguageProductDf[col] = new_df_prod_language[col]
                            # print({"country":countryName ,"platform":platform,"language": prod, "product":prod_language_wise,"data": eachPlatformlanguageProductDf})
                            list_of_dataframes_eachcountry_by_platform_by_language_by_product.append({"country":countryName ,"platform":platform,"asset":prod_asset,"language": prod, "product":prod_language_wise,"data": eachPlatformlanguageProductDf })
                            # print(prod_asset)
    # print("Total Dataframes:",len(list_of_dataframes_eachcountry_by_platform_by_language))
        # -----------------------------------------------------------------------
    
    # Then in each market (Country), separate them using platform wise
    
    # print(list_of_dataframes_eachcountry_by_platform)
    
    # Loop through list of countries 
    
    for entry in list_of_dataframes_eachcountry_by_platform:
        country = entry['country']
        platform = entry['platform']
        data_frame = entry['data']
        

        # print(platform)


        # Create or open the Excel file for the country
        file_name = f'{country}_data.xlsx'
        start_date=pd.to_datetime(start_date)


        end_date=pd.to_datetime(end_date)
   
        
       
        temp_start_date = start_date
        temp_end_date = end_date
        
        data_values = []
        while temp_start_date <= temp_end_date:
            # Your logic ....
            # Add row with date 
            
            new_data_temp_start_date = [temp_start_date.strftime('%Y-%m-%d')]
            # Created new array with end date which will be incremented one by one  
            for column in list_of_columns_for_platform:
                # Going to each column and if the column is not date then add 0 to the array
                if column != 'date':
                    new_data_temp_start_date.append(0)
      
            
            data_values.append(new_data_temp_start_date)
  
            temp_start_date = temp_start_date + timedelta(days=1)
        
        newDF= pd.DataFrame(data_values, columns=list_of_columns_for_platform)
        data_frame = pd.concat([data_frame,newDF ], ignore_index=True)

        data_frame.sort_values(by='date', inplace=True)
        data_frame = data_frame.groupby(["date"]).sum()
        
        


        # Add extra columns here and show the data
        list_of_formula_columns = ['Budget Spent', 'CTR','CPC','CPV'] 
        
        i=1
        for _, row in data_frame.iterrows():
            row['CTR'] = f'=(D{i}/C{i})'

        # ----------------------------------------
        df_weekly = CalculateWeekly(data_frame,list_of_columns_for_platform)


        # printing_daily_total = data_frame.copy()  # For printing daily + weekly records

        # print(['date'] + data_frame.columns.append())
        columns_ = data_frame.columns.tolist()

        # Add a new column name to the beginning of the list
        new_column_name = 'date'
        columns_.insert(0, new_column_name)
        # print(columns_)
        print_daily_total_df = pd.DataFrame(columns=columns_)
        # for _, row in df_weekly.iterrows():
        #     values_list = row.values.tolist()
            
        #     print(values_list)
            # This is row of week


        records_added = 0

        # Define the number of records after which to insert a values_list
        insert_every = 7

        weeklyTable = []
        # Iterate through rows and create DataFrames for each values_list and separator row
        for _, row in df_weekly.iterrows():
            eachWeek = row.values.tolist() # This is week 1 
            weeklyTable.append([f"{int(_)+1} Weekly Total"]+eachWeek)

        records_added = 0
        weekly_ind=0
        for _, row in data_frame.iterrows():
            eachEntry = row.values.tolist() # This is dailyEntry
            print_daily_total_df.loc[len(print_daily_total_df)] = [_]+eachEntry
            records_added +=1
            if records_added % insert_every == 0:
                print_daily_total_df.loc[len(print_daily_total_df)] = weeklyTable[weekly_ind] 
                weekly_ind+=1
               
           


        # print(printing_daily_total)
        summary_total = df_weekly.sum().to_frame().T
        
        # print(summary_total)
        excel_file_path = f'{country}.xlsx'
        if os.path.exists(excel_file_path):
            # If the file exists, load it and add a new sheet with the DataFrame data
            # print("Check if it is running")
            wb = load_workbook(excel_file_path)
            

        else:
            wb = Workbook()
      
        ws = wb.create_sheet(platform)
        wb.active = ws
        wb,fb_start_row,start_row = media_calculation(wb,ws,1,fb_ig_obj,print_daily_total_df,df_weekly,summary_total)
        
        wb.save(excel_file_path)
        
    # {
    #     country: UAE,
    #     data: {"facebook": 13,"youtube": 13}
    # }
    countryListWithAssets = []
    for countryName in list_of_countries:
        temp = {}
        youtube_ = False
        temp["country"]= countryName
        for plat in list_of_platforms:
            if youtube_ == False and "YouTube" in plat:
                temp["YouTube"] = 13
                youtube_ = True
            elif youtube_ == True:
                pass
            else:        
                temp[plat] = 13
        countryListWithAssets.append(temp)
    
    # print(len(list_of_dataframes_eachcountry_by_platform_by_language))
    # Assets tables 
    # print(len(list_of_dataframes_eachcountry_by_platform_by_language))
    for entry in list_of_dataframes_eachcountry_by_platform_by_language_by_product:
        # print(entry)
        country = entry['country']
        platform = entry['platform']
        asset_type = entry['asset']
        language = entry['language']
        product = entry['product']
        data_frame_language = entry['data']
        
        # print(f"{country}:",language)
        # print(platform)


        # Create or open the Excel file for the country
        file_name = f'{country}_data.xlsx'
        # start_date=pd.to_datetime(start_date)


        # end_date=pd.to_datetime(end_date)
   
        
       
        temp_start_date = start_date
        temp_end_date = end_date
        
        data_values = []
        while temp_start_date <= temp_end_date:
            # Your logic ....
            # Add row with date 
            
            new_data_temp_start_date = [temp_start_date.strftime('%Y-%m-%d')]
            # Created new array with end date which will be incremented one by one  
            for column in list_of_columns_for_platform:
                # Going to each column and if the column is not date then add 0 to the array
                if column != 'date':
                    new_data_temp_start_date.append(0)
      
            
            data_values.append(new_data_temp_start_date)
  
            temp_start_date = temp_start_date + timedelta(days=1)
        
        newDF= pd.DataFrame(data_values, columns=list_of_columns_for_platform)
        data_frame_language = pd.concat([data_frame_language,newDF ], ignore_index=True)

        data_frame_language.sort_values(by='date', inplace=True)
        data_frame_language = data_frame_language.groupby(["date"]).sum()
        
        # print(data_frame_language)


        # Add extra columns here and show the data
        list_of_formula_columns = ['Budget Spent', 'CTR','CPC','CPV'] 
        
        i=1
        for _, row in data_frame_language.iterrows():
            row['CTR'] = f'=(D{i}/C{i})'

        # ----------------------------------------
        df_weekly_assets = CalculateWeekly(data_frame_language,list_of_columns_for_platform)


        # printing_daily_total = data_frame_language.copy()  # For printing daily + weekly records

        # print(['date'] + data_frame_language.columns.append())
        columns_ = data_frame_language.columns.tolist()

        # Add a new column name to the beginning of the list
        new_column_name = 'date'
        columns_.insert(0, new_column_name)
        # print(columns_)
        print_daily_total_df_assets = pd.DataFrame(columns=columns_)
      


        records_added = 0

        # Define the number of records after which to insert a values_list
        insert_every = 7

        weeklyTable = []
        # Iterate through rows and create DataFrames for each values_list and separator row
        for _, row in df_weekly_assets.iterrows():
            eachWeek = row.values.tolist() # This is week 1 
            weeklyTable.append([f"{int(_)+1} Weekly Total"]+eachWeek)

        records_added = 0
        weekly_ind=0
        for _, row in data_frame_language.iterrows():
            eachEntry = row.values.tolist() # This is dailyEntry
            print_daily_total_df_assets.loc[len(print_daily_total_df_assets)] = [_]+eachEntry
            records_added +=1
            if records_added % insert_every == 0:
                print_daily_total_df_assets.loc[len(print_daily_total_df_assets)] = weeklyTable[weekly_ind] 
                weekly_ind+=1
               
           


        # print(print_daily_total_df_assets)
        summary_total_assets = df_weekly_assets.sum().to_frame().T
        
        # print(summary_total_assets)
        excel_file_path = f'{country}.xlsx'
        # if os.path.exists(excel_file_path):
            # If the file exists, load it and add a new sheet with the DataFrame data
            # print("Check if it is running")
        wb = load_workbook(excel_file_path)
        ws = wb[platform]

        # else:
        #     wb = Workbook()
        #     ws = wb.create_sheet(platform)
        wb.active = ws
        # print(countryListWithAssets)
        for list_of_coun in countryListWithAssets:
            if country == list_of_coun["country"]:
                new_column = list_of_coun[platform]
                break

        # print(f"{country}: {platform} : {new_column}")
        wb,new_record_column = media_calculation_assets(wb,ws,1,fb_ig_obj,print_daily_total_df_assets,df_weekly_assets,summary_total_assets,start_row,new_column,product+" ("+language+") ",asset_type)
        for list_of_coun in countryListWithAssets:
            if country == list_of_coun["country"]:
                list_of_coun[platform] = new_record_column
                break
        # print(countryListWithAssets)
        # list_of_platforms_for_assets[platform] = new_column
        # print(country)
        wb.save(excel_file_path)



    
    # Loop here for the dataframe eachcountry_by_platform_by asset
    # print(list_of_dataframes_eachcountry_by_platform_by_language)            

    # ============================================================


    return jsonify({
        "message": "completed"
    }),200
        # try:    
        #     fb_asset[fb_asset_obj.date] = pd.to_datetime(fb_asset[fb_asset_obj.date])

        #     fb_asset.sort_values(by=fb_asset_obj.date, inplace=True)
        #     fb_asset[fb_asset_obj.date] = fb_asset[fb_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Assets dates issues")

        # try:    
        #     fb_size[fb_size_obj.date] = pd.to_datetime(fb_size[fb_size_obj.date])

        #     fb_size.sort_values(by=fb_size_obj.date, inplace=True)
        #     fb_size[fb_size_obj.date] = fb_size[fb_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Size dates issues")

        # try:    
        #     fb_language[fb_language_obj.date] = pd.to_datetime(fb_language[fb_language_obj.date])

        #     fb_language.sort_values(by=fb_language_obj.date, inplace=True)
        #     fb_language[fb_language_obj.date] = fb_language[fb_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Language dates issues")

        # try:    
        #     fb_country[fb_country_obj.date] = pd.to_datetime(fb_country[fb_country_obj.date])

        #     fb_country.sort_values(by=fb_country_obj.date, inplace=True)
        #     fb_country[fb_country_obj.date] = fb_country[fb_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Country dates issues")

        # try:    
        #     fb_message[fb_message_obj.date] = pd.to_datetime(fb_message[fb_message_obj.date])

        #     fb_message.sort_values(by=fb_message_obj.date, inplace=True)
        #     fb_message[fb_message_obj.date] = fb_message[fb_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Facebook Message dates issues")
       

        # try:
        #     ig[ig_obj.date] = pd.to_datetime(ig[ig_obj.date])
        #     ig.sort_values(by=ig_obj.date, inplace=True)
        #     ig[ig_obj.date] = ig[ig_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("ig Dates issues")

        # try:    
        #     ig_asset[ig_asset_obj.date] = pd.to_datetime(ig_asset[ig_asset_obj.date])

        #     ig_asset.sort_values(by=ig_asset_obj.date, inplace=True)
        #     ig_asset[ig_asset_obj.date] = ig_asset[ig_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Assets dates issues")

        # try:    
        #     ig_size[ig_size_obj.date] = pd.to_datetime(ig_size[ig_size_obj.date])

        #     ig_size.sort_values(by=ig_size_obj.date, inplace=True)
        #     ig_size[ig_size_obj.date] = ig_size[ig_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Size dates issues")

        # try:    
        #     ig_language[ig_language_obj.date] = pd.to_datetime(ig_language[ig_language_obj.date])

        #     ig_language.sort_values(by=ig_language_obj.date, inplace=True)
        #     ig_language[ig_language_obj.date] = ig_language[ig_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Language dates issues")

        # try:    
        #     ig_country[ig_country_obj.date] = pd.to_datetime(ig_country[ig_country_obj.date])

        #     ig_country.sort_values(by=ig_country_obj.date, inplace=True)
        #     ig_country[ig_country_obj.date] = ig_country[ig_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Country dates issues")

        # try:    
        #     ig_message[ig_message_obj.date] = pd.to_datetime(ig_message[ig_message_obj.date])

        #     ig_message.sort_values(by=ig_message_obj.date, inplace=True)
        #     ig_message[ig_message_obj.date] = ig_message[ig_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Instagram Message dates issues")




        # try:
        #     sc[sc_obj.date] = pd.to_datetime(sc[sc_obj.date])
        #     rawData = sc
        #     sc.sort_values(by=sc_obj.date, inplace=True)
        #     sc[sc_obj.date] = sc[sc_obj.date].dt.strftime('%Y-%m-%d')

        #     rawData[sc_obj.date] = pd.to_datetime(sc[sc_obj.date])
        # except:
        #     print("sc Dates issues")

        # try:    
        #     sc_asset[sc_asset_obj.date] = pd.to_datetime(sc_asset[sc_asset_obj.date])

        #     sc_asset.sort_values(by=sc_asset_obj.date, inplace=True)
        #     sc_asset[sc_asset_obj.date] = sc_asset[sc_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Assets dates issues")

        # try:    
        #     sc_size[sc_size_obj.date] = pd.to_datetime(sc_size[sc_size_obj.date])

        #     sc_size.sort_values(by=sc_size_obj.date, inplace=True)
        #     sc_size[sc_size_obj.date] = sc_size[sc_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Size dates issues")

        # try:    
        #     sc_language[sc_language_obj.date] = pd.to_datetime(sc_language[sc_language_obj.date])

        #     sc_language.sort_values(by=sc_language_obj.date, inplace=True)
        #     sc_language[sc_language_obj.date] = sc_language[sc_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Language dates issues")

        # try:    
        #     sc_country[sc_country_obj.date] = pd.to_datetime(sc_country[sc_country_obj.date])

        #     sc_country.sort_values(by=sc_country_obj.date, inplace=True)
        #     sc_country[sc_country_obj.date] = sc_country[sc_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Country dates issues")

        # try:    
        #     sc_message[sc_message_obj.date] = pd.to_datetime(sc_message[sc_message_obj.date])

        #     sc_message.sort_values(by=sc_message_obj.date, inplace=True)
        #     sc_message[sc_message_obj.date] = sc_message[sc_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("Snapchat Message dates issues")

      
        

        # # Dv 360 -------------------------------------------
        # try:
        #     dv[dv_obj.date] = pd.to_datetime(dv[dv_obj.date])
        #     dv.sort_values(by=dv_obj.date, inplace=True)
        #     dv[dv_obj.date] = dv[ig_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("dv Dates issues")

        # try:    
        #     dv_asset[dv_asset_obj.date] = pd.to_datetime(dv_asset[dv_asset_obj.date])

        #     dv_asset.sort_values(by=dv_asset_obj.date, inplace=True)
        #     dv_asset[dv_asset_obj.date] = dv_asset[dv_asset_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Assets dates issues")

        # try:    
        #     dv_size[dv_size_obj.date] = pd.to_datetime(dv_size[dv_size_obj.date])

        #     dv_size.sort_values(by=dv_size_obj.date, inplace=True)
        #     dv_size[dv_size_obj.date] = dv_size[dv_size_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Size dates issues")

        # try:    
        #     dv_language[dv_language_obj.date] = pd.to_datetime(dv_language[dv_language_obj.date])

        #     dv_language.sort_values(by=dv_language_obj.date, inplace=True)
        #     dv_language[dv_language_obj.date] = dv_language[dv_language_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Language dates issues")

        # try:    
        #     dv_country[dv_country_obj.date] = pd.to_datetime(dv_country[dv_country_obj.date])

        #     dv_country.sort_values(by=dv_country_obj.date, inplace=True)
        #     dv_country[dv_country_obj.date] = dv_country[dv_country_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Country dates issues")

        # try:    
        #     dv_message[dv_message_obj.date] = pd.to_datetime(dv_message[dv_message_obj.date])

        #     dv_message.sort_values(by=dv_message_obj.date, inplace=True)
        #     dv_message[dv_message_obj.date] = dv_message[dv_message_obj.date].dt.strftime('%Y-%m-%d')
        # except:
        #     print("DV360 Message dates issues")

        # --------------------------------------------------------
        
        

       
        # Formatting dates and sorting the values according to dates 
        
        
        
        # Creating path for the final output file 

        
        # ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
    #     try:
    #         fb = facebook_instagram.groupby([fb_obj.date]).sum()

    #     except:
    #         print("fb sum issues")
        
    #     try: 
    #         ig = ig.groupby([ig_obj.date]).sum()
    #     except:
    #         print("ig sum issues")
        
    #     try: 
    #         sc = sc.groupby([sc_obj.date]).sum()
    #     except:
    #         print("sc sum issues")
    #     try: 
    #         dv = dv.groupby([dv_obj.date]).sum()
    #     except:
    #         print("dv sum issues")
        

    #     # ----------------------------------------------
    #     try:
    #         fb_weekly = CalculateWeekly(fb,fb_obj)
    #     except:
    #         print("fb_weekly issue")
    #     try:
    #         ig_weekly = CalculateWeekly(ig,ig_obj)
    #     except:
    #         print("ig_weekly issue")
    #     try:
    #         sc_weekly = CalculateWeekly(sc,sc_obj)
    #     except:
    #         print("sc_weekly issue")
    #     try:
    #         dv_weekly = CalculateWeekly(dv,dv_obj)
    #     except:
    #         print("dv_weekly issue")
        
       
    #     # =================================================
    #     try:
    #         fb_total = fb_weekly.sum()
    #     except:
    #         print("fb_total issue")
    #     try:
    #         ig_total = ig_weekly.sum()
    #     except:
    #         print("ig_total issue")
    #     try:
    #         sc_total = sc_weekly.sum()
    #     except:
    #         print("sc_total issue")
    #     try:
    #         dv_total = dv_weekly.sum()
    #     except:
    #         print("dv_total issue")
   
        
    #     # Summing the data according to the date so we have one data for each data
        
    #     # Until here everything is fine 
    #     # Instead of setting columns using dataframe we will set them in excel anad update the excel sheet according to row and column 
    #     wb = Workbook()
    #     wb.remove(wb.active)
    #     try:
    #         ws = wb.create_sheet('Facebook')
    #         wb.active = ws
    #         fb_start_row = media_calculation(wb,ws,1,fb_obj,fb,fb_weekly,fb_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Facebook sheet skipped----")

    #     try:
    #         ws = wb.create_sheet('Instagram')
    #         wb.active = ws
    #         ig_start_row = media_calculation(wb,ws,1,ig_obj,ig,ig_weekly,ig_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Ig sheet skipped----")

    #     try:
    #         ws = wb.create_sheet('Youtube')
    #         wb.active = ws
    #         youtube_start_row = media_calculation(wb,ws,1,sc_obj,sc,sc_weekly,sc_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("Youtube sheet skipped---")
        
    #     try:
    #         ws = wb.create_sheet('GDN')
    #         wb.active = ws
    #         gdn_start_row = media_calculation(wb,ws,1,dv_obj,dv,dv_weekly,dv_total)
    #     except:
    #         wb.remove(wb.active)
    #         print("GDN sheet skipped---")

       

    #     # Sheet is active ====
    #     daily_total = "="
    #     concat_arr = []
        
    #     try:
    #         universal_fb = SettingUniversalDataframe(fb,fb_obj)
    #         if daily_total == "=":
    #             daily_total+="'Facebook'!T2"
    #         else:
    #             daily_total+="+'Facebook'!T2"
    #         concat_arr.append(universal_fb)
    #     except:
    #         print("Universal Fb skipped----")
        
    #     try:
    #     # print(universal_fb["view"])
    #         universal_ig = SettingUniversalDataframe(ig,ig_obj)
    #         if daily_total == "=":
    #             daily_total+="'Instagram'!T2"
    #         else:
    #             daily_total+="+'Instagram'!T2"
            
    #         concat_arr.append(universal_ig)
    #     except:
    #         print("Universal ig skipped -----")

    #     try:
    #         universal_sc = SettingUniversalDataframe(sc,sc_obj)
    #         if daily_total == "=":
    #             daily_total+="'Youtube'!T2"
    #         else:
    #             daily_total+="+'Youtube'!T2"
    #         concat_arr.append(universal_sc)
    #     except:
    #         print("Universal Youtube skipped ------")
    #     # universal_tw = SettingUniversalDataframe(tw,tw_obj)
    #     # universal_td = SettingUniversalDataframe(td,td_obj)
        
    #     try:
    #         universal_dv = SettingUniversalDataframe(dv,dv_obj)
    #         if daily_total == "=":
    #             daily_total+="'GDN'!T2"
    #         else:
    #             daily_total+="+'GDN'!T2"
    #         concat_arr.append(universal_dv)
    #     except:
    #         print("Universal GDN skipped----")

       

    #     universal_dataframe = pd.concat(concat_arr).groupby(["date"]).sum()
        
        
    #     universal_obj = MatchingColumn(universal_dataframe)
        
    #     universal_dataframe_weekly = CalculateWeekly(universal_dataframe,universal_obj)
    #     universal_dataframe_total = universal_dataframe_weekly.sum()
    #     ws = wb.create_sheet('Daily')
    #     wb.active = ws
    #     media_calculation(wb,ws,1,universal_obj,universal_dataframe,universal_dataframe_weekly,universal_dataframe_total,daily=True)
    #     ws.cell(row=2,column=20,value=daily_total)
    #     wb.save('summary_testing.xlsx')

    #     # Generating summary ----------------------

    #     have_column = True
    #     ws = wb.create_sheet('Summary')
    #     wb.active = ws
    #     daily_media_index=2
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Facebook",fb_obj,fb_total,have_column=have_column,media_label="Facebook")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Facebook summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Instagram",ig_obj,ig_total,have_column=have_column,media_label="Instagram")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Instagram summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Youtube",sc_obj,sc_total,have_column=have_column,media_label="Youtube")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Snapchat summary is missing")
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"GDN",dv_obj,dv_total,have_column=have_column,media_label="GDN")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("GDN summary is missing")
        
    #     try:
    #         summary_calculation(wb,ws,daily_media_index,"Daily",universal_obj,universal_dataframe_total,have_column=have_column,media_label="Total")
    #         have_column = False
    #         daily_media_index+=1
    #     except:
    #         print("Daily summary is missing")
    #     # engine.dispose()
    #     wb.save(pickle_file_path)

    #     # Calculating Asset part



    # # --------------------------------------------------------------------
    #     try:
    #         end_col = LookupsGenerator(wb,"Facebook",fb_asset,fb_asset_obj,"asset_type",13,fb_start_row)
    #         end_col = LookupsGenerator(wb,"Facebook",fb_language,fb_language_obj,"language",end_col,fb_start_row)
    #         # end_col = LookupsGenerator(wb,"Facebook",fb_country,fb_country_obj,"country",end_col,fb_start_row)
    #         end_col = LookupsGenerator(wb,"Facebook",fb_message,fb_message_obj,"message",end_col,fb_start_row)
    #     except:
    #         print("Facebook Lookup failed")
    #     try:
    #         end_col = LookupsGenerator(wb,"Instagram",ig_asset,ig_asset_obj,"asset_type",13,ig_start_row)
    #         end_col = LookupsGenerator(wb,"Instagram",ig_language,ig_language_obj,"language",end_col,ig_start_row)
    #         # end_col = LookupsGenerator(wb,"Instagram",ig_country,ig_country_obj,"country",end_col,ig_start_row)
    #         end_col = LookupsGenerator(wb,"Instagram",ig_message,ig_message_obj,"message",end_col,ig_start_row)
    #     except:
    #         print("Ig lookup failed")
        
    #     try:
    #         end_col = LookupsGenerator(wb,"GDN",dv_asset,dv_asset_obj,"asset_type",13,gdn_start_row)
    #         end_col = LookupsGenerator(wb,"GDN",dv_language,dv_language_obj,"language",end_col,gdn_start_row)
    #         # end_col = LookupsGenerator(wb,"GDN",dv_country,dv_country_obj,"country",end_col,gdn_start_row)
    #         end_col = LookupsGenerator(wb,"GDN",dv_message,dv_message_obj,"message",end_col,gdn_start_row)
    #     except:
    #         print("GDN Lookup failed")

    #     try:
    #         end_col = LookupsGenerator(wb,"Youtube",sc_asset,sc_asset_obj,"asset_type",13,youtube_start_row)
    #         end_col = LookupsGenerator(wb,"Youtube",sc_language,sc_language_obj,"language",end_col,youtube_start_row)
    #         # end_col = LookupsGenerator(wb,"Youtube",sc_country,sc_country_obj,"country",end_col,youtube_start_row)
    #         end_col = LookupsGenerator(wb,"Youtube",sc_message,sc_message_obj,"message",end_col,youtube_start_row)
    #     except:
    #         print("Youtube Lookup failed")

    #     # ---- Assets Ends ---------------------------------

    #     # return jsonify({
    #     #         "message": "Successfully uploaded",
    #     #         "status": 200
    #     # }),200
    #     return send_file(pickle_file_path, as_attachment=True)
        # ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
        
      
        
        # Create a unique dataframe with all the column names are matching 
        # Then sum all the dataframes and generate in excel

        # Function to change all the columns of the dataframe according to the universal columns 
        
        
        # print(universal_dataframe)

    # -------------------------------------------------------------------





# @app.route("/fetch_platforms",methods=["GET","POST"])
# def fetch_platforms():
#     query = '''
#         SELECT DISTINCT(platform) FROM twitter;
#     '''
#     result = engine.execute(query)
#     data = []
#     for res in result:
#         row_dict = dict(res)  # Convert the row to a dictionary
#         data.append(row_dict)

   
#     print(data)
  
#     return json.dumps({
#         "data": data,
#         "status": 200
#     })

# @app.route("/api/all",methods=["GET","POST"])
# def fetch_all():
#     columns = ["date","platform", "objective", "Campaign / Campaign name   / Insertion Order", "ad group", "Ad name /YouTube Ad","Asset type","Size","Language","Country", "Message"]
#     unique_values = {}

#     for column in columns:
#         query = f"SELECT DISTINCT `{column}` FROM facebook;"
#         result = engine.execute(query)
#         values = [{'label': res[0],'value': res[0]} for res in result]
#         unique_values[column] = values

#     return jsonify(unique_values),200



#     query = '''
#         SELECT DISTINCT(platform),Objective FROM twitter;
#     '''
#     result = engine.execute(query)
#     data = []
#     for res in result:
#         row_dict = dict(res)  # Convert the row to a dictionary
#         data.append(row_dict)

   
#     print(data)
  
#     return json.dumps({
#         "data": data,
#         "status": 200
#     })


# @app.route("/api/search",methods=["POST","GET"])
# def search():
#     data = request.get_json()
#     # Write a query and send back the data

#     startDate = data['startDate']
#     endDate = data['endDate']

    
#     startDate = pd.to_datetime(data['startDate']) if data['startDate'] != '' else 1
#     endDate = pd.to_datetime(data['endDate']) if data['endDate'] != '' else 1

#     print(startDate, " - ", endDate) 
#     platform_data =  data['search']['Platform'] 
#     objective_data = data['search']['Objective'] 
#     campaign_data =  data['search']['Campaign'] 
#     adgroup_data =   data['search']['Adgroup'] 
#     adname_data =    data['search']['Adname'] 
#     assetType_data = data['search']['AssetType'] 
#     size_data =      data['search']['Size'] 
#     language_data =  data['search']['Language'] 
#     country_data =   data['search']['Country'] 
#     message_data =   data['search']['Message'] 

#     # print(platform_data)    

#     # We have a list of platforms objects - take the values from the list - use it in the query

#     len_platforms = len(platform_data)

#     platforms = []
#     objective = []
#     adgroup = []
#     adname = []
#     assetType = []
#     campaign = []
#     language = []
#     message = []
#     country = []
#     size = [] 

#     platforms = ["`platform` = '"+ platform_data[i] + "'" for i in range(0,len(platform_data))]
#     objective = ["`objective` = '"+ objective_data[i] + "'" for i in range(0,len(objective_data))]
#     campaign = ["`Campaign / Campaign name   / Insertion Order` = '"+ campaign_data[i] + "'" for i in range(0,len(campaign_data))]
#     adgroup = ["`ad group` = '" + adgroup_data[i] + "'" for i in range(0,len(adgroup_data))]
#     adname = ["`Ad name /YouTube Ad` = '" +adname_data[i]+ "'" for i in range(0,len(adname_data))]
#     assetType = ["`Asset type` = '" +assetType_data[i]+ "'" for i in range(0,len(assetType_data))]
#     size = ["`Size` = '"+size_data[i]+ "'" for i in range(0,len(size_data))]
#     language = ["`language` = '"+ language_data[i]+ "'" for i in range(0,len(language_data))]
#     country = ["`Country` = '"+country_data[i]+ "'" for i in range(0,len(country_data))]
#     message = ["`Message` ='"+message_data[i]+ "'" for i in range(0,len(message_data))]
    

        
#     # print(message)

#     # Now we have to fetch the data based on the query by passing list of attributes inside the query 
#     Where_Platform = ["date","platform", "objective", "Campaign / Campaign name   / Insertion Order", "ad group", "Ad name /YouTube Ad","Asset type","Size","Language","Country", "Message"]
    
#     Where = "WHERE"
#     Where_Date = f"`Date` between '{startDate}' AND '{endDate}'" if (startDate != 1 and endDate != 1) else "1"
#     Where_Platform = '(' + ' OR '.join(platforms)+ ')' if platforms != [] else "1"
#     Where_Objective = '(' + ' OR '.join(objective)+ ')' if objective != [] else "1"
#     Where_campaign = '(' + ' OR '.join(campaign)+ ')' if campaign != [] else "1"
#     Where_adgroup = '(' + ' OR '.join(adgroup)+ ')' if adgroup != [] else "1"
#     Where_adname = '(' + ' OR '.join(adname)+ ')' if adname != [] else "1"
#     Where_assetType = '(' + ' OR '.join(assetType)+ ')' if assetType != [] else "1"
#     Where_size = '(' + ' OR '.join(size)+ ')' if size != [] else "1"
#     Where_language = '(' + ' OR '.join(language)+ ')' if language != [] else "1"
#     Where_country = '(' + ' OR '.join(country)+ ')' if country != [] else "1"
#     Where_message = '(' + ' OR '.join(message)+ ')' if message != [] else "1"

#     # print(Where_Date)
#     Where = Where +' ' +Where_Date + ' AND '+Where_Platform + ' AND ' + Where_Objective + ' AND ' +Where_campaign + ' AND ' +Where_adgroup + ' AND ' +Where_adname + ' AND ' +Where_assetType + ' AND ' +Where_size + ' AND ' +Where_language + ' AND ' +Where_country + ' AND ' +Where_message
#     # Think how we can make filters dynamic 
#     # SELECT `Date`,SUM(`Impression`) AS `Impression`,SUM(`Link clicks`) AS `Clicks`,`Platform` FROM `facebook` GROUP BY `Date`;


#     # Take list of platforms and create a conditional where clause
#     # 
#     # print(Where)
    
#     # select_columns = ', '.join(Columns)
#     # select_tables = ', '.join("facebook")
#     # select_where = "WHERE " + ' AND '.join(Where)
#     # print(select_columns)
    

#     # select_where = "WHERE " + ' OR '.join(Where)


# # -----------------------------------------------------------------
# # Daily Performance Table 
# # Description : Show Platform wise Daily total, Show Daily Total Table for all platforms
#     Columns = ['Date',"SUM(`Impression`) AS `Impression`", "SUM(`Link clicks`) AS `Link clicks`", "SUM(`Amount Spent`) AS `Amount Spent`", 
#                 "SUM(`Conversion`) AS `Conversion`", "SUM(`View (Video)`) AS `View (Video)`", "SUM(`25%%e Views (Video)`) AS `25%% Views (Video)`", "SUM(`50%% View (Video)`) AS `50%% View (Video)`", 
#                 "SUM(`75%% View (Video)`) AS `75%% View (Video)`", "SUM(`Complete Views (Video)`) AS `Complete Views (Video)`"]
#     select_columns = ', '.join(Columns)


#     query = f"""
#         SELECT {select_columns} 
#         FROM `facebook` 
#         {Where}
#         GROUP BY `Date`;
#     """
#     result = engine.execute(query)
#     # Fetch and print the query results

        
#     column_names = result.keys()


#     dailyPerformanceRows = []

#     # Iterate over each row in the result
#     for row in result:
        
#         row_dict = dict(zip(column_names, row))
#         dailyPerformanceRows.append((row_dict))

    
    
#     # Close the result object
#     # print(dailyPerformanceRows)
#     result.close()
# # ------------------------------------------------------------

# # -----------------------------------------------------------------
# # Total Performance Table 
# # Description : Show Platform wise Daily total, Show Daily Total Table for all platforms
#     Columns = ['Platform',"SUM(`Impression`) AS `Impression`", "SUM(`Link clicks`) AS `Link clicks`", "SUM(`Amount Spent`) AS `Amount Spent`", 
#                 "SUM(`Conversion`) AS `Conversion`", "SUM(`View (Video)`) AS `View (Video)`", "SUM(`25%%e Views (Video)`) AS `25%% Views (Video)`", "SUM(`50%% View (Video)`) AS `50%% View (Video)`", 
#                 "SUM(`75%% View (Video)`) AS `75%% View (Video)`", "SUM(`Complete Views (Video)`) AS `Complete Views (Video)`"]
#     select_columns = ', '.join(Columns)


#     query = f"""
#         SELECT {select_columns} 
#         FROM `facebook` 
#         {Where}
#         GROUP BY `Platform`;
#     """
#     result = engine.execute(query)
#     # Fetch and print the query results

        
#     column_names = result.keys()


#     totalPerformanceRows = []

#     # Iterate over each row in the result
#     for row in result:
        
#         row_dict = dict(zip(column_names, row))
#         totalPerformanceRows.append((row_dict))

#     query = f"""
#         SELECT {select_columns} 
#         FROM `facebook` 
#         {Where};
#     """
#     result = engine.execute(query)
#     for row in result:
        
#         row_dict = dict(zip(column_names, row))
#         row_dict['Platform'] = "Total"
#         totalPerformanceRows.append((row_dict))
#     # Close the result object
#     # print(totalPerformanceRows)
#     result.close()
# # ------------------------------------------------------------
#     print(totalPerformanceRows)

    
#     # select_columns
#     # # Find sizes 
#     # query = f"""
#     #     SELECT 'Size' 
#     #     FROM `facebook` 
#     #     {Where};
#     # """





#     # Return the query result as a JSON response
#     return jsonify({"dailyPerformance":dailyPerformanceRows, "totalPerformance": totalPerformanceRows})
#     # --------------------------
    
#     # data = request.json
#     # print (data) 
#     return jsonify({
#         "data": rows,
#         "status": 200
#     }),200
#     # return jsonify(data),200

# @app.route("/",methods=["GET"])
# def index():
#     return render_template("index.html")


@app.route("/global",methods=["GET"])
def globa():
    return render_template("global.html")

def CalculateEachMedia(df,df_obj,budget):
    summary = df.copy()
    summary.drop(summary.index, inplace=True)
    # print(gdn_df_daily)
    for i in range(0, 1):
            row = df.iloc[i:len(df)].select_dtypes(include=['int64','double']).sum()
            row["CTR"] = row[df_obj.clicks] / row[df_obj.impressions]
            row[df_obj.budget] = round(row[df_obj.budget],2)
            # row['SPENT_BUDGET'] = round((row[df_obj.budget]/budget)*100)
            # type(row[df_obj.budget])
            row['SPENT_BUDGET'] = f'=ROUND({row[df_obj.budget]}/V11,2)'
            row['CTR'] = round(row['CTR']*100,2)
        
            summary = summary.append(row, ignore_index=True)
    summary = summary.assign(CPC = round((summary[df_obj.budget] / summary[df_obj.clicks]),2))
    return summary
def SettingUniversalDataframe(df,df_obj):
    # universal_df = pd.DataFrame()
    # print(df)
    universal_df = df.copy()
    
    universal_df.index.name = "date"
    universal_df = universal_df.rename(columns = {df_obj.ad_name:"ad_name"})
    universal_df = universal_df.rename(columns = {df_obj.campaign_name:"campaign_name"})
    universal_df = universal_df.rename(columns = {df_obj.budget:"budget"})
    universal_df = universal_df.rename(columns = {df_obj.impressions:"impressions"})
    universal_df = universal_df.rename(columns = {df_obj.clicks:"clicks"})
    universal_df = universal_df.rename(columns = {df_obj.view:"view"})
    # universal_df = universal_df.rename(columns = {df_obj.reach:"reach"})
    # universal_df = universal_df.rename(columns = {df_obj.thruplay:"thruplay"})
    universal_df = universal_df.rename(columns = {df_obj.percent_25:"percent_25"})
    universal_df = universal_df.rename(columns = {df_obj.percent_50:"percent_50"})
    universal_df = universal_df.rename(columns = {df_obj.percent_75:"percent_75"})
    universal_df = universal_df.rename(columns = {df_obj.percent_100:"percent_100"})
    

    # print(universal_df)
    return universal_df


def SettingOtherColumns(df,df_obj, budget):
    # print(df[df_obj.budget])
    df[df_obj.budget] = (round(df[df_obj.budget],2)).replace([np.inf, -np.inf,np.nan], 0)
    # df = df.assign(DAILY_VIEW_KPI = (round(float(monthly_view_kpi) / 30)))
    # df = df.assign(DAILY_KPI_ACHIEVEMENT = (round((df['View 100%'] / df['DAILY_VIEW_KPI'])*100)).replace([np.inf, -np.inf], 0))
    # df = df.assign(SPENT_BUDGET = (round((df[df_obj.budget].cumsum()/int(budget))*100)).replace([np.inf, -np.inf], 0))
    df = df.assign(SPENT_BUDGET = (round((df[df_obj.budget].cumsum()/int(budget))*100)).replace([np.inf, -np.inf], 0))
    # f'=ROUND({row[df_obj.budget]}/V11,2)'
    df = df.assign(CPV_COMPLETE = (round(df[df_obj.budget]/df[df_obj.percent_100],4)).replace([np.inf, -np.inf], 0))
    # print(df)
    # print(df[df_obj.impressions])
    df = df.assign(CTR = (round(((df[df_obj.clicks]/df[df_obj.impressions])*100),2).replace([np.inf, -np.inf], 0)))
    print(df_obj.clicks)
    print(df_obj.impressions)
    # print(df_obj.thruplay)
    try:
        df = df.assign(CPV_TRUEVIEW = (round(df[df_obj.budget]/df[df_obj.thruplay],4)).replace([np.inf, -np.inf], 0))
    except:
        df = df.assign(CPV_TRUEVIEW = 0)
    
    # print(df["SPENT_BUDGET"])
    return df

def CalculateWeekly(df_daily,list_of_columns):
    # print(df_daily)
    # date_rng = pd.date_range(start=df_daily.index.min(), end=df_daily.index.max(), freq='D')
    # print("Fine")
    # Fetching the start and end date so we can loop through it for weeks 
    df_weekly = df_daily.copy()
    # Making a copy of daily table so we can use the same columns for weekly as well
    # print("Still fine")
    df_weekly.drop(df_weekly.index, inplace=True)
    # We are dropping all the data so we can add new data inside weekly 
    # print("still still fine")
    
    # print(gdn_df_daily)
    # Starting the loop for weeks 
    for i in range(0, len(df_daily), 7):
        
        try:
            row = df_daily.iloc[i:i+7].select_dtypes(include=['int64','double']).sum()
            # print(row[df_daily_obj.impressions])
        except:
            row = df_daily.iloc[i:len(df_daily)].select_dtypes(include=['int64','double']).sum()
        # Doing sum for the row from 1 - 7 days and storing in row 
        try:
            row['budget'] = round(row['budget'],2)
        # Changing the Format of the budget 
        except:
            print("not fine 2")
        # try:
        #     row["CTR"]=round(((row[df_daily_obj.clicks]/row[df_daily_obj.impressions])*100),2)
        # except:
        #     print("not fine 3")
     

        # try:
        #     row["CPV_TRUEVIEW"] = round(row["CPV_TRUEVIEW"],4)
        # except:
        #     print("not fine 4")
        
        # try:
        #     try:
        #         if row[df_daily_obj.thruplay] == False:
        #             row["CPV_TRUEVIEW"] = 0
        #         else:
        #             row["CPV_TRUEVIEW"] = round(row[df_daily_obj.budget]/row[df_daily_obj.thruplay],4)

        #     except:
        #         if row[df_daily_obj.thruplay] == False:
        #             row["CPV_TRUEVIEW"] = 0
        #         else:
        #             row["CPV_TRUEVIEW"] = round(row[df_daily_obj.budget]/row[df_daily_obj.thruplay],4)
        # except:
        #     print("not fine 5")
        # # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
        # if row[df_daily_obj.percent_100] == 0:
        #     row["CPV_COMPLETE"] = 0
        # else:
        #     row["CPV_COMPLETE"] = round(row[df_daily_obj.budget]/row[df_daily_obj.percent_100],4)

        # # print("Position of index:",i+6)
        # # print(len(date_rng))
        # if i+6 >= len(date_rng):
        #     # print(row["SPENT_BUDGET"])
        #     row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][len(date_rng)-1]
        # else : 
        #     # row['Budget'] = round(row['Budget'],2)
        #     row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][i+6]
        #     # row['CTR'] = round(row['CTR'],2)
        #     # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
        #     # row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
        #     # row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
        #     # if type == "fb_video" or type == "fb_image" or type=="ig_image" or type=="ig_video":
        #     # else:
        #     #     row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
        # row["CPC"] = round((row[df_daily_obj.budget] / row[df_daily_obj.clicks]),2)
        # # df_weekly = df_weekly.assign(CPC = round((df_weekly[fb_obj.budget] / fb_summary[fb_obj.clicks]),2))
        df_weekly.loc[len(df_weekly)] = row
    # print("absolutely fine")
    # del df_weekly['type']
    return df_weekly 

def CheckObjectGetValue(media_asset_obj,media_asset_name):
    if hasattr(media_asset_obj, media_asset_name):
            # Retrieve the attribute's value
            value = getattr(media_asset_obj, media_asset_name)
            return value

def LookupsGenerator(wb,sheet_name,media_asset, media_asset_obj, media_asset_name,end_col,start_row = 8):
        
        
        
        media_object_value = CheckObjectGetValue(media_asset_obj,media_asset_name)
        list_of_assets = (media_asset[media_object_value].unique())
        print(f"{media_asset_name}:",list_of_assets)
        size_no = 1
        
        start_col = end_col + 1
        for asset in list_of_assets:
            fb_size_filtered = media_asset[media_asset[media_object_value].isin([asset])]
            # print(fb_size_filtered)
            # First message (calm) --> fb_size_filtered 
            # print(filtered_df)
            try:
                if media_asset_name == "message": 
                    # loop to images in single message
                    # use date sum and show on excel 
                    
                    list_of_image_sizes = (fb_size_filtered[media_asset_obj.size].unique())
                    for image_size in list_of_image_sizes:
                        fb_image_filtered = fb_size_filtered[media_asset[media_asset_obj.size].isin([image_size])]
                        fb_image_filtered = fb_image_filtered.groupby([media_asset_obj.date]).sum()
                        # print(fb_image_filtered)
                        try:
                            fb_image_filtered_weekly = CalculateWeekly(fb_image_filtered,media_asset_obj)
                        except:
                            print("fb size image weekly issues")

                        # Images data is not coming why ?  - Fixed
                        # When you are sending data and its calculating too? did you confirm it ?
                        # Fixedd --------
                        # Show which message it is 
                        # and then do it for all medias 


                        try:
                            fb_image_size_filtered_total = fb_image_filtered_weekly.sum()
                        except:
                            print("fb size image weekly total issue")

                
                        size=image_size
                
                        ws = wb[sheet_name]
                        wb.active = ws
                        start_col = media_calculation_message(asset,size,wb,ws,start_row,start_col,media_asset_obj,fb_image_filtered,fb_image_filtered_weekly,fb_image_size_filtered_total)
                        # except:
                        #     wb.remove(wb.active)
                        #     print(f"{asset} sheet skipped----")
                        end_col = start_col
                        size_no+=1
                    end_col = end_col - 1
                    





                else: 
                    fb_size_filtered = fb_size_filtered.groupby([media_asset_obj.date]).sum()
                    try:
                        fb_size_filtered_weekly = CalculateWeekly(fb_size_filtered,media_asset_obj)
                    except:
                        print("fb size weekly issues")


                    try:
                        fb_size_filtered_total = fb_size_filtered_weekly.sum()
                    except:
                        print("fb size weekly total issue")

            
                    size=asset
            
                    ws = wb[sheet_name]
                    wb.active = ws
                    start_col = media_calculation_asset(size,wb,ws,start_row,start_col,media_asset_obj,fb_size_filtered,fb_size_filtered_weekly,fb_size_filtered_total)
                    # except:
                    #     wb.remove(wb.active)
                    #     print(f"{asset} sheet skipped----")
                    end_col = start_col
                    size_no+=1
            except:
                print("fb size sum issues")
        
        return end_col - 1
        
            
      

            

if __name__ == "__main__":
	app.run(debug = True)




